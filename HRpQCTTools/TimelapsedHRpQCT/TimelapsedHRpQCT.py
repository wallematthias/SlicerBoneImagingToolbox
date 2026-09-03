from __future__ import annotations

import os
import re
import shutil
import signal
import subprocess
import tempfile
import json
import csv
import gc
import sys
import importlib
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import qt
import ctk
import slicer
import vtk

MODULE_VERSION = "0.2.4"
MIN_PIPELINE_VERSION = "2.0.43"
_SCENE_MASK_CHOICE_SEPARATOR = "||"


def _version_tuple(version_text):
    parts = []
    for token in re.split(r"[^0-9]+", str(version_text or "")):
        if token:
            parts.append(int(token))
    return tuple(parts or [0])


def _local_pipeline_version(repo_path):
    try:
        pyproject_text = (Path(repo_path) / "pyproject.toml").read_text(encoding="utf-8")
    except Exception:
        return ""
    match = re.search(r'(?m)^version\s*=\s*"([^"]+)"', pyproject_text)
    return match.group(1) if match else ""


def _local_pipeline_usable(repo_path, src_path):
    version = _local_pipeline_version(repo_path)
    return (
        Path(src_path).exists()
        and bool(version)
        and _version_tuple(version) >= _version_tuple(MIN_PIPELINE_VERSION)
    )


def _resolve_local_pipeline_paths(toolbox_root):
    for base in (Path(toolbox_root).parent, *Path(toolbox_root).parents):
        candidate_repo = base / "TimelapsedHRpQCT"
        candidate_src = candidate_repo / "src"
        if _local_pipeline_usable(candidate_repo, candidate_src):
            return candidate_repo, candidate_src
    return Path(toolbox_root).parent / "TimelapsedHRpQCT", Path(toolbox_root).parent / "TimelapsedHRpQCT" / "src"


def _encode_scene_mask_choice(node_id, segment_id=""):
    node_id = str(node_id or "")
    segment_id = str(segment_id or "")
    return f"{node_id}{_SCENE_MASK_CHOICE_SEPARATOR}{segment_id}" if segment_id else node_id


def _decode_scene_mask_choice(value):
    value = str(value or "")
    if _SCENE_MASK_CHOICE_SEPARATOR not in value:
        return value, ""
    node_id, segment_id = value.split(_SCENE_MASK_CHOICE_SEPARATOR, 1)
    return node_id, segment_id


# Prevent Slicer-specific ITK ImageIO plugin autoloading in this process.
# This avoids repeated MRMLIDImageIO factory noise from SimpleITK calls.
for _itk_env_key in ("ITK_AUTOLOAD_PATH", "SITK_AUTOLOAD_PATH"):
    try:
        os.environ.pop(_itk_env_key, None)
        os.environ[_itk_env_key] = ""
    except Exception:
        pass

import SimpleITK as sitk

_TOOLBOX_ROOT = Path(__file__).resolve().parents[2]
if str(_TOOLBOX_ROOT) not in sys.path:
    sys.path.insert(0, str(_TOOLBOX_ROOT))
_PIPELINE_LOCAL_REPO, _PIPELINE_LOCAL_SRC = _resolve_local_pipeline_paths(_TOOLBOX_ROOT)
if _local_pipeline_usable(_PIPELINE_LOCAL_REPO, _PIPELINE_LOCAL_SRC) and str(_PIPELINE_LOCAL_SRC) not in sys.path:
    sys.path.insert(0, str(_PIPELINE_LOCAL_SRC))
    _existing_pythonpath = os.environ.get("PYTHONPATH", "")
    os.environ["PYTHONPATH"] = (
        str(_PIPELINE_LOCAL_SRC)
        if not _existing_pythonpath
        else str(_PIPELINE_LOCAL_SRC) + os.pathsep + _existing_pythonpath
    )

from bone_imaging_derivatives import discover_manifests  # noqa: E402
from bone_imaging_derivatives import resolve_workflow_plan  # noqa: E402
_timelapsed_scene = importlib.import_module("SlicerBoneImagingToolboxLib.timelapsed_scene")
_timelapsed_scene = importlib.reload(_timelapsed_scene)
TimelapsedSceneNodeCandidate = _timelapsed_scene.TimelapsedSceneNodeCandidate
TimelapsedSceneRoiSelection = _timelapsed_scene.TimelapsedSceneRoiSelection
TimelapsedSceneTimepoint = _timelapsed_scene.TimelapsedSceneTimepoint
build_timelapsed_scene_plan = _timelapsed_scene.build_timelapsed_scene_plan
discover_timelapsed_scene_timepoints = _timelapsed_scene.discover_timelapsed_scene_timepoints
scene_segment_matches_role = _timelapsed_scene.scene_segment_matches_role
timelapsed_scene_run_args = _timelapsed_scene.timelapsed_scene_run_args

_reporting = importlib.import_module("TimelapsedHRpQCTLib.Reporting")
if not hasattr(_reporting, "COHORT_DEFAULT_EXPORT_FIELDS"):
    _reporting = importlib.reload(_reporting)
COHORT_DEFAULT_EXPORT_FIELDS = _reporting.COHORT_DEFAULT_EXPORT_FIELDS
COHORT_EXTRA_EXPORT_FIELD_SPECS = _reporting.COHORT_EXTRA_EXPORT_FIELD_SPECS
PROFILE_DISPLAY_ORDER = _reporting.PROFILE_DISPLAY_ORDER
default_export_filename = _reporting.default_export_filename
enrich_cohort_export_row = _reporting.enrich_cohort_export_row
project_rows_to_fields = _reporting.project_rows_to_fields
from slicer.ScriptedLoadableModule import (
    ScriptedLoadableModule,
    ScriptedLoadableModuleWidget,
    ScriptedLoadableModuleLogic,
    ScriptedLoadableModuleTest,
)

def _suppress_simpleitk_warnings():
    """Reduce known harmless ITK/SimpleITK warning noise in Slicer logs."""
    try:
        if hasattr(sitk, "ProcessObject_SetGlobalWarningDisplay"):
            sitk.ProcessObject_SetGlobalWarningDisplay(False)
        elif hasattr(sitk, "ProcessObject") and hasattr(sitk.ProcessObject, "SetGlobalWarningDisplay"):
            sitk.ProcessObject.SetGlobalWarningDisplay(False)
    except Exception:
        pass


class TimelapsedHRpQCT(ScriptedLoadableModule):
    def __init__(self, parent):
        super().__init__(parent)
        parent.title = "Timelapsed Remodelling"
        parent.categories = ["Bone Imaging.Microstructural Analysis"]
        parent.index = 40
        parent.dependencies = []
        parent.contributors = ["Matthias Walle"]
        parent.helpText = (
            "GUI wrapper for timelapsed-hrpqct pipeline.\n"
            f"Module version: {MODULE_VERSION}"
        )
        parent.acknowledgementText = """Author: Matthias Walle. Built for streamlined longitudinal HR-pQCT workflows.

If you use the main timelapsed HR-pQCT workflow, please cite:
Walle M, Whittier DE, Schenk D, Atkins PR, Blauth M, Zysset P, Lippuner K, Müller R, Collins CJ. Precision of bone mechanoregulation assessment in humans using longitudinal high-resolution peripheral quantitative computed tomography in vivo. Bone. 2023;172:116780. doi: 10.1016/j.bone.2023.116780.

If you use multistack registration, please cite:
Whittier DE, Walle M, Schenk D, Atkins PR, Collins CJ, Zysset P, Lippuner K, Müller R. A multi-stack registration technique to improve measurement accuracy and precision across longitudinal HR-pQCT scans. Bone. 2023;176:116893. doi: 10.1016/j.bone.2023.116893."""


class TimelapsedHRpQCTLogic(ScriptedLoadableModuleLogic):
    def __init__(self):
        super().__init__()
        _suppress_simpleitk_warnings()
        self._proc = None
        self._temp_config_path = None
        self._fallback_default_config_path = None

    def __del__(self):
        try:
            self.cleanup_temp_files(remove_fallback=True)
        except Exception:
            pass

    def is_pipeline_available(self):
        ok, _message = self.pipeline_status()
        return ok

    def pipeline_status(self):
        importlib.invalidate_caches()
        try:
            import timelapsedhrpqct
        except Exception as exc:
            return False, f"Not installed ({exc})"

        version = str(getattr(timelapsedhrpqct, "__version__", "0"))
        package_path = str(Path(getattr(timelapsedhrpqct, "__file__", "")).resolve())
        if _version_tuple(version) < _version_tuple(MIN_PIPELINE_VERSION):
            return (
                False,
                f"Out of date ({version}); install/update to timelapsed-hrpqct >= {MIN_PIPELINE_VERSION}. "
                f"Imported from {package_path}",
            )
        return True, f"Installed ({version}) from {package_path}"

    def run_cli_supports_option(self, option):
        try:
            import inspect
            from timelapsedhrpqct import cli

            return str(option) in inspect.getsource(cli._build_parser)
        except Exception:
            return False

    def install_or_update_pipeline(self):
        if _local_pipeline_usable(_PIPELINE_LOCAL_REPO, _PIPELINE_LOCAL_SRC):
            # Local development: import from the sibling checkout directly, and install the
            # published contour dependency without letting pip replace the local source tree.
            slicer.util.pip_install("hrpqct-geodesic-contour>=0.1.1")
        else:
            # Force-refresh from PyPI so package management pulls the latest release.
            slicer.util.pip_install(
                f"--upgrade --force-reinstall --no-cache-dir timelapsed-hrpqct>={MIN_PIPELINE_VERSION}"
            )
        for name in list(sys.modules):
            if name == "timelapsedhrpqct" or name.startswith("timelapsedhrpqct."):
                sys.modules.pop(name, None)

    def default_config_path(self):
        import timelapsedhrpqct
        import yaml
        from dataclasses import asdict

        package_default = Path(timelapsedhrpqct.__file__).resolve().parent / "configs" / "defaults.yml"
        if package_default.exists():
            return package_default

        # Fallback for environments where package data files were not included.
        if self._fallback_default_config_path and Path(self._fallback_default_config_path).exists():
            return Path(self._fallback_default_config_path)

        from timelapsedhrpqct.config.models import AppConfig

        default_cfg = asdict(AppConfig())
        fd, path = tempfile.mkstemp(prefix="timelapsed_default_", suffix=".yml")
        os.close(fd)
        with open(path, "w", encoding="utf-8") as f:
            yaml.safe_dump(default_cfg, f, sort_keys=False)
        self._fallback_default_config_path = path
        return Path(path)

    def discover_derivative_prerequisites(self, dataset_root):
        dataset_root = Path(dataset_root).expanduser().resolve()
        if dataset_root.name == "derivatives":
            dataset_root = dataset_root.parent
        manifests = discover_manifests(dataset_root)
        available_records = []
        for manifest in manifests:
            available_records.extend(manifest.records)
        available = {record.derivative for record in available_records}
        first = available_records[0] if available_records else None
        plan = resolve_workflow_plan(
            "Timelapse",
            manifests=manifests,
            subject_id=first.subject_id if first else "unknown",
            site=first.site if first else "unknown",
            sessions=sorted({str(record.session_id) for record in available_records if record.session_id}),
            generate_missing=True,
        )
        return {
            "registration_available": "Registration" in available,
            "common_region_available": "CommonRegion" in available,
            "planned_steps": list(plan.steps),
            "blocked": bool(plan.blocked),
            "missing_roles": [requirement.derivative for requirement in plan.missing],
        }

    def create_override_config(self, settings_dict, results_root=None):
        import yaml

        path = None
        if results_root:
            try:
                config_dir = Path(results_root) / "slicer_run_configs"
                config_dir.mkdir(parents=True, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                path = config_dir / f"timelapsed_slicer_{timestamp}.yml"
            except Exception:
                path = None

        if path is None:
            self.cleanup_temp_files(remove_fallback=False)
            fd, temp_path = tempfile.mkstemp(prefix="timelapsed_slicer_", suffix=".yml")
            os.close(fd)
            path = Path(temp_path)
            self._temp_config_path = str(path)
        else:
            self._temp_config_path = None

        with Path(path).open("w", encoding="utf-8") as f:
            yaml.safe_dump(settings_dict, f, sort_keys=False)

        return str(path)

    def cleanup_temp_files(self, remove_fallback=False):
        paths = []
        if self._temp_config_path:
            paths.append(("_temp_config_path", self._temp_config_path))
        if remove_fallback and self._fallback_default_config_path:
            paths.append(("_fallback_default_config_path", self._fallback_default_config_path))
        for attr_name, path in paths:
            try:
                p = Path(path)
                if p.exists():
                    p.unlink()
            except Exception:
                pass
            setattr(self, attr_name, None)

    def parse_input(self, root_path, parse_mode="auto"):
        try:
            from timelapsedhrpqct.config.loader import load_config
            from timelapsedhrpqct.dataset.discovery import discover_raw_sessions

            config = load_config(self.default_config_path())
            mode = str(parse_mode or "auto").strip().lower()
            if mode not in {"auto", "filename", "header"}:
                mode = "auto"

            if mode == "filename":
                sessions = discover_raw_sessions(
                    root_path,
                    config.discovery,
                    force_header_discovery=False,
                    canonicalize_sessions=False,
                )
                return sessions, None, "filename"

            if mode == "header":
                sessions = discover_raw_sessions(
                    root_path,
                    config.discovery,
                    force_header_discovery=True,
                    canonicalize_sessions=True,
                )
                return sessions, None, "header"

            # auto: prefer filename parse, then fall back to header parse.
            try:
                sessions = discover_raw_sessions(
                    root_path,
                    config.discovery,
                    force_header_discovery=False,
                    canonicalize_sessions=False,
                )
                return sessions, None, "filename"
            except Exception:
                sessions = discover_raw_sessions(
                    root_path,
                    config.discovery,
                    force_header_discovery=True,
                    canonicalize_sessions=True,
                )
                return sessions, None, "header"
        except Exception as exc:
            return [], str(exc), None

    def run_cli(self, args, on_output=None, on_finished=None):
        if self._proc is not None:
            raise RuntimeError("A pipeline process is already running")

        proc = qt.QProcess()
        proc.setProcessChannelMode(qt.QProcess.MergedChannels)
        env = qt.QProcessEnvironment.systemEnvironment()
        env.insert("PYTHONUNBUFFERED", "1")
        if os.environ.get("PYTHONPATH"):
            env.insert("PYTHONPATH", os.environ["PYTHONPATH"])
        # Enable Python faulthandler in CLI subprocesses to improve crash traces.
        if not env.contains("TIMELAPSE_FAULTHANDLER"):
            env.insert("TIMELAPSE_FAULTHANDLER", "1")
        # Avoid loading Slicer-specific ITK ImageIO plugins (e.g., MRMLIDImageIO)
        # in the pipeline subprocess, which can cause noisy factory warnings when
        # mixed with pip-installed ITK/ITKElastix components.
        if env.contains("ITK_AUTOLOAD_PATH"):
            env.remove("ITK_AUTOLOAD_PATH")
        if env.contains("SITK_AUTOLOAD_PATH"):
            env.remove("SITK_AUTOLOAD_PATH")
        # Some launcher setups may repopulate these internally; force empty.
        env.insert("ITK_AUTOLOAD_PATH", "")
        env.insert("SITK_AUTOLOAD_PATH", "")
        proc.setProcessEnvironment(env)

        def _read_output():
            raw = proc.readAll()
            # PythonQt/PySide can expose QByteArray differently across versions.
            if isinstance(raw, (bytes, bytearray)):
                data = bytes(raw)
            else:
                try:
                    data = raw.data()
                    if isinstance(data, str):
                        data = data.encode("utf-8", errors="replace")
                    else:
                        data = bytes(data)
                except Exception:
                    try:
                        data = bytes(raw)
                    except Exception:
                        data = str(raw).encode("utf-8", errors="replace")
            text = data.decode("utf-8", errors="replace")
            if on_output and text:
                # Filter recurring harmless ITK/Slicer plugin noise.
                filtered_lines = []
                for line in text.splitlines(keepends=True):
                    if "MRMLIDImageIO" in line:
                        continue
                    if "ImageIO factory did not return an ImageIOBase" in line:
                        continue
                    filtered_lines.append(line)
                filtered = "".join(filtered_lines)
                if filtered:
                    on_output(filtered)

        def _finished(*signal_args):
            self._proc = None
            # Handle different finished signal signatures across bindings.
            if len(signal_args) >= 2:
                exit_code = int(signal_args[0])
                exit_status = signal_args[1]
            elif len(signal_args) == 1:
                exit_code = int(signal_args[0])
                exit_status = 0
            else:
                try:
                    exit_code = int(proc.exitCode())
                except Exception:
                    exit_code = 0
                try:
                    exit_status = proc.exitStatus()
                except Exception:
                    exit_status = 0
            if on_finished:
                on_finished(exit_code, exit_status)

        proc.readyRead.connect(_read_output)
        proc.finished.connect(_finished)

        python_exe = shutil.which("PythonSlicer") or shutil.which("python") or shutil.which("python3")
        if python_exe is None:
            raise RuntimeError("Could not find Python executable in Slicer environment")

        if _local_pipeline_usable(_PIPELINE_LOCAL_REPO, _PIPELINE_LOCAL_SRC):
            bootstrap = (
                "import sys; "
                f"sys.path.insert(0, {str(_PIPELINE_LOCAL_SRC)!r}); "
                "from timelapsedhrpqct.cli import main; "
                "raise SystemExit(main())"
            )
            full_args = ["-c", bootstrap] + args
        else:
            full_args = ["-m", "timelapsedhrpqct.cli"] + args
        if on_output:
            on_output(f"[process] launching: {python_exe} {' '.join(full_args)}\n")
        proc.start(python_exe, full_args)

        if not proc.waitForStarted(3000):
            raise RuntimeError("Failed to start timelapsed-hrpqct process")
        if on_output:
            try:
                on_output(f"[process] started (pid={int(proc.processId())})\n")
            except Exception:
                on_output("[process] started\n")

        self._proc = proc

    def is_running(self):
        return self._proc is not None

    def list_external_run_pids(self):
        try:
            out = subprocess.check_output(
                ["pgrep", "-P", str(os.getpid()), "-f", "timelapsedhrpqct.cli run"],
                text=True,
            ).strip()
        except Exception:
            return []
        if not out:
            return []
        pids = []
        current = os.getpid()
        for line in out.splitlines():
            try:
                pid = int(line.strip())
            except Exception:
                continue
            if pid == current:
                continue
            if self._proc is not None:
                try:
                    if pid == int(self._proc.processId()):
                        continue
                except Exception:
                    pass
            pids.append(pid)
        return sorted(set(pids))

    def kill_external_runs(self):
        pids = self.list_external_run_pids()
        killed = []
        for pid in pids:
            try:
                os.kill(pid, signal.SIGTERM)
                killed.append(pid)
            except Exception:
                pass
        return killed

    def cancel_run(self):
        if self._proc is None:
            return False
        proc = self._proc
        try:
            proc.terminate()
            if not proc.waitForFinished(1500):
                proc.kill()
                proc.waitForFinished(1500)
            return True
        except Exception:
            try:
                proc.kill()
                proc.waitForFinished(1500)
                return True
            except Exception:
                return False


class TimelapsedHRpQCTWidget(ScriptedLoadableModuleWidget):
    def setup(self):
        super().setup()
        self.logic = TimelapsedHRpQCTLogic()
        self._queued_commands = []
        self._queued_stages = []
        self._active_stage = None
        self._stage_states = {}
        self._last_parsed_sessions = []
        self._parsed_baseline_rows = []
        self._patient_keys = []
        self._remodelling_comparison_items = []
        self._sh_tree_hooks_installed = False
        self._is_full_pipeline_run = False
        self._run_skips_mask_generation = False
        self._run_includes_analysis = False
        self._last_parse_mode_used = None
        self._updating_parse_table = False
        self._manual_parse_active = False
        self._temp_input_root = None
        self._mask_method_defaults = {
            "adaptive": (100.0, 300.0),
            "seg_gauss": (320.0, 450.0),
            "laplace_hamming": (320.0, 70.0),
        }
        self._seg_gauss_sigma = 0.8
        self._contour_gaussian_sigma = 1.5
        self._lh_cort_support_threshold = 450.0
        self._analysis_method = "grayscale_and_binary"
        self._analysis_erosion_voxels = 1
        self._interactive_preview_cache = {}
        self._updating_analysis_controls = False
        self._series_summary_pair_checks = {}
        self._latest_series_summary = None
        self._latest_study_summary_rows = []
        self._last_dataset_root_text = ""
        self._last_results_root_text = ""
        self._slice_scale_bars = {}
        self._suppress_interactive_preview_updates = False
        self._scene_subject_id = ""
        self._scene_site = ""
        self._last_scene_plan = None
        self._last_missing_scene_baseline_pairs = []
        self._syncing_scene_mask_policy = False

        self._build_ui()
        self._interactivePreviewTimer = qt.QTimer()
        self._interactivePreviewTimer.setSingleShot(True)
        self._interactivePreviewTimer.setInterval(350)
        self._interactivePreviewTimer.timeout.connect(self._on_apply_interactive_remodelling)
        self._load_defaults_from_pipeline_config()
        self._refresh_patient_list()
        self._refresh_processing_subjects()
        self._set_3d_background_black()
        self._ensure_slice_scale_bars()

    def _qt_object_alive(self, widget):
        if widget is None:
            return False
        try:
            widget.objectName
        except (RuntimeError, ValueError):
            return False
        except Exception:
            pass
        return True

    def _set_widget_enabled_safe(self, widget, enabled):
        if not self._qt_object_alive(widget):
            return
        try:
            widget.enabled = bool(enabled)
        except (RuntimeError, ValueError, AttributeError):
            return

    def _set_label_text_safe(self, label, text):
        if not self._qt_object_alive(label):
            return
        try:
            label.text = str(text)
        except (RuntimeError, ValueError, AttributeError):
            try:
                label.setText(str(text))
            except (RuntimeError, ValueError, AttributeError):
                return

    def _set_widget_style_safe(self, widget, style):
        if not self._qt_object_alive(widget):
            return
        try:
            widget.setStyleSheet(str(style))
        except (RuntimeError, ValueError, AttributeError):
            try:
                widget.styleSheet = str(style)
            except (RuntimeError, ValueError, AttributeError):
                return

    def _set_widget_visible_safe(self, widget, visible):
        if not self._qt_object_alive(widget):
            return
        try:
            widget.show() if visible else widget.hide()
        except (RuntimeError, ValueError, AttributeError):
            try:
                widget.visible = bool(visible)
            except (RuntimeError, ValueError, AttributeError):
                return

    def _set_widget_value_safe(self, widget, value):
        if not self._qt_object_alive(widget):
            return False
        previous = None
        try:
            previous = widget.blockSignals(True)
        except Exception:
            previous = None
        try:
            try:
                widget.setValue(value)
            except (RuntimeError, ValueError, AttributeError):
                widget.value = value
            return True
        except (RuntimeError, ValueError, AttributeError):
            return False
        finally:
            if previous is not None:
                try:
                    widget.blockSignals(previous)
                except Exception:
                    pass

    def _set_checkbox_checked_safe(self, checkbox, checked):
        if not self._qt_object_alive(checkbox):
            return False
        previous = None
        try:
            previous = checkbox.blockSignals(True)
        except Exception:
            previous = None
        try:
            try:
                checkbox.setChecked(bool(checked))
            except (RuntimeError, ValueError, AttributeError):
                checkbox.checked = bool(checked)
            return True
        except (RuntimeError, ValueError, AttributeError):
            return False
        finally:
            if previous is not None:
                try:
                    checkbox.blockSignals(previous)
                except Exception:
                    pass

    def _set_combo_current_data_safe(self, combo, data):
        if not self._qt_object_alive(combo):
            return False
        try:
            index = combo.findData(data)
            if index < 0:
                index = combo.findText(str(data))
            if index < 0:
                return False
            previous = combo.blockSignals(True)
            try:
                combo.setCurrentIndex(index)
            finally:
                combo.blockSignals(previous)
            return True
        except (RuntimeError, ValueError, AttributeError):
            return False

    def _combo_current_data_safe(self, combo):
        if not self._qt_object_alive(combo):
            return None
        try:
            current_data = getattr(combo, "currentData", None)
            if callable(current_data):
                value = current_data()
            else:
                value = current_data
            if value is not None:
                return value
        except (RuntimeError, ValueError, AttributeError):
            pass
        try:
            current_index = getattr(combo, "currentIndex", 0)
            if callable(current_index):
                current_index = current_index()
            return combo.itemData(int(current_index))
        except (RuntimeError, ValueError, AttributeError, TypeError):
            pass
        try:
            current_text = getattr(combo, "currentText", "")
            return current_text() if callable(current_text) else current_text
        except (RuntimeError, ValueError, AttributeError):
            return None

    def _build_ui(self):
        def _cap_width(widget, width=320):
            try:
                widget.setMaximumWidth(width)
            except Exception:
                pass

        def _label(text, help_text):
            label = qt.QLabel(str(text))
            label.toolTip = str(help_text)
            return label

        def _tip(widget, help_text):
            widget.toolTip = str(help_text)
            return widget

        self.timelapsedModeTabs = qt.QTabWidget()
        self.timelapsedModeTabs.setMaximumHeight(16777215)
        self.timelapsedModeTabs.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        self.timelapsedModeTabs.currentChanged.connect(self._on_timelapsed_mode_changed)
        scenePage = qt.QWidget()
        batchPage = qt.QWidget()
        self.timelapsedModeTabs.addTab(scenePage, "Scene")
        self.layout.addWidget(self.timelapsedModeTabs)
        self._build_scene_ui(scenePage)
        self.batchLayout = qt.QVBoxLayout(batchPage)
        self.batchLayout.setContentsMargins(0, 0, 0, 0)
        self.batchLayout.setSpacing(4)

        depBox = ctk.ctkCollapsibleButton()
        depBox.text = "Dependency"
        depBox.collapsed = True
        self.dependencyBox = depBox
        depForm = qt.QFormLayout(depBox)
        depForm.setLabelAlignment(qt.Qt.AlignRight | qt.Qt.AlignVCenter)
        self.pipelineStatusLabel = qt.QLabel()
        self.pipelineStatusLabel.wordWrap = False
        self.pipelineStatusLabel.setMaximumWidth(260)
        self.checkBtn = qt.QPushButton("Check")
        _tip(self.checkBtn, "Check the installed timelapsed-hrpqct package status and version.")
        self.checkBtn.clicked.connect(self._on_check_pipeline)
        rowWidget = qt.QWidget()
        row = qt.QHBoxLayout(rowWidget)
        row.setContentsMargins(0, 0, 0, 0)
        row.addWidget(self.checkBtn)
        depForm.addRow(_label("Status", "Installed timelapsed-hrpqct package status inside Slicer Python."), self.pipelineStatusLabel)
        depForm.addRow(rowWidget)
        self.batchLayout.addWidget(depBox)

        form = qt.QFormLayout()
        form.setLabelAlignment(qt.Qt.AlignRight | qt.Qt.AlignVCenter)
        form.setVerticalSpacing(4)

        self.inputPath = ctk.ctkPathLineEdit()
        self.inputPath.filters = ctk.ctkPathLineEdit.Dirs
        self.inputPath.setCurrentPath("")
        _tip(self.inputPath, "Folder containing raw AIM data or an existing TimelapsedHRpQCT results dataset.")
        _cap_width(self.inputPath, 360)
        form.addRow(_label("Dataset root", "Folder containing raw AIM data or an existing TimelapsedHRpQCT results dataset."), self.inputPath)
        self._connect_path_changed(self.inputPath, self._on_dataset_or_results_root_changed)
        self.derivativePrerequisitesLabel = qt.QLabel("Derivative prerequisites: Registration/CommonRegion derivatives will be discovered when available.")
        self.derivativePrerequisitesLabel.wordWrap = True
        self.derivativePrerequisitesLabel.toolTip = "Shows whether Registration/CommonRegion derivatives are available or need to be generated."
        form.addRow(_label("Prerequisites", "Derivative prerequisites for dependency-aware Timelapsed analysis."), self.derivativePrerequisitesLabel)

        parseBtn = qt.QPushButton("Parse input")
        parseBtn.clicked.connect(self._on_parse)
        _tip(parseBtn, "Discover AIM files, sessions, masks, and stacks before running the pipeline.")
        form.addRow(parseBtn)

        self.processingSubjectCombo = qt.QComboBox()
        self.processingSubjectCombo.addItem("All subjects")
        _cap_width(self.processingSubjectCombo, 260)
        self.processingSubjectCombo.toolTip = (
            "Choose which parsed subject to process. "
            "'All subjects' runs the full cohort."
        )
        self.processingSubjectCombo.currentIndexChanged.connect(self._refresh_processing_sites)
        self.processingSiteCombo = qt.QComboBox()
        self.processingSiteCombo.addItem("All sites")
        _cap_width(self.processingSiteCombo, 220)
        self.processingSiteCombo.toolTip = (
            "Choose which parsed site to process. "
            "'All sites' uses every parsed site for the selected subject scope."
        )

        self.parseSummaryLabel = qt.QLabel("Parse summary: not run")
        self.parseSummaryLabel.wordWrap = True
        self.parseSummaryLabel.toolTip = "Summary of subjects, sites, sessions, and stacks discovered during parsing."
        form.addRow(self.parseSummaryLabel)
        self.userMessageLabel = qt.QLabel("")
        self.userMessageLabel.wordWrap = True
        self.userMessageLabel.toolTip = "Important validation messages for the current dataset or settings."
        self.userMessageLabel.setStyleSheet(
            "QLabel { background:#fff6db; border:1px solid #f0c36d; padding:8px; border-radius:4px; }"
        )
        self.userMessageLabel.hide()
        form.addRow(self.userMessageLabel)

        self.parseTable = qt.QTableWidget()
        self.parseTable.setColumnCount(8)
        self.parseTable.setHorizontalHeaderLabels(
            ["Subject", "Site", "Session", "Stack", "Image", "Masks", "Seg", "Original Session"]
        )
        self.parseTable.horizontalHeader().setStretchLastSection(False)
        self.parseTable.horizontalHeader().setSectionResizeMode(qt.QHeaderView.ResizeToContents)
        self.parseTable.setMinimumHeight(128)
        self.parseTable.setEditTriggers(
            qt.QAbstractItemView.DoubleClicked
            | qt.QAbstractItemView.EditKeyPressed
            | qt.QAbstractItemView.SelectedClicked
        )
        self.parseTable.itemChanged.connect(self._on_parse_table_item_changed)

        parseBox = ctk.ctkCollapsibleButton()
        parseBox.text = "Parse Details"
        parseBox.collapsed = True
        self.parseBox = parseBox
        parseLayout = qt.QVBoxLayout(parseBox)
        parseLayout.setContentsMargins(6, 6, 6, 4)
        parseLayout.setSpacing(4)
        parseLayout.addWidget(self.parseTable)

        quickBox = qt.QGroupBox("Study Profile")
        quickForm = qt.QFormLayout(quickBox)
        quickForm.setContentsMargins(6, 8, 6, 6)
        quickForm.setVerticalSpacing(4)
        self.studyProfileCombo = qt.QComboBox()
        self._populate_study_profiles()
        _cap_width(self.studyProfileCombo, 220)
        self.applyProfileBtn = qt.QPushButton("Apply profile")
        _cap_width(self.applyProfileBtn, 105)
        self.applyProfileBtn.clicked.connect(self._on_apply_study_profile)
        self.studyProfileCombo.currentIndexChanged.connect(self._on_apply_study_profile)
        _tip(self.studyProfileCombo, "Select bundled study defaults for mask generation, registration, and analysis.")
        _tip(self.applyProfileBtn, "Apply the selected profile to visible settings and refresh the loaded preview when possible.")
        profileRow = qt.QWidget()
        profileLayout = qt.QHBoxLayout(profileRow)
        profileLayout.setContentsMargins(0, 0, 0, 0)
        profileLayout.setSpacing(6)
        profileLayout.addWidget(self.studyProfileCombo, 1)
        profileLayout.addWidget(self.applyProfileBtn)
        quickForm.addRow(_label("Profile", "Preset study settings for segmentation and remodelling analysis."), profileRow)

        analysisSectionBox = ctk.ctkCollapsibleButton()
        analysisSectionBox.text = "Analysis Options"
        analysisSectionBox.collapsed = True
        analysisSectionBox.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        self.analysisSectionBox = analysisSectionBox
        analysisSectionLayout = qt.QVBoxLayout(analysisSectionBox)
        analysisSectionLayout.setContentsMargins(6, 6, 6, 4)
        analysisSectionLayout.setSpacing(4)

        settingsBox = ctk.ctkCollapsibleButton()
        settingsBox.text = "Advanced Settings"
        settingsBox.collapsed = True
        settingsBox.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        self.advancedSettingsBox = settingsBox
        settingsLayout = qt.QVBoxLayout(settingsBox)
        settingsLayout.setContentsMargins(6, 6, 6, 4)
        settingsLayout.setSpacing(6)

        discoveryBox = ctk.ctkCollapsibleButton()
        discoveryBox.text = "Discovery / Import"
        discoveryBox.collapsed = True
        discoveryBox.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        discoveryLayout = qt.QFormLayout(discoveryBox)
        discoveryLayout.setLabelAlignment(qt.Qt.AlignRight | qt.Qt.AlignVCenter)

        maskBox = qt.QGroupBox("Mask generation")
        maskForm = qt.QFormLayout(maskBox)

        self.maskMethod = qt.QComboBox()
        self.maskMethod.addItems(["adaptive", "seg_gauss", "laplace_hamming"])
        self.maskMethod.currentTextChanged.connect(self._on_mask_method_changed)
        _tip(self.maskMethod, "Bone segmentation method used during automatic mask generation.")
        _cap_width(self.maskMethod, 220)
        self.maskPeriostealContour = qt.QComboBox()
        self.maskPeriostealContour.addItem("standard", "standard")
        self.maskPeriostealContour.addItem("geodesic_fracture", "geodesic_fracture")
        self.maskPeriostealContour.currentIndexChanged.connect(self._on_periosteal_contour_method_changed)
        _tip(self.maskPeriostealContour, "Full/periosteal contour method used during automatic mask generation.")
        _cap_width(self.maskPeriostealContour, 220)
        self.maskEndostealContour = qt.QComboBox()
        self.maskEndostealContour.addItem("standard", "standard")
        self.maskEndostealContour.addItem("none", "none")
        self.maskEndostealContour.currentIndexChanged.connect(self._on_periosteal_contour_method_changed)
        _tip(self.maskEndostealContour, "Endosteal/trab-cort contour method used when trabecular or cortical masks are generated.")
        _cap_width(self.maskEndostealContour, 220)
        self.maskLow = ctk.ctkDoubleSpinBox()
        self.maskLow.minimum = -10000.0
        self.maskLow.maximum = 10000.0
        self.maskLow.decimals = 1
        self.maskLow.singleStep = 5.0
        self.maskLow.value = 100.0
        self.maskHigh = ctk.ctkDoubleSpinBox()
        self.maskHigh.minimum = -10000.0
        self.maskHigh.maximum = 10000.0
        self.maskHigh.decimals = 1
        self.maskHigh.singleStep = 5.0
        self.maskHigh.value = 300.0
        _tip(self.maskLow, "Lower method-specific threshold used when generating masks.")
        _tip(self.maskHigh, "Upper method-specific threshold or smoothing parameter used when generating masks.")
        _cap_width(self.maskLow, 220)
        _cap_width(self.maskHigh, 220)
        self.maskLowLabel = _label("Mask lower threshold", "Lower method-specific threshold used when generating masks.")
        self.maskHighLabel = _label("Mask higher threshold", "Upper method-specific threshold or smoothing parameter used when generating masks.")
        self.maskContourSupportThreshold = ctk.ctkDoubleSpinBox()
        self.maskContourSupportThreshold.minimum = 0.0
        self.maskContourSupportThreshold.maximum = 5000.0
        self.maskContourSupportThreshold.decimals = 1
        self.maskContourSupportThreshold.singleStep = 5.0
        self.maskContourSupportThreshold.value = 320.0
        self.maskContourSupportThresholdLabel = _label("Contour support threshold", "Density threshold used to build Gaussian support for full/endosteal contours.")
        _tip(self.maskContourSupportThreshold, "Density threshold used to build Gaussian support for full/endosteal contours.")
        _cap_width(self.maskContourSupportThreshold, 220)
        self.maskEndostealThreshold = ctk.ctkDoubleSpinBox()
        self.maskEndostealThreshold.minimum = 0.0
        self.maskEndostealThreshold.maximum = 5000.0
        self.maskEndostealThreshold.decimals = 1
        self.maskEndostealThreshold.singleStep = 5.0
        self.maskEndostealThreshold.value = 500.0
        self.maskEndostealThresholdLabel = _label("Endosteal threshold", "Density threshold used by the standard endosteal/trab-cort contour.")
        _tip(self.maskEndostealThreshold, "Density threshold used by the standard endosteal/trab-cort contour.")
        _cap_width(self.maskEndostealThreshold, 220)
        self.maskEndostealKernel = qt.QSpinBox()
        self.maskEndostealKernel.minimum = 1
        self.maskEndostealKernel.maximum = 101
        self.maskEndostealKernel.singleStep = 2
        self.maskEndostealKernel.value = 3
        self.maskEndostealKernelLabel = _label("Endosteal kernel", "Kernel size used when building the standard endosteal/trab-cort contour.")
        _tip(self.maskEndostealKernel, "Kernel size used when building the standard endosteal/trab-cort contour.")
        _cap_width(self.maskEndostealKernel, 220)
        self.maskSigma = ctk.ctkDoubleSpinBox()
        self.maskSigma.minimum = 0.0
        self.maskSigma.maximum = 10.0
        self.maskSigma.decimals = 2
        self.maskSigma.singleStep = 0.05
        self.maskSigma.value = 0.8
        self.maskSigmaLabel = _label("Segmentation Gaussian sigma", "Smoothing sigma in voxels used by Gaussian bone segmentation.")
        _tip(self.maskSigma, "Smoothing sigma in voxels used by Gaussian bone segmentation.")
        _cap_width(self.maskSigma, 220)
        self.maskContourSigma = ctk.ctkDoubleSpinBox()
        self.maskContourSigma.minimum = 0.0
        self.maskContourSigma.maximum = 10.0
        self.maskContourSigma.decimals = 2
        self.maskContourSigma.singleStep = 0.05
        self.maskContourSigma.value = 1.5
        self.maskContourSigmaLabel = _label("Contour Gaussian sigma", "Smoothing sigma in voxels used by Gaussian contour support.")
        _tip(self.maskContourSigma, "Smoothing sigma in voxels used by Gaussian contour support.")
        _cap_width(self.maskContourSigma, 220)
        self.maskLaplaceThreshold = ctk.ctkDoubleSpinBox()
        self.maskLaplaceThreshold.minimum = 0.0
        self.maskLaplaceThreshold.maximum = 100000.0
        self.maskLaplaceThreshold.decimals = 1
        self.maskLaplaceThreshold.singleStep = 50.0
        self.maskLaplaceThreshold.value = 15564.0
        self.maskLaplaceThresholdLabel = _label("LH threshold", "Laplace-Hamming threshold used for bone segmentation.")
        _tip(self.maskLaplaceThreshold, "Laplace-Hamming threshold used for bone segmentation.")
        _cap_width(self.maskLaplaceThreshold, 220)
        self.maskLaplaceLowPass = ctk.ctkDoubleSpinBox()
        self.maskLaplaceLowPass.minimum = 0.0
        self.maskLaplaceLowPass.maximum = 1.0
        self.maskLaplaceLowPass.decimals = 3
        self.maskLaplaceLowPass.singleStep = 0.01
        self.maskLaplaceLowPass.value = 0.3
        self.maskLaplaceLowPassLabel = _label("LH low-pass cutoff", "Low-pass cutoff for Laplace-Hamming filtering.")
        _tip(self.maskLaplaceLowPass, "Low-pass cutoff for Laplace-Hamming filtering.")
        _cap_width(self.maskLaplaceLowPass, 220)
        self.maskLaplaceHighPass = ctk.ctkDoubleSpinBox()
        self.maskLaplaceHighPass.minimum = 0.0
        self.maskLaplaceHighPass.maximum = 1.0
        self.maskLaplaceHighPass.decimals = 3
        self.maskLaplaceHighPass.singleStep = 0.01
        self.maskLaplaceHighPass.value = 0.0
        self.maskLaplaceHighPassLabel = _label("LH high-pass cutoff", "High-pass cutoff for Laplace-Hamming filtering.")
        _tip(self.maskLaplaceHighPass, "High-pass cutoff for Laplace-Hamming filtering.")
        _cap_width(self.maskLaplaceHighPass, 220)
        self.maskLaplaceEpsilon = ctk.ctkDoubleSpinBox()
        self.maskLaplaceEpsilon.minimum = 0.0
        self.maskLaplaceEpsilon.maximum = 5.0
        self.maskLaplaceEpsilon.decimals = 3
        self.maskLaplaceEpsilon.singleStep = 0.01
        self.maskLaplaceEpsilon.value = 0.45
        self.maskLaplaceEpsilonLabel = _label("LH epsilon", "Epsilon used in Laplace-Hamming filtering.")
        _tip(self.maskLaplaceEpsilon, "Epsilon used in Laplace-Hamming filtering.")
        _cap_width(self.maskLaplaceEpsilon, 220)
        self.maskOuterKernel = qt.QSpinBox()
        self.maskOuterKernel.minimum = 1
        self.maskOuterKernel.maximum = 101
        self.maskOuterKernel.singleStep = 2
        self.maskOuterKernel.value = 5
        self.maskOuterKernelLabel = _label("Outer kernel", "Kernel size used when smoothing/building the full periosteal mask.")
        _tip(self.maskOuterKernel, "Kernel size used when smoothing/building the full periosteal mask.")
        _cap_width(self.maskOuterKernel, 220)
        self.maskOuterOpen = qt.QSpinBox()
        self.maskOuterOpen.minimum = 0
        self.maskOuterOpen.maximum = 50
        self.maskOuterOpen.singleStep = 1
        self.maskOuterOpen.value = 2
        self.maskOuterOpenLabel = _label("Outer opening radius", "Morphological opening radius used for the full periosteal mask.")
        _tip(self.maskOuterOpen, "Morphological opening radius used for the full periosteal mask.")
        _cap_width(self.maskOuterOpen, 220)
        self.maskGeodesicThreshold = ctk.ctkDoubleSpinBox()
        self.maskGeodesicThreshold.minimum = 0.0
        self.maskGeodesicThreshold.maximum = 5000.0
        self.maskGeodesicThreshold.decimals = 1
        self.maskGeodesicThreshold.singleStep = 5.0
        self.maskGeodesicThreshold.value = 250.0
        self.maskGeodesicThresholdLabel = _label("Geodesic bone threshold", "Bone threshold passed to the geodesic fracture periosteal contour method.")
        _tip(self.maskGeodesicThreshold, "Bone threshold passed to the geodesic fracture periosteal contour.")
        _cap_width(self.maskGeodesicThreshold, 220)
        self.maskGeodesicFillHoles = qt.QCheckBox()
        self.maskGeodesicFillHoles.checked = True
        self.maskGeodesicFillHolesLabel = _label("Fill full mask holes", "Fill internal holes in generated full/periosteal masks.")
        _tip(self.maskGeodesicFillHoles, "Fill internal holes in generated full/periosteal masks.")
        _cap_width(self.maskGeodesicFillHoles, 220)
        self.maskAlignedContourSupport = qt.QCheckBox()
        self.maskAlignedContourSupport.checked = False
        self.maskAlignedContourSupportLabel = _label(
            "Aligned contour support",
            "Let contour-support binarization follow the selected segmentation method. Leave off for more stable full/trab/cort masks.",
        )
        _tip(
            self.maskAlignedContourSupport,
            "When enabled, contour-support binarization follows the selected segmentation method. "
            "Leave this off to keep full/trab/cort masks more stable across sessions.",
        )
        _cap_width(self.maskAlignedContourSupport, 220)
        self.maskSegmentationSectionLabel = qt.QLabel("<b>Bone Segmentation</b>")
        self.maskPeriostealSectionLabel = qt.QLabel("<b>Full / Periosteal Contour</b>")
        self.maskEndostealSectionLabel = qt.QLabel("<b>Endosteal / Trab-Cort Contour</b>")
        maskForm.addRow(self.maskSegmentationSectionLabel)
        maskForm.addRow(_label("Bone segmentation method", "Method used when automatic bone segmentation is enabled."), self.maskMethod)
        maskForm.addRow(self.maskLowLabel, self.maskLow)
        maskForm.addRow(self.maskHighLabel, self.maskHigh)
        maskForm.addRow(self.maskSigmaLabel, self.maskSigma)
        maskForm.addRow(self.maskLaplaceThresholdLabel, self.maskLaplaceThreshold)
        maskForm.addRow(self.maskLaplaceLowPassLabel, self.maskLaplaceLowPass)
        maskForm.addRow(self.maskLaplaceHighPassLabel, self.maskLaplaceHighPass)
        maskForm.addRow(self.maskLaplaceEpsilonLabel, self.maskLaplaceEpsilon)
        maskForm.addRow(self.maskAlignedContourSupportLabel, self.maskAlignedContourSupport)
        maskForm.addRow(self.maskPeriostealSectionLabel)
        maskForm.addRow(
            _label("Full/periosteal contour", "Full/periosteal mask contour method used during automatic mask generation."),
            self.maskPeriostealContour,
        )
        maskForm.addRow(self.maskContourSupportThresholdLabel, self.maskContourSupportThreshold)
        maskForm.addRow(self.maskContourSigmaLabel, self.maskContourSigma)
        maskForm.addRow(self.maskOuterKernelLabel, self.maskOuterKernel)
        maskForm.addRow(self.maskOuterOpenLabel, self.maskOuterOpen)
        maskForm.addRow(
            self.maskGeodesicThresholdLabel,
            self.maskGeodesicThreshold,
        )
        maskForm.addRow(
            self.maskGeodesicFillHolesLabel,
            self.maskGeodesicFillHoles,
        )
        maskForm.addRow(self.maskEndostealSectionLabel)
        maskForm.addRow(
            _label("Endosteal/trab-cort contour", "Endosteal contour method used when trabecular or cortical masks are generated."),
            self.maskEndostealContour,
        )
        maskForm.addRow(self.maskEndostealThresholdLabel, self.maskEndostealThreshold)
        maskForm.addRow(self.maskEndostealKernelLabel, self.maskEndostealKernel)
        self.doNotGenerateMasksCheck = qt.QCheckBox()
        self.doNotGenerateMasksCheck.checked = False
        skip_masks_tip = (
            "Skip automatic mask/segmentation generation during Run pipeline. "
            "Use this only when existing generated outputs or raw provided masks/SEG files should be used as-is. "
            "If the selected analysis options require segmentations and none are available, the run will stop before launch."
        )
        self.doNotGenerateMasksCheck.toolTip = skip_masks_tip
        _cap_width(self.doNotGenerateMasksCheck, 220)
        self.doNotGenerateMasksLabel = _label("Do not generate masks", skip_masks_tip)
        maskForm.addRow(self.doNotGenerateMasksLabel, self.doNotGenerateMasksCheck)

        self.resultsRootPath = ctk.ctkPathLineEdit()
        self.resultsRootPath.filters = ctk.ctkPathLineEdit.Dirs
        self.resultsRootPath.setCurrentPath("")
        _tip(
            self.resultsRootPath,
            "Optional output/results root. Leave empty to write Timelapse outputs under dataset/derivatives.",
        )
        _cap_width(self.resultsRootPath, 360)
        discoveryLayout.addRow(
            _label(
                "Results folder (optional)",
                "Optional output/results root. Leave empty to write Timelapse outputs under dataset/derivatives.",
            ),
            self.resultsRootPath,
        )
        self._connect_path_changed(self.resultsRootPath, self._on_dataset_or_results_root_changed)

        self.copyRawInputsCheck = qt.QCheckBox()
        self.copyRawInputsCheck.checked = False
        self.copyRawInputsCheck.toolTip = (
            "Copy raw AIM files into sourcedata/hrpqct during import."
        )
        self.restructureRawCheck = qt.QCheckBox()
        self.restructureRawCheck.checked = False
        self.restructureRawCheck.toolTip = (
            "Move raw AIM files into results root sub-*/site-*/ses-* during import."
        )
        self.parseModeCombo = qt.QComboBox()
        self.parseModeCombo.addItems(["auto", "filename", "header"])
        self.parseModeCombo.setCurrentText("auto")
        self.parseModeCombo.toolTip = (
            "Input parsing mode. 'auto' tries filename parsing first, "
            "then falls back to header parsing."
        )
        self.storageModeCombo = qt.QComboBox()
        self.storageModeCombo.addItem("Minimal", "minimal")
        self.storageModeCombo.addItem("Full debug", "full")
        self.storageModeCombo.setCurrentIndex(0)
        self.storageModeCombo.toolTip = (
            "Derivative storage mode. Minimal keeps imported grayscale stack images as lazy AIM-backed views; "
            "derived analysis products remain cached. Full debug also writes split stack image files."
        )
        _cap_width(self.copyRawInputsCheck, 220)
        _cap_width(self.restructureRawCheck, 220)
        _cap_width(self.parseModeCombo, 220)
        _cap_width(self.storageModeCombo, 220)
        discoveryLayout.addRow(_label("Copy raw inputs", "Copy raw AIM files into sourcedata/hrpqct during import."), self.copyRawInputsCheck)
        discoveryLayout.addRow(_label("Restructure raw inputs", "Move raw AIM files into the results root sub-*/site-*/ses-* layout during import."), self.restructureRawCheck)
        discoveryLayout.addRow(_label("Parse mode", "Input parsing mode. Auto tries filenames first, then AIM headers."), self.parseModeCombo)
        discoveryLayout.addRow(_label("Storage mode", "Minimal avoids copied imported stack images; full debug writes them for inspection."), self.storageModeCombo)

        registrationBox = qt.QGroupBox("Registration")
        registrationForm = qt.QFormLayout(registrationBox)
        registrationForm.setVerticalSpacing(10)
        registrationForm.setHorizontalSpacing(18)

        tlHeader = qt.QLabel("Timelapse registration")
        tlHeader.setStyleSheet("font-weight: 600; color: #3f3f3f; padding-top: 2px;")
        tlHeader.toolTip = "Settings for registration across longitudinal sessions."
        registrationForm.addRow(tlHeader)

        self.regMetric = qt.QComboBox(); self.regMetric.addItems(["mattes", "correlation"])
        _tip(self.regMetric, "Similarity metric for timelapse and multistack registration.")
        _cap_width(self.regMetric, 220)
        registrationForm.addRow(_label("Registration metric", "Similarity metric for timelapse and multistack registration. Options: mattes or correlation."), self.regMetric)
        self.tlSampling = ctk.ctkDoubleSpinBox()
        self.tlSampling.minimum = 0.00001
        self.tlSampling.maximum = 1.0
        self.tlSampling.decimals = 5
        self.tlSampling.singleStep = 0.0001
        self.tlSampling.value = 0.001
        _tip(self.tlSampling, "Random voxel sampling fraction for timelapse registration. Lower is faster; higher may be more robust.")
        _cap_width(self.tlSampling, 220)
        registrationForm.addRow(_label("Timelapse sampling", "Random voxel sampling fraction for timelapse registration. Lower is faster; higher may be more robust."), self.tlSampling)

        self.tlRes = qt.QSpinBox(); self.tlRes.minimum = 1; self.tlRes.maximum = 10; self.tlRes.value = 6
        self.tlIter = qt.QSpinBox(); self.tlIter.minimum = 1; self.tlIter.maximum = 5000; self.tlIter.value = 250
        _tip(self.tlRes, "Number of image pyramid levels used for pairwise timelapse registration.")
        _tip(self.tlIter, "Maximum optimizer iterations per timelapse registration level.")
        _cap_width(self.tlRes, 220)
        _cap_width(self.tlIter, 220)
        registrationForm.addRow(_label("Timelapse resolutions", "Number of image pyramid levels used for pairwise timelapse registration."), self.tlRes)
        registrationForm.addRow(_label("Timelapse iterations", "Maximum optimizer iterations per registration level."), self.tlIter)

        self.msRes = qt.QSpinBox(); self.msRes.minimum = 1; self.msRes.maximum = 10; self.msRes.value = 4
        self.msIter = qt.QSpinBox(); self.msIter.minimum = 1; self.msIter.maximum = 5000; self.msIter.value = 250
        self.msSampling = ctk.ctkDoubleSpinBox()
        self.msSampling.minimum = 0.00001
        self.msSampling.maximum = 1.0
        self.msSampling.decimals = 5
        self.msSampling.singleStep = 0.0001
        self.msSampling.value = 0.005
        _tip(self.msSampling, "Random voxel sampling fraction for multistack correction registration.")
        _tip(self.msRes, "Number of image pyramid levels used for multistack correction.")
        _tip(self.msIter, "Maximum optimizer iterations per multistack correction level.")
        self.useMultistackCheck = qt.QCheckBox()
        self.useMultistackCheck.checked = False
        self.useMultistackCheck.toolTip = (
            "When enabled, Timelapse Pipeline runs with multistack correction."
        )
        self.msOverlapBuffer = qt.QSpinBox()
        self.msOverlapBuffer.minimum = 0
        self.msOverlapBuffer.maximum = 2000
        self.msOverlapBuffer.value = 40
        self.msOverlapBuffer.toolTip = (
            "Extra z-slices (in voxels) added on both sides of overlap crop "
            "for multistack superstack registration."
        )
        msHeader = qt.QLabel("Multistack correction registration")
        msHeader.setStyleSheet("font-weight: 600; color: #3f3f3f; padding-top: 8px;")
        msHeader.toolTip = "Settings for correcting multiple overlapping stacks within a scan."
        registrationForm.addRow(msHeader)
        self.msInitTx = ctk.ctkDoubleSpinBox()
        self.msInitTy = ctk.ctkDoubleSpinBox()
        self.msInitTz = ctk.ctkDoubleSpinBox()
        for sb in (self.msInitTx, self.msInitTy, self.msInitTz):
            sb.minimum = -5000.0
            sb.maximum = 5000.0
            sb.decimals = 3
            sb.singleStep = 1.0
            sb.value = 0.0
            sb.toolTip = "Initial multistack translation offset in voxels (X, Y, Z)."
        self.msInitTz.value = -20.0
        initTranslationRow = qt.QWidget()
        initTranslationLayout = qt.QHBoxLayout(initTranslationRow)
        initTranslationLayout.setContentsMargins(0, 0, 0, 0)
        initTranslationLayout.setSpacing(6)
        for label_text, widget in (("X", self.msInitTx), ("Y", self.msInitTy), ("Z", self.msInitTz)):
            lbl = qt.QLabel(label_text)
            lbl.setMinimumWidth(10)
            lbl.toolTip = "Initial multistack translation component in voxels."
            initTranslationLayout.addWidget(lbl)
            initTranslationLayout.addWidget(widget)
        _cap_width(self.msSampling, 220)
        _cap_width(self.msRes, 220)
        _cap_width(self.msIter, 220)
        _cap_width(self.msOverlapBuffer, 220)
        _cap_width(self.useMultistackCheck, 220)
        _cap_width(self.msInitTx, 90)
        _cap_width(self.msInitTy, 90)
        _cap_width(self.msInitTz, 90)
        registrationForm.addRow(_label("Use multistack correction", "Enable correction/registration for scans acquired as multiple overlapping stacks."), self.useMultistackCheck)
        registrationForm.addRow(_label("Multistack correction sampling", "Random voxel sampling fraction for multistack correction registration."), self.msSampling)
        registrationForm.addRow(_label("Multistack correction resolutions", "Number of image pyramid levels used for multistack correction."), self.msRes)
        registrationForm.addRow(_label("Multistack correction iterations", "Maximum optimizer iterations per multistack correction level."), self.msIter)
        registrationForm.addRow(_label("Multistack overlap crop buffer (voxels)", "Extra z-slices included around overlap regions for multistack registration."), self.msOverlapBuffer)
        registrationForm.addRow(_label("Multistack initial translation (voxels)", "Initial X/Y/Z translation offset for multistack registration."), initTranslationRow)

        advancedAnalysisBox = qt.QGroupBox("Advanced analysis")
        self.advancedAnalysisBox = advancedAnalysisBox
        advancedAnalysisForm = qt.QFormLayout(advancedAnalysisBox)

        analysisBox = qt.QGroupBox("Remodelling Analysis")
        analysisForm = qt.QFormLayout(analysisBox)

        self.analysisThresholdSlider = qt.QSlider(qt.Qt.Horizontal)
        self.analysisThresholdSlider.minimum = 0
        self.analysisThresholdSlider.maximum = 1000
        self.analysisThresholdSlider.singleStep = 5
        self.analysisThresholdSlider.pageStep = 25
        self.analysisThresholdSlider.tickInterval = 25
        self.analysisThresholdSlider.setTickPosition(qt.QSlider.TicksBelow)
        self.analysisThreshold = ctk.ctkDoubleSpinBox()
        self.analysisThreshold.minimum = 0.0
        self.analysisThreshold.maximum = 1000.0
        self.analysisThreshold.decimals = 0
        self.analysisThreshold.singleStep = 5.0
        self.analysisThreshold.value = 225.0
        _tip(self.analysisThresholdSlider, "Adjust absolute density-change threshold for formation/resorption preview.")
        _tip(self.analysisThreshold, "Absolute density-change threshold for formation/resorption detection.")
        self.analysisClusterSlider = qt.QSlider(qt.Qt.Horizontal)
        self.analysisClusterSlider.minimum = 0
        self.analysisClusterSlider.maximum = 30
        self.analysisClusterSlider.singleStep = 1
        self.analysisClusterSlider.pageStep = 5
        self.analysisClusterSlider.tickInterval = 1
        self.analysisClusterSlider.setTickPosition(qt.QSlider.TicksBelow)
        self.analysisCluster = qt.QSpinBox(); self.analysisCluster.minimum = 0; self.analysisCluster.maximum = 30; self.analysisCluster.singleStep = 1; self.analysisCluster.value = 12
        _tip(self.analysisClusterSlider, "Adjust minimum connected event size retained in remodelling maps.")
        _tip(self.analysisCluster, "Minimum connected event size retained in remodelling maps. Use 0 to disable cluster filtering.")
        self.analysisGaussianFilterCheck = qt.QCheckBox()
        self.analysisGaussianFilterCheck.checked = True
        self.analysisGaussianSigma = ctk.ctkDoubleSpinBox()
        self.analysisGaussianSigma.minimum = 0.0
        self.analysisGaussianSigma.maximum = 10.0
        self.analysisGaussianSigma.decimals = 2
        self.analysisGaussianSigma.singleStep = 0.1
        self.analysisGaussianSigma.value = 1.2
        self.analysisMethodCombo = qt.QComboBox()
        self.analysisMethodCombo.addItem("Binary + grayscale", "grayscale_and_binary")
        self.analysisMethodCombo.addItem("Grayscale only", "grayscale_delta_only")
        self.analysisMethodCombo.addItem("Marrow shell + grayscale", "grayscale_marrow_mask")
        self.analysisMethodCombo.visible = False
        self.analysisRestrictBoneSupportCheck = qt.QCheckBox()
        self.analysisRestrictBoneSupportCheck.checked = False
        restrict_tip = (
            "Limit formation/resorption candidates to baseline/follow-up bone support. "
            "Use this for marrow-mask style analysis."
        )
        self.analysisRestrictBoneSupportCheck.toolTip = restrict_tip
        self.analysisBinaryReclassificationCheck = qt.QCheckBox()
        self.analysisBinaryReclassificationCheck.checked = True
        binary_tip = (
            "Require the binary segmentation state to change in addition to the grayscale density change. "
            "Formation must become bone; resorption must stop being bone."
        )
        self.analysisBinaryReclassificationCheck.toolTip = binary_tip
        self.analysisPairModeCombo = qt.QComboBox()
        self.analysisPairModeCombo.addItem("Adjacent", "adjacent")
        self.analysisPairModeCombo.addItem("Baseline", "baseline")
        self.analysisPairModeCombo.addItem("All pairs", "all_pairs")
        self.analysisFullMaskDilation = qt.QSpinBox()
        self.analysisFullMaskDilation.minimum = 0
        self.analysisFullMaskDilation.maximum = 20
        self.analysisFullMaskDilation.value = 2
        self.analysisBoneSupportDilation = qt.QSpinBox()
        self.analysisBoneSupportDilation.minimum = 0
        self.analysisBoneSupportDilation.maximum = 20
        self.analysisBoneSupportDilation.value = 0
        self.analysisMarrowMaskDilation = self.analysisBoneSupportDilation
        self.analysisMarrowMaskErosion = qt.QSpinBox()
        self.analysisMarrowMaskErosion.minimum = 0
        self.analysisMarrowMaskErosion.maximum = 20
        self.analysisMarrowMaskErosion.value = 0
        self.analysisMarrowMaskErosion.visible = False
        _tip(self.analysisGaussianFilterCheck, "Smooth grayscale images before subtraction for remodelling-site detection.")
        _tip(self.analysisGaussianSigma, "Sigma in voxels when Gaussian remodelling-site filtering is enabled.")
        _tip(self.analysisMethodCombo, "Choose grayscale-only or grayscale-plus-binary remodelling event logic.")
        _tip(self.analysisPairModeCombo, "Choose which session comparisons to analyze: adjacent, baseline, or all pairs.")
        _tip(self.analysisFullMaskDilation, "Dilate full masks before common-region construction.")
        _tip(self.analysisBoneSupportDilation, "Dilate baseline/follow-up bone support when restricting changes to bone support.")
        _tip(self.analysisMarrowMaskErosion, "Erode marrow masks before marrow-shell analysis.")
        _cap_width(self.analysisThreshold, 220)
        _cap_width(self.analysisCluster, 220)
        _cap_width(self.analysisGaussianFilterCheck, 220)
        _cap_width(self.analysisGaussianSigma, 220)
        _cap_width(self.analysisMethodCombo, 220)
        _cap_width(self.analysisRestrictBoneSupportCheck, 220)
        _cap_width(self.analysisBinaryReclassificationCheck, 220)
        _cap_width(self.analysisPairModeCombo, 220)
        _cap_width(self.analysisFullMaskDilation, 220)
        _cap_width(self.analysisBoneSupportDilation, 220)
        _cap_width(self.analysisMarrowMaskErosion, 220)
        self.analysisStatusLabel = qt.QLabel("Ready")
        self.analysisStatusLabel.styleSheet = "color: #666666;"
        self.applyAnalysisSettingsBtn = qt.QPushButton("Apply settings")
        self.applyAnalysisSettingsBtn.toolTip = (
            "Apply the current remodelling analysis options to the loaded comparison."
        )
        self.runAnalysisBtn = qt.QPushButton("Rerun cohort analysis")
        self.runAnalysisBtn.toolTip = (
            "Rerun remodelling analysis for all processed samples using the current analysis options."
        )
        self.saveAnalysisScenarioBtn = qt.QPushButton("Save current analysis...")
        self.saveAnalysisScenarioBtn.toolTip = (
            "Save the currently loaded remodelling comparison and current analysis options as a scenario."
        )
        _cap_width(self.applyAnalysisSettingsBtn, 140)
        _cap_width(self.runAnalysisBtn, 180)
        _cap_width(self.saveAnalysisScenarioBtn, 180)
        thresholdRow = qt.QWidget()
        thresholdLayout = qt.QHBoxLayout(thresholdRow)
        thresholdLayout.setContentsMargins(0, 0, 0, 0)
        thresholdLayout.setSpacing(8)
        thresholdLayout.addWidget(self.analysisThresholdSlider, 1)
        thresholdLayout.addWidget(self.analysisThreshold)
        clusterRow = qt.QWidget()
        clusterLayout = qt.QHBoxLayout(clusterRow)
        clusterLayout.setContentsMargins(0, 0, 0, 0)
        clusterLayout.setSpacing(8)
        clusterLayout.addWidget(self.analysisClusterSlider, 1)
        clusterLayout.addWidget(self.analysisCluster)
        metricsBox = qt.QGroupBox("Current Comparison")
        metricsBox.toolTip = "Formation and resorption fractions for the currently loaded or previewed comparison."
        metricsBox.setStyleSheet(
            "QGroupBox { font-weight: 600; border: 1px solid #bdbdbd; border-radius: 4px; "
            "margin-top: 8px; padding: 8px 6px 6px 6px; } "
            "QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 3px; }"
        )
        metricsLayout = qt.QVBoxLayout(metricsBox)
        metricsLayout.setContentsMargins(8, 10, 8, 8)
        metricsLayout.setSpacing(4)
        self.currentComparisonTable = qt.QTableWidget()
        self.currentComparisonTable.setColumnCount(5)
        self.currentComparisonTable.setHorizontalHeaderLabels(["Mask", "FV/BV", "RV/BV", "AV/BV", "NV/BV"])
        self.currentComparisonTable.setEditTriggers(qt.QAbstractItemView.NoEditTriggers)
        self.currentComparisonTable.setSelectionMode(qt.QAbstractItemView.NoSelection)
        self.currentComparisonTable.verticalHeader().setVisible(False)
        self.currentComparisonTable.horizontalHeader().setStretchLastSection(True)
        self.currentComparisonTable.setAlternatingRowColors(True)
        self.currentComparisonTable.setMinimumHeight(82)
        self.currentComparisonTable.setMaximumHeight(120)
        self.currentComparisonTable.toolTip = (
            "Formation, resorption, activity, and net-change fractions for the currently loaded comparison."
        )
        metricsLayout.addWidget(self.currentComparisonTable)
        self._update_current_comparison_table([])
        analysisActionsRow = qt.QWidget()
        analysisActionsLayout = qt.QHBoxLayout(analysisActionsRow)
        analysisActionsLayout.setContentsMargins(0, 0, 0, 0)
        analysisActionsLayout.setSpacing(8)
        analysisActionsLayout.addWidget(self.applyAnalysisSettingsBtn)
        analysisActionsLayout.addWidget(self.runAnalysisBtn)
        analysisActionsLayout.addWidget(self.saveAnalysisScenarioBtn)
        analysisActionsLayout.addStretch(1)
        analysisForm.addRow(_label("Pair mode", "Which session comparisons to analyze: adjacent, baseline, or all pairs."), self.analysisPairModeCombo)
        analysisForm.addRow(_label("Threshold", "Absolute density-change threshold for formation/resorption detection."), thresholdRow)
        analysisForm.addRow(_label("Cluster size", "Minimum connected event size retained in remodelling maps. Use 0 to disable cluster filtering."), clusterRow)
        analysisForm.addRow(_label("Restrict changes to bone support", restrict_tip), self.analysisRestrictBoneSupportCheck)
        analysisForm.addRow(_label("Require binary reclassification", binary_tip), self.analysisBinaryReclassificationCheck)
        analysisForm.addRow(_label("Gaussian filter remodelling sites", "Smooth grayscale images before subtraction for remodelling-site detection."), self.analysisGaussianFilterCheck)
        analysisForm.addRow(_label("Actions", "Apply current options to the loaded comparison, rerun cohort analysis, or save the current analysis as a scenario."), analysisActionsRow)
        analysisForm.addRow(_label("Preview status", "Status of the interactive remodelling preview update."), self.analysisStatusLabel)
        advancedAnalysisForm.addRow(
            _label("Full mask dilation (vox)", "Dilation applied to full masks before common-region construction."),
            self.analysisFullMaskDilation,
        )
        advancedAnalysisForm.addRow(
            _label("Bone support dilation (vox)", "Dilation of baseline/follow-up bone support when restricting changes to bone support."),
            self.analysisBoneSupportDilation,
        )
        advancedAnalysisForm.addRow(
            _label("Gaussian sigma (vox)", "Sigma in voxels when Gaussian remodelling-site filtering is enabled."),
            self.analysisGaussianSigma,
        )
        self.analysisThresholdSlider.valueChanged.connect(
            lambda value: self._set_analysis_threshold_value(value, from_slider=True, queue_update=False)
        )
        self.analysisThresholdSlider.sliderReleased.connect(
            lambda: self._set_analysis_threshold_value(self.analysisThresholdSlider.value, from_slider=True, queue_update=True)
        )
        self.analysisThreshold.editingFinished.connect(
            lambda: self._set_analysis_threshold_value(self.analysisThreshold.value, queue_update=True)
        )
        self.analysisClusterSlider.valueChanged.connect(
            lambda value: self._set_analysis_cluster_value(value, from_slider=True, queue_update=False)
        )
        self.analysisClusterSlider.sliderReleased.connect(
            lambda: self._set_analysis_cluster_value(self.analysisClusterSlider.value, from_slider=True, queue_update=True)
        )
        self.analysisCluster.editingFinished.connect(
            lambda: self._set_analysis_cluster_value(self.analysisCluster.value, queue_update=True)
        )
        self.analysisMethodCombo.currentIndexChanged.connect(self._on_analysis_method_changed)
        self.analysisPairModeCombo.currentIndexChanged.connect(self._on_analysis_pair_mode_changed)
        self.analysisRestrictBoneSupportCheck.toggled.connect(self._on_analysis_option_changed)
        self.analysisBinaryReclassificationCheck.toggled.connect(self._on_analysis_option_changed)
        self.analysisFullMaskDilation.valueChanged.connect(self._on_interactive_preview_control_changed)
        self.analysisBoneSupportDilation.valueChanged.connect(self._on_interactive_preview_control_changed)
        self.analysisMarrowMaskErosion.valueChanged.connect(self._on_interactive_preview_control_changed)
        self.analysisGaussianFilterCheck.toggled.connect(self._on_interactive_preview_control_changed)
        self.analysisGaussianSigma.editingFinished.connect(self._on_interactive_preview_control_changed)
        self._set_analysis_threshold_value(self.analysisThreshold.value)
        self._set_analysis_cluster_value(self.analysisCluster.value)

        self.clearLoadedResultsBtn = qt.QPushButton("Clear loaded")
        self.clearLoadedResultsBtn.toolTip = "Remove loaded Timelapsed result nodes and clear interactive preview cache."
        self.clearLoadedResultsBtn.clicked.connect(self._on_clear_loaded_timelapsed_results)

        self.saveAnalysisScenarioBtn.clicked.connect(self._on_save_analysis_scenario)
        self.applyAnalysisSettingsBtn.clicked.connect(self._on_apply_interactive_remodelling)

        self.maskGenerationBox = maskBox
        self.maskGenerationBox.visible = False
        settingsLayout.addWidget(registrationBox)
        settingsLayout.addWidget(advancedAnalysisBox)
        analysisSectionLayout.addWidget(analysisBox)

        actionBox = qt.QGroupBox("Pipeline")
        actionBox.setStyleSheet(
            "QGroupBox { font-weight: 600; border: 1px solid #b8c7d9; border-radius: 4px; "
            "margin-top: 8px; padding: 8px 6px 6px 6px; } "
            "QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 3px; }"
        )
        actionLayout = qt.QVBoxLayout(actionBox)
        actionLayout.setContentsMargins(8, 10, 8, 8)
        actionLayout.setSpacing(6)
        self.runMasksBtn = qt.QPushButton("Generate masks")
        self.runMasksBtn.toolTip = "Generate/recompute masks from imported stacks."
        self.runMasksBtn.visible = False
        self.runTimelapseBtn = qt.QPushButton("Run pipeline")
        self.runTimelapseBtn.toolTip = (
            "Run the complete workflow using masks/ROIs discovered from the input data or Bone Contouring outputs."
        )
        self.cancelRunBtn = qt.QPushButton("✕ Cancel")
        self.cancelRunBtn.clicked.connect(self._on_cancel_run)
        self.cancelRunBtn.enabled = False
        self.cancelRunBtn.toolTip = "Cancel the currently running pipeline step."
        self._style_primary_run_button(self.runTimelapseBtn)
        self.cancelRunBtn.setStyleSheet(
            "QPushButton { background:#fff5f5; color:#9b1c1c; border:1px solid #e0b4b4; "
            "border-radius:4px; padding:5px 8px; } "
            "QPushButton:disabled { background:#eeeeee; color:#9a9a9a; border-color:#d0d0d0; }"
        )
        secondaryActionRow = qt.QWidget()
        secondaryActionLayout = qt.QHBoxLayout(secondaryActionRow)
        secondaryActionLayout.setContentsMargins(0, 0, 0, 0)
        secondaryActionLayout.setSpacing(6)
        secondaryActionLayout.addWidget(self.clearLoadedResultsBtn)
        secondaryActionLayout.addWidget(self.cancelRunBtn)
        secondaryActionLayout.addStretch(1)
        _cap_width(self.clearLoadedResultsBtn, 104)
        _cap_width(self.cancelRunBtn, 82)
        processingSubjectRow = qt.QWidget()
        processingSubjectLayout = qt.QFormLayout(processingSubjectRow)
        processingSubjectLayout.setContentsMargins(0, 0, 0, 0)
        processingSubjectLayout.setSpacing(6)
        processingSubjectLayout.addRow(
            _label("Processing subject", "Subject selected for pipeline runs. All subjects processes the cohort."),
            self.processingSubjectCombo,
        )
        processingSubjectLayout.addRow(
            _label("Processing site", "Site selected for pipeline runs. All sites processes every parsed site in scope."),
            self.processingSiteCombo,
        )
        actionLayout.addWidget(processingSubjectRow)
        actionLayout.addWidget(self.runTimelapseBtn)
        actionLayout.addWidget(secondaryActionRow)

        self.runTimelapseBtn.clicked.connect(self._on_run_full_pipeline)
        self.runAnalysisBtn.clicked.connect(self._on_run_analysis)

        statusBox = ctk.ctkCollapsibleButton()
        statusBox.text = "Pipeline Status"
        self.statusBox = statusBox
        statusForm = qt.QFormLayout(statusBox)
        self.progressBar = qt.QProgressBar()
        self.progressBar.minimum = 0
        self.progressBar.maximum = 4
        self.progressBar.value = 0
        self.progressBar.toolTip = "Current pipeline stage progress."
        self.currentStepLabel = qt.QLabel("Current step: idle")
        self.currentStepLabel.toolTip = "Currently running pipeline command or idle state."
        statusForm.addRow(_label("Progress", "Current pipeline stage progress."), self.progressBar)
        statusForm.addRow(_label("Current", "Currently running pipeline step."), self.currentStepLabel)
        self.stageLabels = {}
        for key, title in [
            ("dataset", "Dataset"),
            ("parse", "Parse"),
            ("registration", "Registration"),
            ("analysis", "Analysis"),
        ]:
            lbl = qt.QLabel("")
            lbl.wordWrap = True
            lbl.toolTip = f"Status for the {title.lower()} pipeline stage."
            self.stageLabels[key] = lbl
            statusForm.addRow(_label(title, f"Status for the {title.lower()} pipeline stage."), lbl)

        loadBox = ctk.ctkCollapsibleButton()
        loadBox.text = "Load Processed Data"
        loadForm = qt.QFormLayout(loadBox)
        self.patientCombo = qt.QComboBox()
        self.loadTypeCombo = qt.QComboBox()
        self.loadTypeCombo.addItems(
            ["remodelling image", "transformed", "raw"]
        )
        self.remodellingComparisonCombo = qt.QComboBox()
        self.loadDataBtn = qt.QPushButton("Load selected")
        _tip(self.patientCombo, "Processed subject/site available for loading into Slicer.")
        _tip(self.loadTypeCombo, "Processed output type to load into Slicer.")
        _tip(self.remodellingComparisonCombo, "Pairwise remodelling comparison to load when loading remodelling images.")
        _tip(self.loadDataBtn, "Load the selected processed output into the current Slicer scene.")
        _cap_width(self.patientCombo, 260)
        _cap_width(self.loadTypeCombo, 260)
        _cap_width(self.remodellingComparisonCombo, 260)
        _cap_width(self.loadDataBtn, 180)
        self.loadDataBtn.clicked.connect(self._on_load_selected)
        self.patientCombo.currentIndexChanged.connect(self._refresh_remodelling_comparison_list)
        self.loadTypeCombo.currentIndexChanged.connect(self._refresh_remodelling_comparison_list)
        loadForm.addRow(_label("Patient", "Processed subject/site available for loading into Slicer."), self.patientCombo)
        loadForm.addRow(_label("Data type", "Processed output type to load into Slicer."), self.loadTypeCombo)
        loadForm.addRow(_label("Comparison", "Pairwise remodelling comparison to load when loading remodelling images."), self.remodellingComparisonCombo)
        loadForm.addRow(self.loadDataBtn)

        # Internal controls used by legacy result-refresh helpers; these are not shown.
        self.remodellingFullSegCombo = qt.QComboBox()
        self.remodellingRefreshBtn = qt.QPushButton("")
        _tip(self.remodellingRefreshBtn, "Refresh the list of loaded remodelling segmentations.")
        self.remodellingRefreshBtn.clicked.connect(self._refresh_remodelling_full_selector)
        self.remodellingFullSegCombo.currentIndexChanged.connect(self._on_remodelling_selection_changed)
        self.remodellingAutoUpdateCheck = qt.QCheckBox()
        self.remodellingAutoUpdateCheck.checked = False
        self.remodellingApplyInteractiveBtn = qt.QPushButton("")
        self.remodellingApplyInteractiveBtn.clicked.connect(self._on_apply_interactive_remodelling)

        self.logText = qt.QPlainTextEdit()
        self.logText.readOnly = True
        self.logText.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        self.logText.setMinimumHeight(140)
        self.logText.setMaximumHeight(200)

        self.batchLayout.addLayout(form)
        self.batchLayout.addWidget(quickBox)
        self.batchLayout.addWidget(analysisSectionBox)
        self.batchLayout.addWidget(parseBox)
        self.batchLayout.addWidget(discoveryBox)
        self.batchLayout.addWidget(actionBox)
        self.batchLayout.addWidget(loadBox)
        self.batchLayout.addWidget(metricsBox)
        self.batchLayout.addStretch(1)
        self.layout.addWidget(statusBox)
        self.layout.addWidget(settingsBox)
        self.layout.addWidget(self.logText)
        self.layout.addStretch(1)
        self._on_mask_method_changed(self.maskMethod.currentText)
        self._on_periosteal_contour_method_changed()
        self._update_dependency_ui()
        self._set_stage_status("dataset", "pending")
        self._set_stage_status("parse", "pending")
        self._set_stage_status("registration", "pending")
        self._set_stage_status("analysis", "pending")
        self._update_progress_ui()
        self._on_timelapsed_mode_changed()

    def _style_primary_run_button(self, button):
        button.setMinimumHeight(34)
        button.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Fixed)
        button.setStyleSheet(
            "QPushButton { background:#1f6feb; color:white; border:1px solid #175cc5; "
            "border-radius:4px; padding:7px 10px; font-weight:600; } "
            "QPushButton:hover { background:#1a5fd0; } "
            "QPushButton:pressed { background:#154ea8; } "
            "QPushButton:disabled { background:#9aaec8; border-color:#8fa2ba; }"
        )

    def _build_scene_ui(self, parent):
        def _cap_width(widget, width=220):
            try:
                widget.setMaximumWidth(width)
            except Exception:
                pass

        def _label(text, help_text):
            label = qt.QLabel(str(text))
            label.toolTip = str(help_text)
            return label

        layout = qt.QVBoxLayout(parent)
        self.sceneLayout = layout
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        sceneRegistrationBox = qt.QGroupBox("Timepoints")
        sceneRegistrationLayout = qt.QVBoxLayout(sceneRegistrationBox)
        sceneRegistrationLayout.setContentsMargins(6, 8, 6, 6)
        sceneRegistrationLayout.setSpacing(4)

        sceneDiscoveryRow = qt.QWidget()
        sceneDiscoveryLayout = qt.QHBoxLayout(sceneDiscoveryRow)
        sceneDiscoveryLayout.setContentsMargins(0, 0, 0, 0)
        sceneDiscoveryLayout.setSpacing(6)
        self.sceneDiscoverButton = qt.QPushButton("Discover Loaded Timepoints")
        self.sceneAddTimepointButton = qt.QPushButton("Add timepoint")
        self.sceneRemoveTimepointButton = qt.QPushButton("Remove timepoint")
        self.sceneMoveUpButton = qt.QPushButton("Move up")
        self.sceneMoveDownButton = qt.QPushButton("Move down")
        sceneDiscoveryLayout.addWidget(self.sceneDiscoverButton)
        sceneDiscoveryLayout.addWidget(self.sceneAddTimepointButton)
        sceneDiscoveryLayout.addWidget(self.sceneRemoveTimepointButton)
        sceneDiscoveryLayout.addWidget(self.sceneMoveUpButton)
        sceneDiscoveryLayout.addWidget(self.sceneMoveDownButton)
        sceneDiscoveryLayout.addStretch(1)
        sceneRegistrationLayout.addWidget(sceneDiscoveryRow)

        self.sceneAppendDiscoveryCheck = qt.QCheckBox("Append to table")
        self.sceneAppendDiscoveryCheck.checked = False
        self.sceneAppendDiscoveryCheck.toolTip = "Append discovered loaded timepoints instead of replacing the current table."
        sceneRegistrationLayout.addWidget(self.sceneAppendDiscoveryCheck)

        self.sceneInputHintLabel = qt.QLabel(
            "Scene mode works from loaded Slicer nodes. Use Discover to guess timepoints, then edit the table if needed."
        )
        self.sceneInputHintLabel.wordWrap = True
        self.sceneInputHintLabel.toolTip = (
            "Loaded nodes may have arbitrary names. The module writes a standardized processing dataset before running."
        )
        sceneRegistrationLayout.addWidget(self.sceneInputHintLabel)

        sceneWorkspaceRow = qt.QWidget()
        sceneWorkspaceLayout = qt.QFormLayout(sceneWorkspaceRow)
        sceneWorkspaceLayout.setContentsMargins(0, 0, 0, 0)
        sceneWorkspaceLayout.setVerticalSpacing(4)
        sceneWorkspaceLayout.setLabelAlignment(qt.Qt.AlignRight | qt.Qt.AlignVCenter)
        self.sceneResultsRootPath = ctk.ctkPathLineEdit()
        self.sceneResultsRootPath.filters = ctk.ctkPathLineEdit.Dirs
        self.sceneResultsRootPath.setCurrentPath(str(self._default_scene_results_root()))
        self.sceneResultsRootPath.toolTip = (
            "Folder used for standardized scene-run inputs and outputs. "
            "Selected scene nodes are written into pipeline-readable names before running."
        )
        _cap_width(self.sceneResultsRootPath, 360)
        sceneWorkspaceLayout.addRow(
            _label("Processing workspace", "Folder used for standardized scene-run inputs and outputs."),
            self.sceneResultsRootPath,
        )
        sceneRegistrationLayout.addWidget(sceneWorkspaceRow)

        self.sceneRegistrationTable = qt.QTableWidget()
        self.sceneRegistrationTable.setColumnCount(3)
        self.sceneRegistrationTable.setHorizontalHeaderLabels(
            [
                "Session",
                "Image",
                "Masks / segments",
            ]
        )
        self._configure_scene_registration_table_columns()
        self.sceneRegistrationTable.setVerticalScrollBarPolicy(qt.Qt.ScrollBarAsNeeded)
        sceneRegistrationLayout.addWidget(self.sceneRegistrationTable)
        self.sceneTimepointTable = self.sceneRegistrationTable

        sceneRoiBox = ctk.ctkCollapsibleButton()
        sceneRoiBox.text = "Role Mapping"
        sceneRoiBox.collapsed = True
        sceneRoiBox.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        sceneRoiBox.toolTip = (
            "Advanced scene role assignments. Use this only when automatic matching or segment-order mapping needs correction."
        )
        sceneRoiLayout = qt.QVBoxLayout(sceneRoiBox)
        sceneRoiLayout.setContentsMargins(8, 10, 8, 8)
        sceneRoiLayout.setSpacing(4)
        self.sceneRoiTable = qt.QTableWidget()
        self.sceneRoiTable.setColumnCount(3)
        self.sceneRoiTable.setHorizontalHeaderLabels(["Use", "Role", "Status"])
        self.sceneRoiTable.horizontalHeader().setSectionResizeMode(qt.QHeaderView.Stretch)
        self.sceneRoiTable.setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Maximum)
        self.sceneRoiTable.setMinimumWidth(420)
        self.sceneRoiTable.setVerticalScrollBarPolicy(qt.Qt.ScrollBarAsNeeded)
        sceneRoiLayout.addWidget(self.sceneRoiTable)
        sceneRoiLayout.addSpacing(6)
        self.sceneManualMappingCheck = qt.QCheckBox("Edit per-timepoint mapping")
        self.sceneManualMappingCheck.checked = False
        self.sceneManualMappingCheck.toolTip = "Show one mapping dropdown per timepoint. Leave hidden when automatic matching is correct."
        sceneRoiLayout.addWidget(self.sceneManualMappingCheck)
        roiActions = qt.QHBoxLayout()
        self.sceneAddRoiButton = qt.QPushButton("Add ROI")
        self.sceneRemoveRoiButton = qt.QPushButton("Remove ROI")
        self.sceneAutoRoiButton = qt.QPushButton("Auto-detect roles")
        roiActions.addWidget(self.sceneAddRoiButton)
        roiActions.addWidget(self.sceneRemoveRoiButton)
        roiActions.addWidget(self.sceneAutoRoiButton)
        roiActions.addStretch(1)
        sceneRoiLayout.addLayout(roiActions)
        sceneRegistrationLayout.addWidget(sceneRoiBox)
        self._resize_scene_timepoint_table()
        layout.addWidget(sceneRegistrationBox)

        sceneProfileBox = qt.QGroupBox("Study Profile")
        self.sceneProfileBox = sceneProfileBox
        form = qt.QFormLayout(sceneProfileBox)
        form.setContentsMargins(6, 8, 6, 6)
        form.setVerticalSpacing(4)
        self.sceneProfileCombo = qt.QComboBox()
        self._populate_study_profiles(self.sceneProfileCombo)
        self.sceneProfileCombo.currentIndexChanged.connect(self._on_scene_profile_changed)
        try:
            self.sceneProfileCombo.activated.connect(self._on_scene_profile_changed)
        except Exception:
            pass
        try:
            self.sceneProfileCombo.currentTextChanged.connect(self._on_scene_profile_changed)
        except Exception:
            pass
        self.sceneProfileCombo.toolTip = "Study defaults applied to scene runs."
        _cap_width(self.sceneProfileCombo)
        form.addRow(_label("Profile", "Study defaults used for registration, mask generation, and remodelling analysis."), self.sceneProfileCombo)
        layout.addWidget(sceneProfileBox)

        sceneActionBox = qt.QGroupBox("Pipeline")
        sceneActionLayout = qt.QVBoxLayout(sceneActionBox)
        sceneActionLayout.setContentsMargins(6, 8, 6, 6)
        sceneActionLayout.setSpacing(6)
        self.sceneRunButton = qt.QPushButton("Run")
        self.sceneInterruptButton = qt.QPushButton("✕ Cancel")
        self.sceneExportCsvButton = qt.QPushButton("Export CSV")
        self.sceneClearLoadedButton = qt.QPushButton("Clear loaded")
        self._style_primary_run_button(self.sceneRunButton)
        self.sceneInterruptButton.enabled = False
        self.sceneInterruptButton.toolTip = "Interrupt the currently running scene pipeline."
        self.sceneExportCsvButton.toolTip = "Export the current scene comparison rows to CSV."
        self.sceneClearLoadedButton.toolTip = "Remove loaded Timelapsed result nodes and clear interactive preview cache."
        self.sceneExportCsvButton.setStyleSheet(
            "QPushButton { background:#f5f7fa; color:#222; border:1px solid #b8c0ca; "
            "border-radius:4px; padding:5px 8px; } "
            "QPushButton:hover { background:#edf2f7; }"
        )
        self.sceneInterruptButton.setStyleSheet(
            "QPushButton { background:#fff5f5; color:#9b1c1c; border:1px solid #e0b4b4; "
            "border-radius:4px; padding:5px 8px; } "
            "QPushButton:disabled { background:#eeeeee; color:#9a9a9a; border-color:#d0d0d0; }"
        )
        self.sceneDiscoverButton.clicked.connect(self._on_discover_scene_timepoints)
        self.sceneAddTimepointButton.clicked.connect(self._add_scene_timepoint)
        self.sceneRemoveTimepointButton.clicked.connect(self._remove_scene_timepoint)
        self.sceneMoveUpButton.clicked.connect(lambda _checked=False: self._move_scene_timepoint(-1))
        self.sceneMoveDownButton.clicked.connect(lambda _checked=False: self._move_scene_timepoint(1))
        self.sceneRunButton.clicked.connect(self._on_run_scene_pipeline)
        self.sceneInterruptButton.clicked.connect(self._on_cancel_run)
        self.sceneExportCsvButton.clicked.connect(self._on_export_scene_comparison_csv)
        self.sceneClearLoadedButton.clicked.connect(self._on_clear_loaded_timelapsed_results)
        self.sceneAddRoiButton.clicked.connect(self._add_scene_roi)
        self.sceneRemoveRoiButton.clicked.connect(self._remove_scene_roi)
        self.sceneAutoRoiButton.clicked.connect(self._auto_detect_scene_rois)
        self.sceneManualMappingCheck.toggled.connect(lambda _checked=False: self._update_scene_role_mapping_visibility())
        sceneSecondaryActionRow = qt.QWidget()
        sceneSecondaryActionLayout = qt.QHBoxLayout(sceneSecondaryActionRow)
        sceneSecondaryActionLayout.setContentsMargins(0, 0, 0, 0)
        sceneSecondaryActionLayout.setSpacing(6)
        sceneSecondaryActionLayout.addWidget(self.sceneInterruptButton)
        sceneSecondaryActionLayout.addWidget(self.sceneClearLoadedButton)
        sceneSecondaryActionLayout.addStretch(1)
        _cap_width(self.sceneExportCsvButton, 96)
        _cap_width(self.sceneClearLoadedButton, 104)
        _cap_width(self.sceneInterruptButton, 92)
        sceneActionLayout.addWidget(self.sceneRunButton)
        sceneActionLayout.addWidget(sceneSecondaryActionRow)
        layout.addWidget(sceneActionBox)

        self.sceneComparisonBox = qt.QGroupBox("Current Comparisons")
        sceneComparisonLayout = qt.QVBoxLayout(self.sceneComparisonBox)
        sceneComparisonLayout.setContentsMargins(8, 10, 8, 8)
        sceneComparisonLayout.setSpacing(4)
        self.sceneComparisonTable = qt.QTableWidget()
        self.sceneComparisonTable.setColumnCount(6)
        self.sceneComparisonTable.setHorizontalHeaderLabels(["Pair", "Mask", "FV/BV", "RV/BV", "AV/BV", "NV/BV"])
        self.sceneComparisonTable.setEditTriggers(qt.QAbstractItemView.NoEditTriggers)
        self.sceneComparisonTable.setSelectionMode(qt.QAbstractItemView.NoSelection)
        self.sceneComparisonTable.verticalHeader().setVisible(False)
        self.sceneComparisonTable.horizontalHeader().setStretchLastSection(True)
        self.sceneComparisonTable.setAlternatingRowColors(True)
        self.sceneComparisonTable.setMinimumHeight(82)
        self.sceneComparisonTable.setMaximumHeight(140)
        self.sceneComparisonTable.toolTip = "Pairwise remodelling fractions from the current scene run."
        sceneComparisonLayout.addWidget(self.sceneComparisonTable)
        sceneComparisonLayout.addWidget(self.sceneExportCsvButton)
        self._set_scene_comparison_rows([])

        self.sceneStatusBox = qt.QGroupBox("Pipeline Status")
        sceneStatusBox = self.sceneStatusBox
        sceneStatusLayout = qt.QVBoxLayout(sceneStatusBox)
        sceneStatusLayout.setContentsMargins(8, 10, 8, 8)
        sceneStatusLayout.setSpacing(2)
        self.sceneStageItems = {}
        self.sceneStageRows = {}
        for key, title in [
            ("dataset", "Dataset"),
            ("parse", "Parse"),
            ("registration", "Registration"),
            ("analysis", "Analysis"),
        ]:
            rowWidget = qt.QWidget()
            rowWidget.setFixedHeight(22)
            rowLayout = qt.QHBoxLayout(rowWidget)
            rowLayout.setContentsMargins(0, 0, 0, 0)
            rowLayout.setSpacing(8)
            stage_label = qt.QLabel(title)
            stage_label.setMinimumWidth(72)
            stage_label.setSizePolicy(qt.QSizePolicy.Fixed, qt.QSizePolicy.Fixed)
            status_label = qt.QLabel("● Pending")
            status_label.setMinimumWidth(84)
            status_label.setSizePolicy(qt.QSizePolicy.Fixed, qt.QSizePolicy.Fixed)
            rowLayout.addWidget(stage_label)
            rowLayout.addWidget(status_label)
            rowLayout.addStretch(1)
            sceneStatusLayout.addWidget(rowWidget)
            self.sceneStageRows[key] = rowWidget
            self.sceneStageItems[key] = status_label
        layout.addWidget(sceneStatusBox)
        self.sceneStatusLabel = qt.QLabel("")
        self.sceneStatusLabel.wordWrap = True
        layout.addWidget(self.sceneStatusLabel)

        layout.addWidget(self.sceneComparisonBox)

    def _configure_scene_registration_table_columns(self):
        header = self.sceneRegistrationTable.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, qt.QHeaderView.Fixed)
        header.setSectionResizeMode(1, qt.QHeaderView.Stretch)
        header.setSectionResizeMode(2, qt.QHeaderView.Stretch)
        self.sceneRegistrationTable.setColumnWidth(0, 68)

    def _configure_scene_roi_table_columns(self):
        header = self.sceneRoiTable.horizontalHeader()
        header.setStretchLastSection(False)
        if self.sceneRoiTable.columnCount > 0:
            header.setSectionResizeMode(0, qt.QHeaderView.Fixed)
            self.sceneRoiTable.setColumnWidth(0, 42)
        if self.sceneRoiTable.columnCount > 1:
            header.setSectionResizeMode(1, qt.QHeaderView.Fixed)
            self.sceneRoiTable.setColumnWidth(1, 118)
        for column in range(2, max(2, self.sceneRoiTable.columnCount - 1)):
            header.setSectionResizeMode(column, qt.QHeaderView.Stretch)
        if self.sceneRoiTable.columnCount > 2:
            header.setSectionResizeMode(self.sceneRoiTable.columnCount - 1, qt.QHeaderView.Fixed)
            self.sceneRoiTable.setColumnWidth(self.sceneRoiTable.columnCount - 1, 92)
        self._update_scene_role_mapping_visibility()

    def _update_scene_role_mapping_visibility(self):
        if not hasattr(self, "sceneRoiTable"):
            return
        show_mapping = bool(getattr(self.sceneManualMappingCheck, "checked", False)) if hasattr(self, "sceneManualMappingCheck") else False
        status_column = self.sceneRoiTable.columnCount - 1
        header = self.sceneRoiTable.horizontalHeader()
        for column in range(self.sceneRoiTable.columnCount):
            hidden = 2 <= column < status_column and not show_mapping
            try:
                self.sceneRoiTable.setColumnHidden(column, hidden)
            except Exception:
                pass
        try:
            if show_mapping:
                header.setSectionResizeMode(1, qt.QHeaderView.Fixed)
                self.sceneRoiTable.setColumnWidth(1, 118)
                for column in range(2, status_column):
                    header.setSectionResizeMode(column, qt.QHeaderView.Stretch)
                header.setSectionResizeMode(status_column, qt.QHeaderView.Fixed)
                self.sceneRoiTable.setColumnWidth(status_column, 92)
            else:
                header.setSectionResizeMode(1, qt.QHeaderView.Stretch)
                header.setSectionResizeMode(status_column, qt.QHeaderView.Fixed)
                self.sceneRoiTable.setColumnWidth(status_column, 92)
        except Exception:
            pass

    def _resize_scene_timepoint_table(self):
        timepoint_visible_rows = self._scene_timepoint_visible_rows()
        def _resize_table(table, visible_rows):
            try:
                header_height = int(table.horizontalHeader().height())
                row_height = int(table.verticalHeader().defaultSectionSize())
            except Exception:
                header_height = 28
                row_height = 30
            roi_extra_padding = 8 if table is self.sceneRoiTable else 12
            height = header_height + visible_rows * row_height + roi_extra_padding
            table.setMinimumHeight(height)
            table.setMaximumHeight(height)
            table.resizeRowsToContents()
            try:
                table.viewport().update()
                table.updateGeometry()
                if table.parent() is not None and table.parent().layout() is not None:
                    table.parent().layout().activate()
            except Exception:
                pass
            return height
        height = _resize_table(self.sceneRegistrationTable, timepoint_visible_rows)
        if hasattr(self, "sceneRoiTable"):
            roi_visible_rows = max(4, min(int(self.sceneRoiTable.rowCount), 8))
            _resize_table(self.sceneRoiTable, roi_visible_rows)
        self._scene_timepoint_table_height = height
        self._resize_timelapsed_mode_tabs()

    def _scene_timepoint_visible_rows(self):
        row_count = int(self.sceneRegistrationTable.rowCount)
        return max(2, min(row_count, 8))

    def _on_timelapsed_mode_changed(self, *_args):
        self._resize_timelapsed_mode_tabs()
        if not hasattr(self, "timelapsedModeTabs"):
            return
        self._place_analysis_options_for_mode()
        scene_mode = self._timelapsed_scene_mode_selected()
        if hasattr(self, "runAnalysisBtn"):
            self.runAnalysisBtn.visible = not scene_mode
        if hasattr(self, "statusBox"):
            self.statusBox.visible = not scene_mode
        if hasattr(self, "sceneStatusBox"):
            self.sceneStatusBox.visible = scene_mode
        if hasattr(self, "doNotGenerateMasksCheck"):
            self.doNotGenerateMasksCheck.visible = not scene_mode
        if hasattr(self, "doNotGenerateMasksLabel"):
            self.doNotGenerateMasksLabel.visible = not scene_mode
        self._update_batch_analysis_options_visibility()

    def _place_analysis_options_for_mode(self):
        if not hasattr(self, "analysisSectionBox"):
            return
        if hasattr(self, "sceneLayout"):
            self.sceneLayout.removeWidget(self.analysisSectionBox)
        if hasattr(self, "batchLayout"):
            self.batchLayout.removeWidget(self.analysisSectionBox)
        if self._timelapsed_scene_mode_selected() and hasattr(self, "sceneLayout"):
            self.sceneLayout.insertWidget(2, self.analysisSectionBox)
        elif hasattr(self, "batchLayout"):
            self.batchLayout.insertWidget(2, self.analysisSectionBox)

    def _timelapsed_scene_mode_selected(self):
        if not hasattr(self, "timelapsedModeTabs"):
            return False
        current_index = self.timelapsedModeTabs.currentIndex
        if callable(current_index):
            current_index = current_index()
        return int(current_index) == 0

    def _update_batch_analysis_options_visibility(self):
        if not hasattr(self, "analysisSectionBox"):
            return
        scene_mode = self._timelapsed_scene_mode_selected()
        custom = self._selected_profile_is_custom()
        self.analysisSectionBox.visible = scene_mode or custom
        if custom:
            self.analysisSectionBox.collapsed = False

    def _resize_timelapsed_mode_tabs(self):
        if not hasattr(self, "timelapsedModeTabs"):
            return
        current_index = self.timelapsedModeTabs.currentIndex
        if callable(current_index):
            current_index = current_index()
        self.timelapsedModeTabs.setMaximumHeight(16777215)
        try:
            self.timelapsedModeTabs.updateGeometry()
            if self.timelapsedModeTabs.parent() is not None and self.timelapsedModeTabs.parent().layout() is not None:
                self.timelapsedModeTabs.parent().layout().activate()
        except Exception:
            pass

    def _default_scene_results_root(self):
        return Path(tempfile.gettempdir()) / "SlicerBoneImagingToolbox" / "TimelapsedScene"

    def _on_scene_profile_changed(self, *_args):
        scene_combo = getattr(self, "sceneProfileCombo", None)
        study_combo = getattr(self, "studyProfileCombo", None)
        if not self._qt_object_alive(scene_combo) or not self._qt_object_alive(study_combo):
            return
        try:
            selected = self._combo_current_data_safe(scene_combo)
            index = study_combo.findData(selected)
            if index >= 0 and study_combo.currentIndex != index:
                previous = study_combo.blockSignals(True)
                try:
                    study_combo.setCurrentIndex(index)
                finally:
                    study_combo.blockSignals(previous)
            self._apply_profile_analysis_controls(selected)
            self._on_apply_study_profile(profile=selected)
        except (RuntimeError, ValueError):
            return

    def _sync_scene_profile_from_batch_profile(self):
        scene_combo = getattr(self, "sceneProfileCombo", None)
        study_combo = getattr(self, "studyProfileCombo", None)
        if not self._qt_object_alive(scene_combo) or not self._qt_object_alive(study_combo):
            return
        try:
            selected = study_combo.currentData
            index = scene_combo.findData(selected)
            if index < 0 or scene_combo.currentIndex == index:
                return
            previous = scene_combo.blockSignals(True)
            try:
                scene_combo.setCurrentIndex(index)
            finally:
                scene_combo.blockSignals(previous)
        except (RuntimeError, ValueError):
            return

    def _scene_node_selector(self, node_types):
        selector = slicer.qMRMLNodeComboBox()
        selector.nodeTypes = node_types
        selector.noneEnabled = True
        selector.addEnabled = False
        selector.setMRMLScene(slicer.mrmlScene)
        return selector

    def _scene_segmentation_node_selector(self):
        selector = self._scene_node_selector(["vtkMRMLSegmentationNode"])
        try:
            selector.currentNodeChanged.connect(lambda _node=None, selector=selector: self._on_scene_mask_source_changed(selector))
        except Exception:
            pass
        return selector

    def _scene_transform_is_supported_initial_transform(self, transform_node):
        if transform_node is None:
            return False
        if transform_node.IsA("vtkMRMLLinearTransformNode"):
            return True
        stored_transform_path = self._scene_transform_storage_path(transform_node)
        return stored_transform_path is not None and stored_transform_path.suffix.lower() == ".tfm"

    def _add_scene_combo_item(self, selector, label, data, tooltip=None):
        selector.addItem(label, data)
        index = selector.count - 1
        tooltip_text = str(tooltip or label)
        try:
            selector.setItemData(index, tooltip_text, qt.Qt.ToolTipRole)
        except Exception:
            try:
                selector.model().setData(selector.model().index(index, 0), tooltip_text, qt.Qt.ToolTipRole)
            except Exception:
                pass

    def _scene_mask_selector(self, timepoint_index=None):
        selector = qt.QComboBox()
        self._add_scene_combo_item(selector, "None", "__none__", "No mask/ROI selected for this timepoint.")
        source_node_id = self._scene_mask_source_node_id_for_timepoint(timepoint_index)
        if source_node_id:
            source_node = slicer.mrmlScene.GetNodeByID(source_node_id)
            for segment_id, segment_name, segment_role in self._scene_segmentation_segment_choices(source_node):
                suffix = f" ({segment_role})" if segment_role else ""
                label = f"{segment_name}{suffix}"
                tooltip = (
                    f"Segmentation: {source_node.GetName() or source_node_id}\n"
                    f"Segment: {segment_name}\n"
                    f"Role: {segment_role or 'unlabeled'}"
                )
                self._add_scene_combo_item(
                    selector,
                    label,
                    _encode_scene_mask_choice(source_node_id, segment_id),
                    tooltip,
                )
            selector.toolTip = "Use a segment from this timepoint's selected Masks / segments node."
            return selector
        for candidate in self._scene_node_candidates():
            node_class = str(candidate.node_class or "")
            if "LabelMapVolume" in node_class:
                label = str(candidate.name or candidate.node_id)
                self._add_scene_combo_item(selector, label, str(candidate.node_id), label)
                continue
            if "Segmentation" not in node_class:
                continue
            node = slicer.mrmlScene.GetNodeByID(str(candidate.node_id))
            for segment_id, segment_name, segment_role in self._scene_segmentation_segment_choices(node):
                suffix = f" ({segment_role})" if segment_role else ""
                label = f"{segment_name}{suffix}"
                tooltip = (
                    f"Segmentation: {candidate.name}\n"
                    f"Segment: {segment_name}\n"
                    f"Role: {segment_role or 'unlabeled'}"
                )
                self._add_scene_combo_item(
                    selector,
                    label,
                    _encode_scene_mask_choice(candidate.node_id, segment_id),
                    tooltip,
                )
        selector.toolTip = "Use a loaded mask/ROI segment or leave it absent. Generate missing masks in Bone Contouring first."
        return selector

    def _scene_mask_source_node_id_for_timepoint(self, timepoint_index):
        if timepoint_index is None or not hasattr(self, "sceneRegistrationTable"):
            return ""
        try:
            node = self._scene_selected_table_node(int(timepoint_index), 2, self.sceneRegistrationTable)
        except Exception:
            node = None
        if node is not None and node.IsA("vtkMRMLSegmentationNode"):
            return str(node.GetID())
        return ""

    def _on_scene_mask_source_changed(self, selector=None):
        if not hasattr(self, "sceneRoiTable"):
            return
        role_rows = self._scene_role_rows()
        self._refresh_scene_roi_columns()
        for row, role_row in enumerate(role_rows):
            if row >= self.sceneRoiTable.rowCount:
                break
            self._restore_scene_role_row(row, role_row)
            self._update_scene_role_status(row)
        timepoint_index = self._scene_timepoint_index_for_mask_source_selector(selector)
        if timepoint_index >= 0:
            source_node = self._scene_selected_table_node(timepoint_index, 2, self.sceneRegistrationTable)
            self._apply_scene_detected_roles_for_timepoint(timepoint_index, source_node)
        self._update_scene_role_mapping_visibility()

    def _scene_timepoint_index_for_mask_source_selector(self, selector):
        if selector is None or not hasattr(self, "sceneRegistrationTable"):
            return -1
        for row in range(self.sceneRegistrationTable.rowCount):
            if self.sceneRegistrationTable.cellWidget(row, 2) is selector:
                return row
        return -1

    def _apply_scene_detected_roles_for_timepoint(self, timepoint_index, source_node):
        if source_node is None or not source_node.IsA("vtkMRMLSegmentationNode"):
            return
        source_node_id = str(source_node.GetID())
        column = 2 + int(timepoint_index)
        for role in ("registration_roi", "segmentation", "roi1", "roi2", "roi3"):
            role_row = self._scene_role_row_index(role)
            if role_row < 0:
                self._add_scene_required_role(role) if not self._scene_role_is_analysis_roi(role) else self._add_scene_roi(role)
                role_row = self._scene_role_row_index(role)
            if role_row < 0:
                continue
            lookup_role = "full" if self._normalize_scene_role_name(role) == "registration_roi" else role
            segment_id = self._scene_segment_id_for_node_role(source_node_id, lookup_role)
            if not segment_id and lookup_role != role:
                segment_id = self._scene_segment_id_for_node_role(source_node_id, role)
            if not segment_id:
                continue
            self._set_scene_mask_row_node(role_row, column, source_node_id, self.sceneRoiTable, role=role, segment_id=segment_id)
            self._set_scene_mask_row_policy(role_row, column, "node", self.sceneRoiTable)
            self._update_scene_role_status(role_row)

    def _scene_segmentation_segment_choices(self, segmentation_node):
        if segmentation_node is None or not segmentation_node.IsA("vtkMRMLSegmentationNode"):
            return []
        segmentation = segmentation_node.GetSegmentation()
        choices = []
        for index in range(segmentation.GetNumberOfSegments()):
            segment_id = segmentation.GetNthSegmentID(index)
            segment = segmentation.GetSegment(segment_id)
            if segment is None:
                continue
            name = str(segment.GetName() or segment_id)
            role = ""
            try:
                role = str(segment.GetTag("HRpQCT.Role") or segment.GetTag("BoneContouring.Role") or "")
            except Exception:
                role = ""
            choices.append((str(segment_id), name, role))
        return choices

    def _refresh_scene_roi_columns(self):
        if not hasattr(self, "sceneRoiTable"):
            return
        session_labels = []
        for row in range(self.sceneRegistrationTable.rowCount):
            item = self.sceneRegistrationTable.item(row, 0)
            session_labels.append(item.text() if item is not None else f"ses-{row + 1}")
        current_roles = self._scene_role_rows() if self.sceneRoiTable.columnCount > 3 else []
        self.sceneRoiTable.setColumnCount(3 + len(session_labels))
        self.sceneRoiTable.setHorizontalHeaderLabels(["Use", "Role"] + session_labels + ["Status"])
        self._configure_scene_roi_table_columns()
        for row, role_row in enumerate(current_roles):
            if row >= self.sceneRoiTable.rowCount:
                break
            self._configure_scene_role_row(row, role_row["role"], role_row.get("include", True))
            self._restore_scene_role_row(row, role_row)
            self._update_scene_role_status(row)
        self._ensure_scene_required_role_rows()

    def _add_scene_roi(self, role=None, nodes_by_session=None):
        if not hasattr(self, "sceneRoiTable"):
            return
        row = self.sceneRoiTable.rowCount
        self.sceneRoiTable.insertRow(row)
        role_text = str(role or f"roi{row + 1}")
        if self.sceneRoiTable.columnCount != self.sceneRegistrationTable.rowCount + 3:
            self._refresh_scene_roi_columns()
        self._configure_scene_role_row(row, role_text, True)
        nodes_by_session = dict(nodes_by_session or {})
        for timepoint_index in range(self.sceneRegistrationTable.rowCount):
            column = 2 + timepoint_index
            selector = self._scene_mask_selector(timepoint_index=timepoint_index)
            self._connect_scene_role_selector_status(selector)
            self.sceneRoiTable.setCellWidget(row, column, selector)
            node_id = nodes_by_session.get(timepoint_index, "")
            if node_id:
                self._set_scene_mask_row_node(
                    row,
                    column,
                    node_id,
                    self.sceneRoiTable,
                    role=role_text,
                )
        self._configure_scene_roi_table_columns()
        self._update_scene_role_status(row)
        self._resize_scene_timepoint_table()

    def _add_scene_required_role(self, role):
        if self._scene_role_row_index(role) >= 0:
            return
        row = self.sceneRoiTable.rowCount
        self.sceneRoiTable.insertRow(row)
        self._configure_scene_role_row(row, role, True)
        self._update_scene_role_status(row)

    def _ensure_scene_required_role_rows(self):
        if not hasattr(self, "sceneRoiTable"):
            return
        for role in ("registration_roi", "initial_transform", "segmentation"):
            self._add_scene_required_role(role)

    def _configure_scene_role_row(self, row, role, include=True):
        role = self._normalize_scene_role_name(role)
        use_item = qt.QTableWidgetItem("")
        if self._scene_role_is_analysis_roi(role):
            use_item.setFlags(use_item.flags() | qt.Qt.ItemIsUserCheckable)
            use_item.setCheckState(qt.Qt.Checked if include else qt.Qt.Unchecked)
        else:
            use_item.setFlags(use_item.flags() & ~qt.Qt.ItemIsEditable)
        self.sceneRoiTable.setItem(row, 0, use_item)
        role_item = qt.QTableWidgetItem(self._scene_role_label(role))
        role_item.setData(qt.Qt.UserRole, role)
        if not self._scene_role_is_analysis_roi(role):
            role_item.setFlags(role_item.flags() & ~qt.Qt.ItemIsEditable)
        self.sceneRoiTable.setItem(row, 1, role_item)
        for timepoint_index in range(self.sceneRegistrationTable.rowCount):
            column = 2 + timepoint_index
            if role == "initial_transform":
                self.sceneRoiTable.setCellWidget(row, column, self._scene_node_selector(["vtkMRMLLinearTransformNode"]))
            else:
                selector = self._scene_mask_selector(timepoint_index=timepoint_index)
                self._connect_scene_role_selector_status(selector)
                self.sceneRoiTable.setCellWidget(row, column, selector)
        self.sceneRoiTable.setItem(row, self._scene_status_column(), qt.QTableWidgetItem(""))

    def _connect_scene_role_selector_status(self, selector):
        try:
            selector.currentIndexChanged.connect(
                lambda _index=0, selector=selector: self._update_scene_role_status_for_selector(selector)
            )
        except Exception:
            pass

    def _update_scene_role_status_for_selector(self, selector):
        if selector is None or not hasattr(self, "sceneRoiTable"):
            return
        status_column = self._scene_status_column()
        for row in range(self.sceneRoiTable.rowCount):
            for column in range(2, status_column):
                if self.sceneRoiTable.cellWidget(row, column) is selector:
                    self._update_scene_role_status(row)
                    return

    def _restore_scene_role_row(self, row, role_row):
        role = role_row["role"]
        for timepoint_index in range(self.sceneRegistrationTable.rowCount):
            column = 2 + timepoint_index
            if role == "initial_transform":
                node_id = role_row["node_ids"][timepoint_index] if timepoint_index < len(role_row["node_ids"]) else ""
                self._set_scene_row_node(row, column, node_id, self.sceneRoiTable)
                continue
            node_id = role_row["node_ids"][timepoint_index] if timepoint_index < len(role_row["node_ids"]) else ""
            segment_id = role_row["segment_ids"][timepoint_index] if timepoint_index < len(role_row["segment_ids"]) else ""
            policy = role_row["policies"][timepoint_index] if timepoint_index < len(role_row["policies"]) else "none"
            self._set_scene_mask_row_node(row, column, node_id, self.sceneRoiTable, role=role, segment_id=segment_id)
            self._set_scene_mask_row_policy(row, column, policy, self.sceneRoiTable)

    def _scene_role_rows(self):
        rows = []
        if not hasattr(self, "sceneRoiTable"):
            return rows
        for row in range(self.sceneRoiTable.rowCount):
            role = self._scene_role_at_row(row)
            if not role:
                continue
            node_ids = []
            segment_ids = []
            policies = []
            for timepoint_index in range(self.sceneRegistrationTable.rowCount):
                column = 2 + timepoint_index
                if role == "initial_transform":
                    node_ids.append(self._scene_selected_node_id(row, column, self.sceneRoiTable))
                    segment_ids.append("")
                    policies.append("node" if node_ids[-1] else "none")
                else:
                    node_ids.append(self._scene_selected_mask_node_id(row, column, self.sceneRoiTable))
                    segment_ids.append(self._scene_selected_mask_segment_id(row, column, self.sceneRoiTable))
                    policies.append(self._scene_selected_mask_policy(row, column, self.sceneRoiTable))
            rows.append(
                {
                    "role": role,
                    "include": self._scene_role_included(row),
                    "node_ids": tuple(node_ids),
                    "segment_ids": tuple(segment_ids),
                    "policies": tuple(policies),
                }
            )
        return rows

    def _scene_role_at_row(self, row):
        item = self.sceneRoiTable.item(row, 1)
        if item is None:
            return ""
        role = item.data(qt.Qt.UserRole)
        role_data = self._normalize_scene_role_name(role)
        if self._scene_role_is_analysis_roi(role_data):
            if str(item.text()).strip() == str(self._scene_role_label(role_data)).strip():
                return role_data
            return self._normalize_scene_role_name(item.text())
        return self._normalize_scene_role_name(role_data or item.text())

    def _scene_role_row_index(self, role):
        normalized = self._normalize_scene_role_name(role)
        for row in range(self.sceneRoiTable.rowCount):
            if self._scene_role_at_row(row) == normalized:
                return row
        return -1

    def _scene_role_included(self, row):
        role = self._scene_role_at_row(row)
        if not self._scene_role_is_analysis_roi(role):
            return True
        item = self.sceneRoiTable.item(row, 0)
        return item is None or item.checkState() == qt.Qt.Checked

    def _scene_role_is_analysis_roi(self, role):
        return self._normalize_scene_role_name(role) not in {"registration_roi", "initial_transform", "segmentation"}

    def _scene_role_display_name(self, role):
        return self._scene_role_label(role)

    def _scene_role_display_labels(self):
        labels = {}
        if not hasattr(self, "sceneRoiTable"):
            return labels
        for row in range(self.sceneRoiTable.rowCount):
            role = self._scene_role_at_row(row)
            if not self._scene_role_is_analysis_roi(role):
                continue
            item = self.sceneRoiTable.item(row, 1)
            label = str(item.text() if item is not None else "").strip()
            if label:
                labels[role] = label
        return labels

    def _scene_display_compartment_name(self, compartment):
        normalized = self._normalize_scene_role_name(compartment)
        table_labels = self._scene_role_display_labels()
        if normalized in table_labels:
            return table_labels[normalized]
        labels = {
            "roi1": "full",
            "roi2": "trab",
            "roi3": "cort",
            "roi_union": "ROI union",
        }
        return labels.get(normalized, str(compartment or "full"))

    def _scene_role_label(self, role):
        labels = dict(getattr(self, "_scene_role_labels", {}) or {})
        labels.update({
            "registration_roi": "Registration ROI",
            "initial_transform": "Initial transform",
            "segmentation": "Segmentation",
            "roi1": "full",
            "roi2": "trab",
            "roi3": "cort",
        })
        normalized = self._normalize_scene_role_name(role)
        return labels.get(normalized, normalized)

    def _normalize_scene_role_name(self, role):
        value = re.sub(r"[^a-zA-Z0-9]+", "_", str(role or "").strip().lower()).strip("_")
        if value in {"registration", "reg", "regmask", "registration_mask"}:
            return "registration_roi"
        if value in {"seg", "bone_seg", "bone_segmentation", "bone_mask"}:
            return "segmentation"
        if value in {"transform", "initial_registration", "initial_tfm"}:
            return "initial_transform"
        return value or "roi"

    def _scene_compartment_is_interactive_source(self, compartment):
        normalized = self._normalize_scene_role_name(compartment)
        return normalized not in {"roi_union"}

    def _scene_remodelling_context_matches_pair(self, candidate_ctx, target_ctx):
        if not candidate_ctx or not target_ctx:
            return False
        for key in ("subject_id", "site", "t0", "t1"):
            if str(candidate_ctx.get(key, "")) != str(target_ctx.get(key, "")):
                return False
        return True

    def _find_loaded_interactive_remodelling_node_for_context(self, target_ctx, preferred_compartment="full"):
        scene = slicer.mrmlScene
        matches = []
        preferred = str(preferred_compartment or "full")
        for class_name in ("vtkMRMLScalarVolumeNode", "vtkMRMLLabelMapVolumeNode", "vtkMRMLSegmentationNode"):
            for index in range(scene.GetNumberOfNodesByClass(class_name)):
                node = scene.GetNthNodeByClass(index, class_name)
                if node is None:
                    continue
                if str(node.GetAttribute("TimelapsedHRpQCT.RemodellingFull") or "") != "1":
                    continue
                source_path = str(node.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "")
                if not source_path:
                    continue
                ctx = self._parse_remodelling_source_context(source_path)
                if not self._scene_remodelling_context_matches_pair(ctx, target_ctx):
                    continue
                compartment = str((ctx or {}).get("compartment", ""))
                if not self._scene_compartment_is_interactive_source(compartment):
                    continue
                rank = 0 if compartment == preferred else 1
                matches.append((rank, self._remodelling_source_sort_key(source_path), node))
        if not matches:
            return None
        return sorted(matches, key=lambda item: (item[0], item[1]))[0][2]

    def _compute_pair_union_remodelling_preview(self, target_ctx, source_path=None):
        from timelapsedhrpqct.analysis import build_series_common_masks

        source_node = self._find_loaded_interactive_remodelling_node_for_context(
            target_ctx,
            preferred_compartment="full",
        )
        source_path = str(source_path or "")
        if not source_path and source_node is None:
            raise ValueError("No loaded per-ROI remodelling source is available for this pair.")
        if not source_path:
            source_path = str(source_node.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "")
        preview_inputs = self._get_interactive_preview_inputs(source_path)
        support_t0 = np.asarray(preview_inputs["support_mask_t0"], dtype=bool)
        support_t1 = np.asarray(preview_inputs["support_mask_t1"], dtype=bool)
        valid_union = np.zeros_like(support_t0, dtype=bool)
        compartments = self._pair_metric_compartments(preview_inputs)
        for compartment in compartments:
            comp_t0, comp_t1 = self._preview_compartment_masks(preview_inputs, compartment)
            valid = build_series_common_masks(
                {
                    "full": [support_t0, support_t1],
                    compartment: [comp_t0, comp_t1],
                },
                [compartment],
                int(self._analysis_erosion_voxels),
                full_mask_dilation_voxels=int(self.analysisFullMaskDilation.value),
            )[compartment]
            valid_union |= np.asarray(valid, dtype=bool)
        preview = self._compute_pair_remodelling_preview_from_cached_delta(
            preview_inputs,
            valid_mask=valid_union,
            label_map=self._interactive_preview_label_map(),
        )
        preview_inputs["current_label_arr"] = preview.label_image
        return preview, preview_inputs, source_node, compartments

    def _scene_status_column(self):
        return self.sceneRoiTable.columnCount - 1

    def _update_scene_role_status(self, row):
        if not hasattr(self, "sceneRoiTable") or row < 0:
            return
        role = self._scene_role_at_row(row)
        total = int(self.sceneRegistrationTable.rowCount)
        present = 0
        tooltip_lines = []
        for timepoint_index in range(total):
            column = 2 + timepoint_index
            session_item = self.sceneRegistrationTable.item(timepoint_index, 0)
            session_label = session_item.text() if session_item is not None else f"ses-{timepoint_index + 1}"
            if role == "initial_transform":
                node_id = self._scene_selected_node_id(row, column, self.sceneRoiTable)
                present += 1 if node_id else 0
                tooltip_lines.append(f"{session_label}: {self._scene_node_name(node_id) if node_id else 'None'}")
            else:
                node_id = self._scene_selected_mask_node_id(row, column, self.sceneRoiTable)
                segment_id = self._scene_selected_mask_segment_id(row, column, self.sceneRoiTable)
                selected = self._scene_selected_mask_policy(row, column, self.sceneRoiTable) != "none"
                present += 1 if selected else 0
                tooltip_lines.append(f"{session_label}: {self._scene_mask_choice_name(node_id, segment_id) if selected else 'None'}")
        if role == "initial_transform":
            text = f"Optional {present}/{total}"
        elif present == total and total:
            text = f"Ready {present}/{total}"
        elif present:
            text = f"Missing {total - present}/{total}"
        else:
            text = f"Missing {total}/{total}" if total else "No sessions"
        status_item = self.sceneRoiTable.item(row, self._scene_status_column())
        if status_item is None:
            status_item = qt.QTableWidgetItem("")
            self.sceneRoiTable.setItem(row, self._scene_status_column(), status_item)
        status_item.setText(text)
        tooltip = "\n".join(tooltip_lines)
        for column in (1, self._scene_status_column()):
            item = self.sceneRoiTable.item(row, column)
            if item is not None:
                item.setToolTip(tooltip)
        self._update_scene_role_mapping_visibility()

    def _scene_node_name(self, node_id):
        node = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id else None
        return str(node.GetName() or node_id) if node is not None else str(node_id or "")

    def _scene_mask_choice_name(self, node_id, segment_id):
        node = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id else None
        if node is None:
            return str(node_id or "")
        if segment_id and node.IsA("vtkMRMLSegmentationNode"):
            segment = node.GetSegmentation().GetSegment(str(segment_id))
            if segment is not None:
                return str(segment.GetName() or segment_id)
        return str(node.GetName() or node_id)

    def _remove_scene_roi(self):
        if not hasattr(self, "sceneRoiTable"):
            return
        row = self.sceneRoiTable.currentRow()
        if row < 0:
            row = self.sceneRoiTable.rowCount - 1
        if row >= 0 and self._scene_role_is_analysis_roi(self._scene_role_at_row(row)):
            self.sceneRoiTable.removeRow(row)
            self._resize_scene_timepoint_table()

    def _auto_detect_scene_rois(self):
        discovery = discover_timelapsed_scene_timepoints(self._scene_node_candidates())
        timepoints = list(discovery.timepoints)
        self._populate_scene_roi_rows_from_timepoints(timepoints, reapply_existing=True)
        for timepoint_index in range(self.sceneRegistrationTable.rowCount):
            source_node = self._scene_selected_table_node(timepoint_index, 2, self.sceneRegistrationTable)
            self._apply_scene_detected_roles_for_timepoint(timepoint_index, source_node)

    def _scene_selected_table_node(self, row, column, table):
        selector = table.cellWidget(row, column)
        if selector is None:
            return None
        try:
            return selector.currentNode()
        except Exception:
            return None

    def _populate_scene_roi_rows_from_timepoints(self, timepoints, reapply_existing=False):
        if not hasattr(self, "sceneRoiTable"):
            return
        existing_roles = []
        for row in range(self.sceneRoiTable.rowCount):
            role = self._scene_role_at_row(row)
            if role:
                existing_roles.append(role)
        for role, attr in (
            ("roi1", "full_mask_node_id"),
            ("roi2", "trab_mask_node_id"),
            ("roi3", "cort_mask_node_id"),
        ):
            nodes_by_session = {
                index: getattr(timepoint, attr, "")
                for index, timepoint in enumerate(timepoints)
                if getattr(timepoint, attr, "")
            }
            if nodes_by_session and (role not in existing_roles or reapply_existing):
                if role in existing_roles:
                    role_row = self._scene_role_row_index(role)
                    for timepoint_index, node_id in nodes_by_session.items():
                        column = 2 + timepoint_index
                        segment_attr = attr.replace("_node_id", "_segment_id")
                        segment_id = getattr(timepoints[timepoint_index], segment_attr, "")
                        self._set_scene_mask_row_node(role_row, column, node_id, self.sceneRoiTable, role=role, segment_id=segment_id)
                        self._set_scene_mask_row_policy(role_row, column, "node", self.sceneRoiTable)
                    continue
                self._add_scene_roi(role, nodes_by_session)
                existing_roles.append(role)
        for row in range(self.sceneRoiTable.rowCount):
            self._update_scene_role_status(row)
        self._update_scene_role_mapping_visibility()

    def _add_scene_timepoint(self, timepoint=None):
        if not isinstance(timepoint, TimelapsedSceneTimepoint):
            timepoint = None
        row = self.sceneRegistrationTable.rowCount
        self.sceneRegistrationTable.insertRow(row)
        session_id = timepoint.session_id if timepoint is not None else f"ses-{row + 1}"
        self.sceneRegistrationTable.setItem(row, 0, qt.QTableWidgetItem(session_id))
        image_selector = self._scene_node_selector(["vtkMRMLScalarVolumeNode"])
        self.sceneRegistrationTable.setCellWidget(row, 1, image_selector)
        self.sceneRegistrationTable.setCellWidget(row, 2, self._scene_segmentation_node_selector())
        if timepoint is not None:
            self._set_scene_row_node(row, 1, timepoint.image_node_id, self.sceneRegistrationTable)
            source_node_id = (
                timepoint.seg_mask_node_id
                or timepoint.full_mask_node_id
                or timepoint.trab_mask_node_id
                or timepoint.cort_mask_node_id
                or timepoint.reg_mask_node_id
            )
            self._set_scene_row_node(row, 2, source_node_id, self.sceneRegistrationTable)
        self._refresh_scene_roi_columns()
        if timepoint is not None:
            session_index = self.sceneRegistrationTable.rowCount - 1
            mappings = {
                "registration_roi": (timepoint.reg_mask_node_id, timepoint.reg_mask_segment_id, timepoint.reg_mask_policy),
                "segmentation": (timepoint.seg_mask_node_id, timepoint.seg_mask_segment_id, timepoint.seg_mask_policy),
            }
            for role, (node_id, segment_id, policy) in mappings.items():
                role_row = self._scene_role_row_index(role)
                if role_row >= 0:
                    column = 2 + session_index
                    self._set_scene_mask_row_node(role_row, column, node_id, self.sceneRoiTable, role=role, segment_id=segment_id)
                    self._set_scene_mask_row_policy(role_row, column, policy, self.sceneRoiTable)
                    self._update_scene_role_status(role_row)
            transform_row = self._scene_role_row_index("initial_transform")
            if transform_row >= 0:
                self._set_scene_row_node(transform_row, 2 + session_index, timepoint.transform_node_id, self.sceneRoiTable)
                self._update_scene_role_status(transform_row)
        self._resize_scene_timepoint_table()

    def _set_scene_row_node(self, row, column, node_id, table=None):
        table = table or self.sceneRegistrationTable
        selector = table.cellWidget(row, column)
        if selector is None:
            return
        if not node_id:
            selector.setCurrentNode(None)
            return
        node = slicer.mrmlScene.GetNodeByID(str(node_id)) if selector is not None else None
        if node is not None:
            selector.setCurrentNode(node)

    def _scene_transform_nodes_by_pair(self):
        matches = {}
        scene = slicer.mrmlScene
        for index in range(scene.GetNumberOfNodes()):
            node = scene.GetNthNode(index)
            if node is None or not node.IsA("vtkMRMLTransformNode"):
                continue
            if not self._scene_transform_is_supported_initial_transform(node):
                continue
            storage_path = self._scene_transform_storage_path(node)
            if storage_path is None:
                continue
            parsed = self._scene_transform_pair_and_kind_from_path(storage_path)
            if parsed is None:
                continue
            moving_session, fixed_session, kind = parsed
            matches.setdefault((moving_session, fixed_session), []).append((kind, node))
        return matches

    def _scene_transform_rank_for_registration_reuse(self, kind):
        kind = str(kind or "").lower()
        ranks = {"pairwise": 0}
        return ranks.get(kind, 99)

    def _select_scene_initial_transforms_for_registration_reuse(self):
        transform_row = self._scene_role_row_index("initial_transform")
        if transform_row < 0 or self.sceneRegistrationTable.rowCount < 2:
            return
        transforms_by_pair = self._scene_transform_nodes_by_pair()
        session_ids = []
        for row in range(self.sceneRegistrationTable.rowCount):
            item = self.sceneRegistrationTable.item(row, 0)
            session_ids.append(str(item.text() if item is not None else "").strip())
        for row in range(1, self.sceneRegistrationTable.rowCount):
            moving_session = session_ids[row]
            fixed_session = session_ids[row - 1]
            candidates = transforms_by_pair.get((moving_session, fixed_session), [])
            candidates = [(kind, node) for kind, node in candidates if str(kind).lower() == "pairwise"]
            if not candidates:
                self._set_scene_row_node(transform_row, 2 + row, "", self.sceneRoiTable)
                continue
            _rank, selected_node = min(
                (
                    (self._scene_transform_rank_for_registration_reuse(kind), node)
                    for kind, node in candidates
                ),
                key=lambda item: item[0],
            )
            self._set_scene_row_node(transform_row, 2 + row, selected_node.GetID(), self.sceneRoiTable)
        self._update_scene_role_status(transform_row)

    def _set_scene_mask_row_node(self, row, column, node_id, table=None, role=None, segment_id=""):
        table = table or self.sceneRoiTable
        selector = table.cellWidget(row, column)
        if selector is None:
            return
        if not node_id:
            index = selector.findData("__none__")
            if index >= 0:
                selector.setCurrentIndex(index)
            return
        index = selector.findData(_encode_scene_mask_choice(node_id, segment_id))
        if index >= 0:
            selector.setCurrentIndex(index)
            return
        role_segment_id = self._scene_segment_id_for_node_role(node_id, role)
        if role_segment_id:
            index = selector.findData(_encode_scene_mask_choice(node_id, role_segment_id))
            if index >= 0:
                selector.setCurrentIndex(index)
                return
        for item_index in range(selector.count):
            value_node_id, _segment_id = _decode_scene_mask_choice(selector.itemData(item_index))
            if value_node_id == str(node_id):
                selector.setCurrentIndex(item_index)
                return

    def _scene_segment_id_for_node_role(self, node_id, role):
        node = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id else None
        if node is None or not node.IsA("vtkMRMLSegmentationNode"):
            return ""
        return (
            self._segmentation_segment_id_for_role(node, role)
            or self._segmentation_segment_id_for_role(node, self._scene_default_segment_match_role(role))
            or ""
        )

    def _scene_default_segment_match_role(self, role):
        normalized = self._normalize_scene_role_name(role)
        return {
            "roi1": "full",
            "roi2": "trab",
            "roi3": "cort",
        }.get(normalized, normalized)

    def _set_scene_mask_row_policy(self, row, column, policy, table=None):
        table = table or self.sceneRoiTable
        selector = table.cellWidget(row, column)
        if selector is None:
            return
        normalized = str(policy or "").strip().lower()
        policy_value = ""
        if normalized == "none":
            policy_value = "__none__"
        if policy_value:
            index = selector.findData(policy_value)
            if index >= 0:
                previous = self._syncing_scene_mask_policy
                self._syncing_scene_mask_policy = True
                try:
                    selector.setCurrentIndex(index)
                finally:
                    self._syncing_scene_mask_policy = previous

    def _on_scene_mask_policy_changed(self, selector, column, table=None):
        if self._syncing_scene_mask_policy:
            return
        table = table or self.sceneRoiTable
        if hasattr(self, "sceneRoiTable") and table is self.sceneRoiTable:
            return
        value = str(selector.currentData or "") if selector is not None else ""
        if value != "__none__":
            return
        self._syncing_scene_mask_policy = True
        try:
            for row in range(table.rowCount):
                other = table.cellWidget(row, column)
                if other is None or other is selector:
                    continue
                index = other.findData(value)
                if index >= 0:
                    other.setCurrentIndex(index)
        finally:
            self._syncing_scene_mask_policy = False

    def _remove_scene_timepoint(self):
        row = self.sceneRegistrationTable.currentRow()
        if row < 0:
            row = self.sceneRegistrationTable.rowCount - 1
        if row >= 0:
            self.sceneRegistrationTable.removeRow(row)
            self._refresh_scene_roi_columns()
            self._resize_scene_timepoint_table()

    def _move_scene_timepoint(self, offset):
        row = self.sceneRegistrationTable.currentRow()
        target = row + int(offset)
        if row < 0 or target < 0 or target >= self.sceneRegistrationTable.rowCount:
            return
        timepoints = list(self._scene_timepoints())
        rois = list(self._scene_roi_selections())
        moving = timepoints.pop(row)
        timepoints.insert(target, moving)
        self.sceneRegistrationTable.setRowCount(0)
        self.sceneRoiTable.setRowCount(0)
        for timepoint in timepoints:
            self._add_scene_timepoint(timepoint)
        self._set_scene_roi_selections(rois)
        self.sceneRegistrationTable.selectRow(target)
        self._resize_scene_timepoint_table()

    def _scene_node_candidates(self):
        candidates = []
        scene = slicer.mrmlScene
        for index in range(scene.GetNumberOfNodes()):
            node = scene.GetNthNode(index)
            if node is None:
                continue
            if node.IsA("vtkMRMLTransformNode") and not self._scene_transform_is_supported_initial_transform(node):
                continue
            if not (
                node.IsA("vtkMRMLScalarVolumeNode")
                or node.IsA("vtkMRMLLabelMapVolumeNode")
                or node.IsA("vtkMRMLSegmentationNode")
                or node.IsA("vtkMRMLTransformNode")
            ):
                continue
            attributes = {}
            try:
                names = vtk.vtkStringArray()
                node.GetAttributeNames(names)
                for attribute_index in range(names.GetNumberOfValues()):
                    name = names.GetValue(attribute_index)
                    value = node.GetAttribute(name)
                    if value:
                        attributes[str(name)] = str(value)
            except Exception:
                attributes = {}
            try:
                storage_node = node.GetStorageNode()
                if storage_node is not None and storage_node.GetFileName():
                    attributes["StorageFileName"] = str(storage_node.GetFileName())
            except Exception:
                pass
            candidates.append(
                TimelapsedSceneNodeCandidate(
                    node_id=str(node.GetID()),
                    name=str(node.GetName() or ""),
                    node_class=str(node.GetClassName() if hasattr(node, "GetClassName") else ""),
                    attributes=attributes,
                )
            )
        return candidates

    def _on_discover_scene_timepoints(self):
        self._remove_scene_run_nonlinear_transform_nodes()
        discovery = discover_timelapsed_scene_timepoints(self._scene_node_candidates())
        if not bool(self.sceneAppendDiscoveryCheck.checked):
            self.sceneRegistrationTable.setRowCount(0)
            self.sceneRoiTable.setRowCount(0)
        for timepoint in discovery.timepoints:
            self._add_scene_timepoint(timepoint)
        self._populate_scene_roi_rows_from_timepoints(discovery.timepoints)
        for timepoint_index in range(self.sceneRegistrationTable.rowCount):
            source_node = self._scene_selected_table_node(timepoint_index, 2, self.sceneRegistrationTable)
            self._apply_scene_detected_roles_for_timepoint(timepoint_index, source_node)
        self._select_scene_initial_transforms_for_registration_reuse()
        self._resize_scene_timepoint_table()
        if discovery.subject_id:
            self._scene_subject_id = discovery.subject_id
        if discovery.site:
            self._scene_site = discovery.site
        count = len(discovery.timepoints)
        self.sceneStatusLabel.text = (
            f"Best guess: {discovery.image_count} image node(s), {discovery.mask_count} mask node(s), "
            f"{discovery.matched_mask_count} matched mask role(s); added {count} timepoint row(s)."
        )

    def _scene_selected_node_id(self, row, column, table=None):
        table = table or self.sceneRegistrationTable
        selector = table.cellWidget(row, column)
        node = selector.currentNode() if selector is not None else None
        if node is not None and node.IsA("vtkMRMLTransformNode") and not self._scene_transform_is_supported_initial_transform(node):
            return ""
        return node.GetID() if node is not None else ""

    def _scene_selected_mask_node_id(self, row, column, table=None):
        table = table or self.sceneRoiTable
        selector = table.cellWidget(row, column)
        value = str(selector.currentData or "") if selector is not None else ""
        if value == "__none__":
            return ""
        node_id, _segment_id = _decode_scene_mask_choice(value)
        if slicer.mrmlScene.GetNodeByID(node_id) is None:
            return ""
        return node_id

    def _scene_selected_mask_segment_id(self, row, column, table=None):
        table = table or self.sceneRoiTable
        selector = table.cellWidget(row, column)
        value = str(selector.currentData or "") if selector is not None else ""
        if value == "__none__":
            return ""
        node_id, segment_id = _decode_scene_mask_choice(value)
        if not segment_id or slicer.mrmlScene.GetNodeByID(node_id) is None:
            return ""
        return segment_id

    def _scene_selected_mask_policy(self, row, column, table=None):
        table = table or self.sceneRoiTable
        selector = table.cellWidget(row, column)
        value = str(selector.currentData or "") if selector is not None else ""
        if value == "__none__":
            return "none"
        if not value:
            return "none"
        node_id, _segment_id = _decode_scene_mask_choice(value)
        if slicer.mrmlScene.GetNodeByID(node_id) is None:
            return "none"
        return "node"

    def _scene_requested_mask_roles(self):
        roles = [roi.role for roi in self._scene_roi_selections() if any(policy != "none" for policy in roi.policies)]
        return roles or ["full"]

    def _scene_segmentation_requested(self):
        row = self._scene_role_row_index("segmentation")
        if row < 0:
            return False
        for timepoint_index in range(self.sceneRegistrationTable.rowCount):
            if self._scene_selected_mask_policy(row, 2 + timepoint_index, self.sceneRoiTable) != "none":
                return True
        return False

    def _scene_analysis_compartments(self):
        return self._scene_requested_mask_roles()

    def _scene_settings_override(self):
        settings = self._settings_override(force_analysis_controls=True)
        masks_cfg = dict(settings.get("masks") or {})
        masks_cfg["generate"] = False
        masks_cfg["overwrite"] = False
        masks_cfg["roles"] = self._scene_requested_mask_roles()
        masks_cfg["generate_segmentation"] = False
        if not any(role in masks_cfg["roles"] for role in ("trab", "cort", "trab_roi", "cort_roi")):
            inner_cfg = dict(masks_cfg.get("inner") or {})
            inner_cfg["contour_method"] = "none"
            masks_cfg["inner"] = inner_cfg
        settings["masks"] = masks_cfg

        analysis_cfg = dict(settings.get("analysis") or {})
        analysis_cfg["compartments"] = self._scene_analysis_compartments()
        visualization_cfg = dict(settings.get("visualization") or {})
        label_map = dict(visualization_cfg.get("label_map") or {})
        if not self._scene_segmentation_requested():
            binary_cfg = dict(analysis_cfg.get("binary_reclassification") or {})
            binary_cfg["enabled"] = False
            analysis_cfg["binary_reclassification"] = binary_cfg
            label_map.update({"demineralisation": 0, "quiescent": 0, "mineralisation": 0})
        else:
            label_map.update({"demineralisation": 2, "quiescent": 2, "mineralisation": 2})
        visualization_cfg["label_map"] = label_map
        settings["visualization"] = visualization_cfg
        settings["analysis"] = analysis_cfg
        return settings

    def _scene_timepoints(self):
        timepoints = []
        registration_row = self._scene_role_row_index("registration_roi")
        segmentation_row = self._scene_role_row_index("segmentation")
        transform_row = self._scene_role_row_index("initial_transform")
        for row in range(self.sceneRegistrationTable.rowCount):
            session_item = self.sceneRegistrationTable.item(row, 0)
            role_column = 2 + row
            timepoints.append(
                TimelapsedSceneTimepoint(
                    session_id=(
                        session_item.text()
                        if session_item is not None
                        else ""
                    ),
                    image_node_id=self._scene_selected_node_id(row, 1, self.sceneRegistrationTable),
                    reg_mask_node_id=(
                        self._scene_selected_mask_node_id(registration_row, role_column, self.sceneRoiTable)
                        if registration_row >= 0
                        else ""
                    ),
                    seg_mask_node_id=(
                        self._scene_selected_mask_node_id(segmentation_row, role_column, self.sceneRoiTable)
                        if segmentation_row >= 0
                        else ""
                    ),
                    reg_mask_segment_id=(
                        self._scene_selected_mask_segment_id(registration_row, role_column, self.sceneRoiTable)
                        if registration_row >= 0
                        else ""
                    ),
                    seg_mask_segment_id=(
                        self._scene_selected_mask_segment_id(segmentation_row, role_column, self.sceneRoiTable)
                        if segmentation_row >= 0
                        else ""
                    ),
                    transform_node_id=(
                        self._scene_selected_node_id(transform_row, role_column, self.sceneRoiTable)
                        if transform_row >= 0
                        else ""
                    ),
                    reg_mask_policy=(
                        self._scene_selected_mask_policy(registration_row, role_column, self.sceneRoiTable)
                        if registration_row >= 0
                        else "none"
                    ),
                    seg_mask_policy=(
                        self._scene_selected_mask_policy(segmentation_row, role_column, self.sceneRoiTable)
                        if segmentation_row >= 0
                        else "none"
                    ),
                )
            )
        return timepoints

    def _scene_roi_selections(self):
        rois = []
        if not hasattr(self, "sceneRoiTable"):
            return rois
        timepoint_count = self.sceneRegistrationTable.rowCount
        for row in range(self.sceneRoiTable.rowCount):
            role = self._scene_role_at_row(row)
            if not self._scene_role_is_analysis_roi(role) or not self._scene_role_included(row):
                continue
            node_ids = []
            segment_ids = []
            policies = []
            for timepoint_index in range(timepoint_count):
                column = 2 + timepoint_index
                node_ids.append(self._scene_selected_mask_node_id(row, column, self.sceneRoiTable))
                segment_ids.append(self._scene_selected_mask_segment_id(row, column, self.sceneRoiTable))
                policies.append(self._scene_selected_mask_policy(row, column, self.sceneRoiTable))
            rois.append(
                TimelapsedSceneRoiSelection(
                    role=role,
                    node_ids=tuple(node_ids),
                    segment_ids=tuple(segment_ids),
                    policies=tuple(policies),
                )
            )
        return rois

    def _set_scene_roi_selections(self, rois):
        if not hasattr(self, "sceneRoiTable"):
            return
        self.sceneRoiTable.setRowCount(0)
        self._refresh_scene_roi_columns()
        for roi in rois:
            self._add_scene_roi(roi.role)
            row = self.sceneRoiTable.rowCount - 1
            for timepoint_index in range(self.sceneRegistrationTable.rowCount):
                column = 2 + timepoint_index
                node_id = roi.node_ids[timepoint_index] if timepoint_index < len(roi.node_ids) else ""
                segment_id = roi.segment_ids[timepoint_index] if timepoint_index < len(roi.segment_ids) else ""
                policy = roi.policies[timepoint_index] if timepoint_index < len(roi.policies) else "none"
                self._set_scene_mask_row_node(
                    row,
                    column,
                    node_id,
                    self.sceneRoiTable,
                    role=roi.role,
                    segment_id=segment_id,
                )
                self._set_scene_mask_row_policy(row, column, policy, self.sceneRoiTable)

    def _segmentation_segment_id_for_role(self, segmentation_node, role):
        segmentation = segmentation_node.GetSegmentation()
        role_norm = str(role or "").strip().lower()
        if not role_norm:
            return None
        for index in range(segmentation.GetNumberOfSegments()):
            segment_id = segmentation.GetNthSegmentID(index)
            segment = segmentation.GetSegment(segment_id)
            tag_role = ""
            try:
                tag_role = str(segment.GetTag("HRpQCT.Role") or segment.GetTag("BoneContouring.Role") or "")
            except Exception:
                tag_role = ""
            name = str(segment.GetName() or "")
            if scene_segment_matches_role(name, tag_role, role_norm):
                return segment_id
        return None

    def _export_segmentation_role_to_labelmap(self, segmentation_node, temporary_node, reference_node, role, segment_id=""):
        segmentation = segmentation_node.GetSegmentation()
        selected_segment_id = str(segment_id or "")
        segment = segmentation.GetSegment(selected_segment_id) if selected_segment_id else None
        segment_id = selected_segment_id if segment is not None else self._segmentation_segment_id_for_role(segmentation_node, role)
        if not segment_id and str(role).strip().lower().replace("-", "_") == "regmask":
            segment_id = self._segmentation_segment_id_for_role(segmentation_node, "full")
        if not segment_id:
            raise ValueError(f"Segmentation {segmentation_node.GetName()} does not contain segment role '{role}'.")
        segment_ids = vtk.vtkStringArray()
        segment_ids.InsertNextValue(segment_id)
        logic = slicer.modules.segmentations.logic()
        try:
            return logic.ExportSegmentsToLabelmapNode(
                segmentation_node,
                segment_ids,
                temporary_node,
                reference_node,
            )
        except TypeError:
            return logic.ExportSegmentsToLabelmapNode(
                segmentation_node,
                segment_ids,
                temporary_node,
                reference_node,
                slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
            )

    def _detach_scene_export_transforms(self, *nodes):
        detached = []
        seen = set()
        for node in nodes:
            if node is None:
                continue
            node_id = str(node.GetID() or "")
            if not node_id or node_id in seen:
                continue
            seen.add(node_id)
            if not hasattr(node, "GetTransformNodeID") or not hasattr(node, "SetAndObserveTransformNodeID"):
                continue
            transform_id = node.GetTransformNodeID()
            detached.append((node, transform_id))
            if transform_id:
                node.SetAndObserveTransformNodeID(None)
        return detached

    def _restore_scene_export_transforms(self, detached_transforms):
        for node, transform_id in reversed(list(detached_transforms or [])):
            if node is None or not hasattr(node, "SetAndObserveTransformNodeID"):
                continue
            try:
                node.SetAndObserveTransformNodeID(transform_id)
            except Exception as exc:
                self._show(f"[scene] could not restore display transform on {self._scene_node_name(node.GetID())}: {exc}")

    def _scene_transform_storage_path(self, transform_node):
        if transform_node is None:
            return None
        try:
            storage_node = transform_node.GetStorageNode()
            if storage_node is not None and storage_node.GetFileName():
                path = Path(str(storage_node.GetFileName()))
                if path.exists():
                    return path
        except Exception:
            pass
        return None

    def _remove_scene_run_nonlinear_transform_nodes(self):
        removed = 0
        nodes_to_remove = []
        scene = slicer.mrmlScene
        for index in range(scene.GetNumberOfNodes()):
            node = scene.GetNthNode(index)
            if node is None or not node.IsA("vtkMRMLTransformNode"):
                continue
            storage_path = self._scene_transform_storage_path(node)
            if storage_path is None or storage_path.suffix.lower() != ".h5":
                continue
            storage_text = str(storage_path).replace("\\", "/")
            if "/TimelapsedScene/derivatives/Timelapsed/scene_runs/" not in storage_text:
                continue
            nodes_to_remove.append(node)
        for node in nodes_to_remove:
            try:
                scene.RemoveNode(node)
                removed += 1
            except Exception:
                pass
        if removed:
            self._show(f"[scene] removed {removed} generated non-linear transform node(s) from the scene.")
        return removed

    def _export_scene_node(self, node_id, path, reference_node_id=None, role=None, segment_id=""):
        node = slicer.mrmlScene.GetNodeByID(node_id)
        if node is None:
            raise ValueError(f"Selected scene node is no longer available: {node_id}")
        node_to_save = node
        temporary_node = None
        detached_transforms = []
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            reference_node = (
                slicer.mrmlScene.GetNodeByID(reference_node_id)
                if reference_node_id
                else None
            )
            if node.IsA("vtkMRMLTransformNode"):
                if not self._scene_transform_is_supported_initial_transform(node):
                    raise ValueError(
                        f"Initial transform {node.GetName()} is not a linear .tfm transform. "
                        "Use a Timelapsed .tfm output or a Slicer linear transform."
                    )
                stored_transform_path = self._scene_transform_storage_path(node)
                if stored_transform_path is not None and Path(stored_transform_path).suffix.lower() == ".tfm":
                    shutil.copy2(stored_transform_path, path)
                    return
            detached_transforms = self._detach_scene_export_transforms(node, reference_node)
            if node.IsA("vtkMRMLSegmentationNode"):
                if reference_node is None:
                    raise ValueError(
                        f"Segmentation export for {node.GetName()} requires a timepoint image geometry."
                    )
                temporary_node = slicer.modules.volumes.logic().CreateAndAddLabelVolume(
                    reference_node,
                    f"{node.GetName()}_timelapsed_scene_labelmap",
                )
                if role:
                    ok = self._export_segmentation_role_to_labelmap(
                        node,
                        temporary_node,
                        reference_node,
                        role,
                        segment_id=segment_id,
                    )
                else:
                    ok = slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(
                        node,
                        temporary_node,
                        slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY,
                    )
                if not ok:
                    raise RuntimeError(f"Could not convert {node.GetName()} to a labelmap.")
                node_to_save = temporary_node
            if not slicer.util.saveNode(node_to_save, str(path)):
                raise RuntimeError(f"Could not export {node.GetName()} to {path}")
        finally:
            self._restore_scene_export_transforms(detached_transforms)
            if temporary_node is not None:
                slicer.mrmlScene.RemoveNode(temporary_node)

    def _read_scene_mask_array(self, path):
        return sitk.GetArrayFromImage(sitk.ReadImage(str(path))) > 0

    def _validate_scene_analysis_roi_overlap(self, plan):
        warnings = []
        for roi in plan.rois:
            roi_label = self._scene_display_compartment_name(roi.role)
            for index, (timepoint, roi_path, policy) in enumerate(
                zip(plan.timepoints, roi.paths, roi.policies)
            ):
                if policy == "none" or roi_path is None or timepoint.seg_mask_path is None:
                    continue
                if not Path(roi_path).exists() or not Path(timepoint.seg_mask_path).exists():
                    continue
                roi_arr = self._read_scene_mask_array(roi_path)
                seg_arr = self._read_scene_mask_array(timepoint.seg_mask_path)
                if roi_arr.shape != seg_arr.shape:
                    warnings.append(
                        f"Scene ROI '{roi_label}' in session {timepoint.session_id} has shape "
                        f"{roi_arr.shape}, but the selected segmentation has shape {seg_arr.shape}. "
                        "Check Role Mapping for that timepoint."
                    )
                    continue
                if np.count_nonzero(roi_arr & seg_arr) == 0:
                    roi_vox = int(np.count_nonzero(roi_arr))
                    seg_vox = int(np.count_nonzero(seg_arr))
                    overlap_vox = 0
                    warnings.append(
                        f"Scene ROI '{roi_label}' in session {timepoint.session_id} "
                        "does not overlap the selected segmentation. "
                        f"roi_vox={roi_vox}, seg_vox={seg_vox}, overlap_vox={overlap_vox}. "
                        "Check Role Mapping for that timepoint or choose "
                        "an ROI that contains bone."
                    )
        if warnings:
            message = "\n".join(warnings)
            self._show(message)
            self.sceneStatusLabel.text = "Scene ROI/segmentation overlap warning. Check log."
            slicer.util.warningDisplay(message)

    def _on_run_scene_pipeline(self):
        if not self._require_pipeline_installed():
            return
        results_root = self._path_text(self.sceneResultsRootPath)
        if not results_root:
            results_root = str(self._default_scene_results_root())
            self.sceneResultsRootPath.setCurrentPath(results_root)
        try:
            self.sceneStatusLabel.text = "Preparing scene run..."
            self._show("[timelapsed-slicer] preparing scene run from loaded nodes.")
            plan = build_timelapsed_scene_plan(
                results_root=results_root,
                subject_id=self._scene_subject_id,
                site=self._scene_site,
                timepoints=self._scene_timepoints(),
                rois=self._scene_roi_selections(),
                run_id=datetime.now().strftime("%Y%m%d_%H%M%S_%f"),
            )
            for timepoint in plan.timepoints:
                self._export_scene_node(timepoint.image_node_id, timepoint.image_path)
                for node_id, path, role, segment_id in [
                    (timepoint.reg_mask_node_id, timepoint.reg_mask_path, "regmask", timepoint.reg_mask_segment_id),
                    (timepoint.seg_mask_node_id, timepoint.seg_mask_path, "seg", timepoint.seg_mask_segment_id),
                    (timepoint.transform_node_id, timepoint.transform_path, None, ""),
                ]:
                    if node_id and path is not None:
                        self._export_scene_node(
                            node_id,
                            path,
                            reference_node_id=timepoint.image_node_id,
                            role=role,
                            segment_id=segment_id,
                        )
            for roi in plan.rois:
                for timepoint, node_id, path, segment_id in zip(
                    plan.timepoints,
                    roi.node_ids,
                    roi.paths,
                    roi.segment_ids,
                ):
                    if node_id and path is not None:
                        self._export_scene_node(
                            node_id,
                            path,
                            reference_node_id=timepoint.image_node_id,
                            role=roi.role,
                            segment_id=segment_id,
                        )
            self._validate_scene_analysis_roi_overlap(plan)
            seeded_transforms = self._seed_scene_transform_registry(plan)
            if seeded_transforms:
                self._show(f"[timelapsed-slicer] seeded {seeded_transforms} scene transform(s) for registration reuse.")
            scene_settings = self._scene_settings_override()
            cfg = self.logic.create_override_config(scene_settings, results_root=plan.output_root)
        except Exception as exc:
            self.sceneStatusLabel.text = f"Scene run could not start: {exc}"
            self._show(f"[timelapsed-slicer] scene run could not start: {exc}")
            slicer.util.errorDisplay(str(exc))
            return

        self._is_full_pipeline_run = True
        self._run_skips_mask_generation = True
        self._run_includes_analysis = True
        self._last_scene_plan = plan
        self._set_stage_status("dataset", "done")
        self._set_stage_status("parse", "done")
        for stage in ("registration", "analysis"):
            self._set_stage_status(stage, "pending")
        self._active_stage = "registration"
        self._run(
            timelapsed_scene_run_args(
                plan,
                mode="regular",
                config_path=cfg,
                generate_missing_masks=False,
            )
        )

    def _dataset_root(self):
        p = self._path_text(getattr(self, "inputPath", None))
        root = Path(p) if p else None
        self._set_stage_status("dataset", "done" if root is not None else "pending")
        return root

    def _imported_dataset_root(self):
        root = self._dataset_root()
        plan = getattr(self, "_last_scene_results_plan", None) or getattr(self, "_last_scene_plan", None)
        if root is None and plan is not None:
            output_root = Path(plan.output_root)
            if output_root.name == "Timelapse":
                return output_root
            return output_root / "derivatives" / "Timelapse"
        if root is None:
            return None
        override = self._path_text(getattr(self, "resultsRootPath", None))
        selected = Path(override).expanduser() if override else root
        if selected.name == "Timelapse":
            return selected
        if selected.name == "derivatives":
            return selected / "Timelapse"
        return selected / "derivatives" / "Timelapse"

    def _derivatives_root(self):
        imported = self._imported_dataset_root()
        if imported is None:
            return None
        return imported

    def _show(self, text):
        message = text.rstrip()
        self._update_stage_from_pipeline_output(message)
        if hasattr(self, "logText") and self.logText is not None:
            self.logText.appendPlainText(message)
        else:
            print(message)

    def _update_stage_from_pipeline_output(self, message):
        if not getattr(self, "_is_full_pipeline_run", False):
            return
        text = str(message or "").strip().lower()
        if not text:
            return

        def mark_running(stage):
            order = ["dataset", "parse", "registration", "analysis"]
            if stage not in order:
                return
            for previous in order[: order.index(stage)]:
                if self._stage_states.get(previous) not in {"done", "error"}:
                    self._set_stage_status(previous, "done")
            if self._stage_states.get(stage) not in {"done", "error"}:
                self._set_stage_status(stage, "running")
                self._active_stage = stage

        if "discovered" in text or "scene run from loaded nodes" in text or "import" in text:
            mark_running("parse")
            return

        if (
            "register:" in text
            or "timelapse registration" in text
            or "stackcorrect:" in text
            or "stack correction for" in text
            or "transform:" in text
            or "[apply]" in text
        ):
            mark_running("registration")
            return

        if "[analysis]" in text or "analyse:" in text:
            mark_running("analysis")

    def _set_scene_stage_message(self, text):
        if not getattr(self, "_last_scene_plan", None):
            return
        self._set_label_text_safe(getattr(self, "sceneStatusLabel", None), text)

    def _set_user_message(self, level, title, body):
        palette = {
            "info": ("#eaf2ff", "#7ea6f7"),
            "warn": ("#fff6db", "#f0c36d"),
            "error": ("#ffeceb", "#e68a87"),
            "success": ("#eaf8ea", "#8aca8a"),
        }
        bg, border = palette.get(level, palette["info"])
        label = getattr(self, "userMessageLabel", None)
        self._set_widget_style_safe(
            label,
            f"QLabel {{ background:{bg}; border:1px solid {border}; padding:8px; border-radius:4px; }}"
        )
        self._set_label_text_safe(label, f"<b>{title}</b><br>{body}")
        self._set_widget_visible_safe(label, True)

    def _clear_user_message(self):
        label = getattr(self, "userMessageLabel", None)
        self._set_widget_visible_safe(label, False)
        self._set_label_text_safe(label, "")

    def _set_stage_status(self, stage_key, status):
        if stage_key not in self.stageLabels:
            return
        self._stage_states[stage_key] = status
        style = {
            "pending": ("●", "#888888", "Pending"),
            "running": ("●", "#2f7ed8", "Running"),
            "done": ("●", "#2d9a4b", "Done"),
            "error": ("●", "#c73a3a", "Needs attention"),
        }
        dot, color, label = style.get(status, style["pending"])
        self.stageLabels[stage_key].setText(f"<span style='color:{color}; font-weight:700'>{dot}</span> {label}")
        if hasattr(self, "sceneStageItems") and stage_key in self.sceneStageItems:
            self.sceneStageItems[stage_key].setText(f"<span style='color:{color}; font-weight:700'>{dot}</span> {label}")
        self._update_progress_ui()

    def _update_progress_ui(self):
        order = ["dataset", "parse", "registration", "analysis"]
        done = sum(1 for k in order if self._stage_states.get(k) == "done")
        if hasattr(self, "progressBar") and self.progressBar is not None:
            self.progressBar.value = int(done)

        running = [k for k in order if self._stage_states.get(k) == "running"]
        errors = [k for k in order if self._stage_states.get(k) == "error"]
        label_map = {
            "dataset": "dataset selection",
            "parse": "parse",
            "registration": "registration",
            "analysis": "analysis",
        }
        if errors:
            text = f"Current step: blocked at {label_map.get(errors[0], errors[0])}"
        elif running:
            text = f"Current step: {label_map.get(running[0], running[0])}"
        elif done == len(order):
            text = "Current step: complete"
        else:
            pending = [k for k in order if self._stage_states.get(k) != "done"]
            text = (
                f"Current step: waiting for {label_map.get(pending[0], pending[0])}"
                if pending
                else "Current step: idle"
            )
        if hasattr(self, "currentStepLabel") and self.currentStepLabel is not None:
            self.currentStepLabel.text = text
        self._set_scene_stage_message(text)

    def _connect_path_changed(self, widget, callback):
        for signal_name in ("currentPathChanged", "pathChanged"):
            signal = getattr(widget, signal_name, None)
            if signal is None:
                continue
            try:
                signal.connect(callback)
                return
            except Exception:
                pass
        for signature in ("currentPathChanged(QString)", "pathChanged(QString)"):
            try:
                widget.connect(signature, callback)
                return
            except Exception:
                pass

    def _path_text(self, widget):
        try:
            return str(widget.currentPath or "").strip()
        except Exception:
            return ""

    def _path_exists(self, path):
        try:
            return path is not None and Path(path).exists()
        except Exception:
            return False

    def _record_image_exists(self, record):
        return self._path_exists(getattr(record, "image_path", None))

    def _imported_stack_masks_complete(self, record):
        mask_paths = getattr(record, "mask_paths", {}) or {}
        required_roles = ("full", "trab", "cort")
        if not all(self._path_exists(mask_paths.get(role)) for role in required_roles):
            return False
        seg_path = getattr(record, "seg_path", None)
        return self._path_exists(seg_path)

    def _fused_session_complete(self, record):
        if not self._record_image_exists(record):
            return False
        metadata_path = getattr(record, "metadata_path", None)
        if not self._path_exists(metadata_path):
            return False
        mask_paths = getattr(record, "mask_paths", {}) or {}
        return self._path_exists(mask_paths.get("full"))

    def _analysis_outputs_complete(self, imported, fused_records):
        groups = {}
        for record in fused_records:
            if not self._fused_session_complete(record):
                continue
            groups.setdefault(
                (str(getattr(record, "subject_id", "")), str(getattr(record, "site", "radius"))),
                [],
            ).append(record)

        processable_groups = [
            (subject_id, site)
            for (subject_id, site), records in groups.items()
            if subject_id and len(records) >= 2
        ]
        if not processable_groups:
            return False

        try:
            from timelapsedhrpqct.dataset.derivative_paths import (
                pairwise_remodelling_csv_path,
                trajectory_metrics_csv_path,
            )
        except Exception as exc:
            self._show(f"[progress] could not import analysis path helpers: {exc}")
            return False

        for subject_id, site in processable_groups:
            if not pairwise_remodelling_csv_path(imported, subject_id, site).exists():
                return False
            if not trajectory_metrics_csv_path(imported, subject_id, site).exists():
                return False
        return True

    def _infer_stage_statuses_from_artifacts(self):
        statuses = {
            "parse": "pending",
            "registration": "pending",
            "analysis": "pending",
        }
        imported = self._imported_dataset_root()
        if imported is None or not imported.exists():
            return statuses

        try:
            from timelapsedhrpqct.dataset.artifacts import (
                iter_filled_session_records,
                iter_fused_session_records,
                iter_imported_stack_records,
            )
        except Exception as exc:
            self._show(f"[progress] artifact status lookup unavailable: {exc}")
            return statuses

        try:
            imported_records = list(iter_imported_stack_records(imported))
            fused_records = list(iter_fused_session_records(imported))
            filled_records = list(iter_filled_session_records(imported))
        except Exception as exc:
            self._show(f"[progress] artifact status lookup failed: {exc}")
            return statuses

        existing_imported = [record for record in imported_records if self._record_image_exists(record)]
        if existing_imported:
            statuses["parse"] = "done"

        existing_fused = [record for record in fused_records if self._fused_session_complete(record)]
        existing_filled = [
            record for record in filled_records
            if self._record_image_exists(record)
            and self._path_exists(getattr(record, "full_mask_path", None))
            and self._path_exists(getattr(record, "metadata_path", None))
        ]
        if existing_fused or existing_filled:
            statuses["registration"] = "done"

        if self._analysis_outputs_complete(imported, fused_records):
            statuses["analysis"] = "done"

        self._show(
            "[progress] artifact scan: "
            f"imported={len(existing_imported)}/{len(imported_records)}, "
            f"fused={len(existing_fused)}/{len(fused_records)}, "
            f"filled={len(existing_filled)}/{len(filled_records)}, "
            f"analysis={'done' if statuses['analysis'] == 'done' else 'pending'}"
        )
        return statuses

    def _reset_progress_for_dataset_root(self):
        dataset_text = self._path_text(self.inputPath)
        results_text = self._path_text(self.resultsRootPath) if hasattr(self, "resultsRootPath") else ""
        self._last_dataset_root_text = dataset_text
        self._last_results_root_text = results_text

        self._last_parsed_sessions = []
        self._parsed_baseline_rows = []
        self._last_parse_mode_used = None
        self._manual_parse_active = False
        self._patient_keys = []
        self._remodelling_comparison_items = []
        self._interactive_preview_cache = {}
        self._last_scene_results_plan = None
        self._latest_series_summary = None
        self._latest_study_summary_rows = []
        self._series_summary_pair_checks = {}
        self.parseTable.setRowCount(0)
        self.parseSummaryLabel.text = "Parse summary: not run"
        self.parseSummaryLabel.styleSheet = ""
        self._rebuild_series_summary_pair_selector([])
        self._set_series_summary_labels(None)
        self._refresh_processing_subjects()
        self._refresh_patient_list()

        self._set_stage_status("dataset", "done" if dataset_text else "pending")
        artifact_statuses = self._infer_stage_statuses_from_artifacts()
        for stage in ("parse", "registration", "analysis"):
            self._set_stage_status(stage, artifact_statuses.get(stage, "pending"))
        self._update_progress_ui()

    def _on_dataset_or_results_root_changed(self, *_args):
        dataset_text = self._path_text(self.inputPath)
        results_text = self._path_text(self.resultsRootPath) if hasattr(self, "resultsRootPath") else ""
        if (
            dataset_text == self._last_dataset_root_text
            and results_text == self._last_results_root_text
        ):
            self._set_stage_status("dataset", "done" if dataset_text else "pending")
            return
        self._reset_progress_for_dataset_root()

    def _selected_config_profile(self):
        combo = getattr(self, "studyProfileCombo", None)
        if not self._qt_object_alive(combo):
            return "standard"
        try:
            data = self._combo_current_data_safe(combo)
            return str(data or "standard")
        except (RuntimeError, ValueError):
            return "standard"

    def _selected_profile_is_custom(self):
        return self._selected_config_profile() == "__custom__"

    def _available_config_profiles(self):
        try:
            from timelapsedhrpqct.config.profiles import list_config_profiles

            profiles = list_config_profiles()
        except Exception:
            profiles = [profile for profile in PROFILE_DISPLAY_ORDER if profile != "xct1-standard"]
        known = [profile for profile in PROFILE_DISPLAY_ORDER if profile in profiles]
        extra = sorted(profile for profile in profiles if profile not in PROFILE_DISPLAY_ORDER)
        return known + extra

    def _populate_study_profiles(self, combo=None):
        combo = combo or self.studyProfileCombo
        if not self._qt_object_alive(combo):
            return
        combo.clear()
        for profile in self._available_config_profiles():
            combo.addItem(profile, profile)
        combo.addItem("Custom", "__custom__")

    def _profile_cli_args(self):
        if self._selected_profile_is_custom():
            return []
        profile = self._selected_config_profile()
        return ["--profile", profile] if profile else []

    def _profile_enables_multistack(self, profile=None):
        profile = str(profile if profile is not None else self._selected_config_profile())
        if not profile:
            return False
        try:
            from timelapsedhrpqct.config.loader import load_config

            cfg_obj = load_config(None, profile=profile)
            ms_cfg = getattr(cfg_obj, "multistack_correction", None)
            return bool(getattr(ms_cfg, "enabled", False))
        except Exception:
            return profile in {"multistack", "ped-fx"}

    def _selected_profile_enables_multistack(self):
        return self._profile_enables_multistack()

    def _selected_run_mode(self, sessions=None):
        if self._selected_profile_enables_multistack():
            return "multistack"
        if bool(getattr(self, "useMultistackCheck", None) and self.useMultistackCheck.checked):
            return "multistack"
        return self._auto_mode_from_sessions(sessions)

    def _selected_profile_multistack_metric(self):
        profile = self._selected_config_profile()
        if not profile:
            return None
        try:
            from timelapsedhrpqct.config.loader import load_config

            cfg_obj = load_config(None, profile=profile)
            ms_cfg = getattr(cfg_obj, "multistack_correction", None)
            metric = str(getattr(ms_cfg, "metric", "") or "").strip()
            return metric or None
        except Exception:
            return None

    def _cohort_export_profile(self):
        profile = self._selected_config_profile()
        return str(profile or "").strip()

    def _selected_profile_config_dict(self):
        if self._selected_profile_is_custom():
            return {}
        try:
            from dataclasses import asdict
            from timelapsedhrpqct.config.loader import load_config

            return asdict(load_config(None, profile=self._selected_config_profile()))
        except Exception:
            return {}

    def _apply_profile_analysis_controls(self, profile):
        selected_profile = str(profile or "").strip()
        if not selected_profile or selected_profile == "__custom__":
            return False
        try:
            from dataclasses import asdict
            from timelapsedhrpqct.config.loader import load_config

            cfg = asdict(load_config(None, profile=selected_profile))
            self._apply_analysis_config_to_controls(cfg.get("analysis") or {})
            return True
        except Exception as exc:
            self._show(f"[settings] could not apply analysis controls for profile {selected_profile}: {exc}")
            return False

    def _apply_config_dict_to_controls(self, cfg, *, source_label):
        seg_cfg = ((cfg.get("masks") or {}).get("segmentation") or {})
        outer_cfg = ((cfg.get("masks") or {}).get("outer") or {})
        inner_cfg = ((cfg.get("masks") or {}).get("inner") or {})
        method = str(seg_cfg.get("method", self.maskMethod.currentText) or self.maskMethod.currentText)
        if method == "global":
            method = "seg_gauss"
        mask_idx = self.maskMethod.findText(method)
        if mask_idx < 0 and method == "seg_gauss":
            mask_idx = self.maskMethod.findText("seg_gauss")

        adaptive_low = float(seg_cfg.get("adaptive_low_threshold", 100.0))
        adaptive_high = float(seg_cfg.get("adaptive_high_threshold", 300.0))
        seg_gauss_threshold = float(seg_cfg.get("seg_gauss_threshold", seg_cfg.get("trab_threshold", 320.0)))
        seg_gauss_cort_threshold = float(seg_cfg.get("cort_threshold", 450.0))
        self._lh_cort_support_threshold = seg_gauss_cort_threshold
        self.maskContourSupportThreshold.value = float(outer_cfg.get("periosteal_threshold", seg_gauss_threshold))
        self._seg_gauss_sigma = float(seg_cfg.get("gaussian_sigma", seg_cfg.get("seg_gauss_sigma", 0.8)))
        self.maskSigma.value = self._seg_gauss_sigma
        self._contour_gaussian_sigma = float(outer_cfg.get("gaussian_sigma", self._contour_gaussian_sigma))
        self.maskContourSigma.value = self._contour_gaussian_sigma
        laplace_hamming_min_size = float(seg_cfg.get("laplace_hamming_min_size_voxels", 70.0))
        self.maskLaplaceThreshold.value = float(seg_cfg.get("laplace_hamming_threshold", 15564.0))
        self.maskLaplaceLowPass.value = float(seg_cfg.get("laplace_hamming_low_pass_cutoff", 0.3))
        self.maskLaplaceHighPass.value = float(seg_cfg.get("laplace_hamming_high_pass_cutoff", 0.0))
        self.maskLaplaceEpsilon.value = float(seg_cfg.get("laplace_hamming_epsilon", 0.45))
        self._mask_method_defaults = {
            "adaptive": (adaptive_low, adaptive_high),
            "seg_gauss": (seg_gauss_threshold, seg_gauss_cort_threshold),
            "laplace_hamming": (seg_gauss_threshold, laplace_hamming_min_size),
        }
        if mask_idx >= 0:
            self.maskMethod.setCurrentIndex(mask_idx)
        periosteal_method = str(outer_cfg.get("contour_method", "standard") or "standard")
        periosteal_idx = self.maskPeriostealContour.findData(periosteal_method)
        if periosteal_idx < 0:
            periosteal_idx = self.maskPeriostealContour.findText(periosteal_method)
        if periosteal_idx >= 0:
            self.maskPeriostealContour.setCurrentIndex(periosteal_idx)
        endosteal_method = str(inner_cfg.get("contour_method", "standard") or "standard")
        endosteal_idx = self.maskEndostealContour.findData(endosteal_method)
        if endosteal_idx < 0:
            endosteal_idx = self.maskEndostealContour.findText(endosteal_method)
        if endosteal_idx >= 0:
            self.maskEndostealContour.setCurrentIndex(endosteal_idx)
        self.maskEndostealThreshold.value = float(inner_cfg.get("endosteal_threshold", seg_gauss_cort_threshold))
        self.maskEndostealKernel.value = int(inner_cfg.get("endosteal_kernelsize", int(self.maskEndostealKernel.value)))
        self.maskOuterKernel.value = int(outer_cfg.get("periosteal_kernelsize", int(self.maskOuterKernel.value)))
        self.maskOuterOpen.value = int(outer_cfg.get("periosteal_open_radius", int(self.maskOuterOpen.value)))
        self.maskGeodesicThreshold.value = float(outer_cfg.get("geodesic_bone_threshold", 250.0))
        self.maskGeodesicFillHoles.checked = bool(
            outer_cfg.get("fill_holes", outer_cfg.get("geodesic_fill_holes", True))
        )
        self.maskAlignedContourSupport.checked = bool(
            seg_cfg.get("use_segmentation_aligned_contour_support", False)
        )
        self._on_mask_method_changed(self.maskMethod.currentText)

        tl_cfg = cfg.get("timelapsed_registration") or {}
        ms_cfg = cfg.get("multistack_correction") or {}
        self.tlSampling.value = float(tl_cfg.get("sampling_percentage", float(self.tlSampling.value)))
        self.tlRes.value = int(tl_cfg.get("number_of_resolutions", int(self.tlRes.value)))
        self.tlIter.value = int(tl_cfg.get("number_of_iterations", int(self.tlIter.value)))
        self.msSampling.value = float(ms_cfg.get("sampling_percentage", float(self.msSampling.value)))
        self.msRes.value = int(ms_cfg.get("number_of_resolutions", int(self.msRes.value)))
        self.msIter.value = int(ms_cfg.get("number_of_iterations", int(self.msIter.value)))
        self.msOverlapBuffer.value = int(
            ms_cfg.get("overlap_crop_buffer_voxels", int(self.msOverlapBuffer.value))
        )
        init_vox = ms_cfg.get("initial_translation_voxels", [0.0, 0.0, -20.0])
        if isinstance(init_vox, (list, tuple)) and len(init_vox) >= 3:
            self.msInitTx.value = float(init_vox[0])
            self.msInitTy.value = float(init_vox[1])
            self.msInitTz.value = float(init_vox[2])

        self._apply_analysis_config_to_controls(cfg.get("analysis") or {})
        self._on_mask_method_changed(self.maskMethod.currentText)
        self._on_periosteal_contour_method_changed()
        self._sync_scene_profile_from_batch_profile()
        if source_label and hasattr(self, "userMessageLabel"):
            self._set_user_message("info", "Profile applied", source_label)

    def _apply_analysis_config_to_controls(self, analysis_cfg):
        analysis_cfg = dict(analysis_cfg or {})
        self._analysis_method = self._analysis_method_from_config(analysis_cfg)
        self._set_combo_current_data_safe(self.analysisMethodCombo, self._analysis_method)
        change_region_cfg = analysis_cfg.get("change_region") or {}
        binary_cfg = analysis_cfg.get("binary_reclassification") or {}
        if str(analysis_cfg.get("method", "") or "") == "auto" or change_region_cfg or binary_cfg:
            restrict_bone = (
                str(change_region_cfg.get("source", "common_mask")).strip().lower()
                in {"bone_union", "segmentation_union"}
            )
            binary_enabled = bool(binary_cfg.get("enabled", False))
        else:
            restrict_bone = self._analysis_method == "grayscale_marrow_mask"
            binary_enabled = self._analysis_method == "grayscale_and_binary"
        self._set_checkbox_checked_safe(self.analysisRestrictBoneSupportCheck, restrict_bone)
        self._set_checkbox_checked_safe(self.analysisBinaryReclassificationCheck, binary_enabled)
        self._set_combo_current_data_safe(self.analysisPairModeCombo, str(analysis_cfg.get("pair_mode", "adjacent")))
        thresholds = analysis_cfg.get("thresholds") or [float(self.analysisThreshold.value)]
        clusters = analysis_cfg.get("cluster_sizes") or [int(self.analysisCluster.value)]
        if thresholds:
            self._set_analysis_threshold_value(float(thresholds[0]), queue_update=False, force=True)
        if clusters:
            self._set_analysis_cluster_value(int(clusters[0]), queue_update=False, force=True)
        self._set_checkbox_checked_safe(
            self.analysisGaussianFilterCheck,
            analysis_cfg.get("gaussian_filter", bool(self.analysisGaussianFilterCheck.checked)),
        )
        self._set_widget_value_safe(
            self.analysisGaussianSigma,
            float(analysis_cfg.get("gaussian_sigma", float(self.analysisGaussianSigma.value))),
        )
        self._set_widget_value_safe(
            self.analysisFullMaskDilation,
            int(analysis_cfg.get("full_mask_dilation_voxels", int(self.analysisFullMaskDilation.value))),
        )
        self._set_widget_value_safe(
            self.analysisMarrowMaskDilation,
            int(
                change_region_cfg.get(
                    "dilation_voxels",
                    analysis_cfg.get("marrow_mask_dilation_voxels", int(self.analysisMarrowMaskDilation.value)),
                )
            ),
        )
        self._set_widget_value_safe(
            self.analysisMarrowMaskErosion,
            int(
                change_region_cfg.get(
                    "erosion_voxels",
                    analysis_cfg.get("marrow_mask_erosion_voxels", int(self.analysisMarrowMaskErosion.value)),
                )
            ),
        )
        self._analysis_erosion_voxels = int(
            ((analysis_cfg.get("valid_region") or {}).get("erosion_voxels", self._analysis_erosion_voxels))
        )
        self._on_analysis_option_changed()

    def _on_apply_study_profile(self, *_args, profile=None):
        selected_profile = str(profile if profile is not None else self._selected_config_profile())
        if selected_profile == "__custom__":
            self._update_batch_analysis_options_visibility()
            self._set_user_message(
                "info",
                "Custom analysis settings",
                "Batch runs will use the analysis options shown below instead of a bundled profile preset.",
            )
            return
        applied = False
        try:
            from dataclasses import asdict
            from timelapsedhrpqct.config.loader import load_config

            try:
                cfg_obj = load_config(None, profile=selected_profile)
            except TypeError as exc:
                if "profile" not in str(exc):
                    raise
                _ok, detail = self.logic.pipeline_status()
                raise RuntimeError(
                    "The installed timelapsed-hrpqct package does not support built-in profiles yet. "
                    f"Install/update to timelapsed-hrpqct >= {MIN_PIPELINE_VERSION}. {detail}"
                ) from exc
            cfg = asdict(cfg_obj)
            self._suppress_interactive_preview_updates = True
            self.useMultistackCheck.checked = self._profile_enables_multistack(selected_profile)
            self._apply_config_dict_to_controls(
                cfg,
                source_label=f"Using built-in profile <b>{selected_profile}</b> for new runs and analysis reruns.",
            )
            applied = True
        except Exception as exc:
            slicer.util.warningDisplay(f"Could not apply profile:\n{exc}")
        finally:
            self._suppress_interactive_preview_updates = False
            self._update_batch_analysis_options_visibility()
        if applied:
            self._on_apply_interactive_remodelling()

    def _load_defaults_from_pipeline_config(self):
        if not self.logic.is_pipeline_available():
            return
        try:
            import yaml

            with open(self.logic.default_config_path(), "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f) or {}

            self._apply_config_dict_to_controls(cfg, source_label="Loaded defaults from timelapsed-hrpqct.")
        except Exception as exc:
            self._show(f"[settings] could not load defaults from pipeline config: {exc}")

    def _current_analysis_method(self):
        restrict_bone = bool(getattr(self, "analysisRestrictBoneSupportCheck", None) and self.analysisRestrictBoneSupportCheck.checked)
        enforce_binary = bool(getattr(self, "analysisBinaryReclassificationCheck", None) and self.analysisBinaryReclassificationCheck.checked)
        if enforce_binary:
            return "grayscale_and_binary"
        if restrict_bone:
            return "grayscale_marrow_mask"
        return "grayscale_delta_only"

    def _legacy_analysis_combo_method(self):
        data = self.analysisMethodCombo.currentData
        if data:
            return str(data)
        return "grayscale_and_binary"

    def _analysis_method_from_config(self, analysis_cfg):
        method = str((analysis_cfg or {}).get("method", "") or "").strip()
        if method and method != "auto":
            return method
        change_region = (analysis_cfg or {}).get("change_region") or {}
        binary = (analysis_cfg or {}).get("binary_reclassification") or {}
        if bool(binary.get("enabled", False)):
            return "grayscale_and_binary"
        if str(change_region.get("source", "common_mask")).strip().lower() in {"bone_union", "segmentation_union"}:
            return "grayscale_marrow_mask"
        return "grayscale_delta_only"

    def _installed_core_supports_explicit_analysis_config(self):
        try:
            from dataclasses import fields
            from timelapsedhrpqct.config.models import AnalysisConfig

            names = {field.name for field in fields(AnalysisConfig)}
        except Exception:
            return False
        return {"change_detection", "change_region", "binary_reclassification"}.issubset(names)

    def _analysis_config_from_controls(self, pair_mode):
        use_bone_union = bool(self.analysisRestrictBoneSupportCheck.checked)
        enforce_binary = bool(self.analysisBinaryReclassificationCheck.checked)
        base = {
            "method": "auto",
            "pair_mode": pair_mode,
            "compartments": ["full", "trab", "cort"],
            "thresholds": [float(self.analysisThreshold.value)],
            "cluster_sizes": [int(self.analysisCluster.value)],
            "use_filled_images": False,
            "gaussian_filter": bool(self.analysisGaussianFilterCheck.checked),
            "gaussian_sigma": float(self.analysisGaussianSigma.value),
            "full_mask_dilation_voxels": int(self.analysisFullMaskDilation.value),
        }
        if self._installed_core_supports_explicit_analysis_config():
            base.update(
                {
                    "change_detection": "grayscale_delta",
                    "change_region": {
                        "source": "bone_union" if use_bone_union else "common_mask",
                        "dilation_voxels": int(self.analysisMarrowMaskDilation.value) if use_bone_union else 0,
                        "erosion_voxels": int(self.analysisMarrowMaskErosion.value) if use_bone_union else 0,
                    },
                    "binary_reclassification": {
                        "enabled": bool(enforce_binary),
                    },
                }
            )
        else:
            # Older released cores warn on the explicit nested analysis keys. Keep legacy keys only.
            base["method"] = self._current_analysis_method()
            base["marrow_mask_dilation_voxels"] = int(self.analysisMarrowMaskDilation.value) if use_bone_union else 0
            base["marrow_mask_erosion_voxels"] = 0
        return base

    def _on_analysis_method_changed(self, *_args):
        method = self._legacy_analysis_combo_method()
        self.analysisRestrictBoneSupportCheck.checked = method == "grayscale_marrow_mask"
        self.analysisBinaryReclassificationCheck.checked = method == "grayscale_and_binary"
        self._on_analysis_option_changed()

    def _on_analysis_option_changed(self, *_args):
        use_bone_union = bool(self.analysisRestrictBoneSupportCheck.checked)
        self.analysisMarrowMaskDilation.enabled = use_bone_union
        self.analysisMarrowMaskErosion.enabled = use_bone_union
        self._on_interactive_preview_control_changed()

    def _on_mask_method_changed(self, method_name):
        method = str(method_name).strip().lower()
        if method not in self._mask_method_defaults:
            return
        is_lh = method == "laplace_hamming"
        for label, widget in [
            (self.maskLaplaceThresholdLabel, self.maskLaplaceThreshold),
            (self.maskLaplaceLowPassLabel, self.maskLaplaceLowPass),
            (self.maskLaplaceHighPassLabel, self.maskLaplaceHighPass),
            (self.maskLaplaceEpsilonLabel, self.maskLaplaceEpsilon),
        ]:
            label.visible = is_lh
            widget.visible = is_lh
        self.maskSigmaLabel.visible = method == "seg_gauss"
        self.maskSigma.visible = method == "seg_gauss"
        self.maskAlignedContourSupportLabel.visible = method in {"adaptive", "laplace_hamming"}
        self.maskAlignedContourSupport.visible = method in {"adaptive", "laplace_hamming"}
        if method == "laplace_hamming":
            self.maskLowLabel.visible = False
            self.maskLow.visible = False
            self.maskHighLabel.text = "Min component voxels"
            self.maskHighLabel.toolTip = "Minimum connected component size retained by Laplace-Hamming segmentation."
            self.maskHigh.minimum = 0.0
            self.maskHigh.maximum = 1000000.0
            self.maskHigh.decimals = 0
            self.maskHigh.singleStep = 1.0
        elif method == "seg_gauss":
            self.maskLowLabel.visible = True
            self.maskLow.visible = True
            self.maskLowLabel.text = "Trab threshold"
            self.maskLowLabel.toolTip = "Trabecular density threshold used after Gaussian smoothing for standard segmentation."
            self.maskHighLabel.text = "Cort threshold"
            self.maskHighLabel.toolTip = "Cortical density threshold used after Gaussian smoothing for standard segmentation."
            self.maskLow.minimum = 0.0
            self.maskLow.maximum = 5000.0
            self.maskLow.decimals = 1
            self.maskLow.singleStep = 5.0
            self.maskHigh.minimum = 0.0
            self.maskHigh.maximum = 5000.0
            self.maskHigh.decimals = 1
            self.maskHigh.singleStep = 5.0
        else:
            self.maskLowLabel.visible = True
            self.maskLow.visible = True
            self.maskLowLabel.text = "Adaptive low threshold"
            self.maskLowLabel.toolTip = "Lower adaptive segmentation threshold."
            self.maskHighLabel.text = "Adaptive high threshold"
            self.maskHighLabel.toolTip = "Upper adaptive segmentation threshold."
            self.maskLow.minimum = -1000.0
            self.maskLow.maximum = 5000.0
            self.maskLow.decimals = 1
            self.maskLow.singleStep = 5.0
            self.maskHigh.minimum = -1000.0
            self.maskHigh.maximum = 5000.0
            self.maskHigh.decimals = 1
            self.maskHigh.singleStep = 5.0
        low, high = self._mask_method_defaults[method]
        self.maskLow.value = float(low)
        self.maskHigh.value = float(high)
        self._on_periosteal_contour_method_changed()

    def _on_periosteal_contour_method_changed(self, *_args):
        method = str(self.maskPeriostealContour.currentData or self.maskPeriostealContour.currentText or "standard")
        is_geodesic = method == "geodesic_fracture"
        endosteal_method = str(self.maskEndostealContour.currentData or self.maskEndostealContour.currentText or "standard")
        any_standard_contour = (not is_geodesic) or endosteal_method != "none"
        contour_support_visible = any_standard_contour
        for label, widget in [
            (self.maskOuterKernelLabel, self.maskOuterKernel),
            (self.maskOuterOpenLabel, self.maskOuterOpen),
        ]:
            label.visible = not is_geodesic
            widget.visible = not is_geodesic
        self.maskGeodesicThresholdLabel.visible = is_geodesic
        self.maskGeodesicThreshold.visible = is_geodesic
        self.maskContourSupportThresholdLabel.visible = contour_support_visible
        self.maskContourSupportThreshold.visible = contour_support_visible
        self.maskContourSigmaLabel.visible = contour_support_visible
        self.maskContourSigma.visible = contour_support_visible
        self.maskEndostealThresholdLabel.visible = endosteal_method != "none"
        self.maskEndostealThreshold.visible = endosteal_method != "none"
        self.maskEndostealKernelLabel.visible = endosteal_method != "none"
        self.maskEndostealKernel.visible = endosteal_method != "none"
        self.maskGeodesicFillHolesLabel.visible = True
        self.maskGeodesicFillHoles.visible = True

    def _queue_interactive_preview_update(self):
        if self.logic.is_running():
            return
        if not self.remodellingAutoUpdateCheck.checked:
            self._mark_analysis_settings_dirty()
            return
        self._interactivePreviewTimer.start()

    def _mark_analysis_settings_dirty(self):
        label = getattr(self, "analysisStatusLabel", None)
        self._set_label_text_safe(label, "Settings changed - click Apply settings")
        self._set_widget_style_safe(label, "color: #996600;")

    def _set_analysis_threshold_value(self, value, *, from_slider=False, queue_update=False, force=False):
        clamped = max(0, min(1000, int(round(float(value) / 5.0) * 5)))
        if self._updating_analysis_controls and not force:
            return
        self._updating_analysis_controls = True
        try:
            if int(self.analysisThresholdSlider.value) != clamped:
                self._set_widget_value_safe(self.analysisThresholdSlider, clamped)
            if float(self.analysisThreshold.value) != float(clamped):
                self._set_widget_value_safe(self.analysisThreshold, float(clamped))
        finally:
            self._updating_analysis_controls = False
        if queue_update:
            self._queue_interactive_preview_update()

    def _set_analysis_cluster_value(self, value, *, from_slider=False, queue_update=False, force=False):
        clamped = max(0, min(30, int(round(float(value)))))
        if self._updating_analysis_controls and not force:
            return
        self._updating_analysis_controls = True
        try:
            if int(self.analysisClusterSlider.value) != clamped:
                self._set_widget_value_safe(self.analysisClusterSlider, clamped)
            if int(self.analysisCluster.value) != clamped:
                self._set_widget_value_safe(self.analysisCluster, clamped)
        finally:
            self._updating_analysis_controls = False
        if queue_update:
            self._queue_interactive_preview_update()

    def _settings_override(self, multistack_enabled=None, force_analysis_controls=False):
        label_map = {
            "resorption": 1,
            "demineralisation": 2,
            "quiescent": 2,
            "formation": 3,
            "mineralisation": 2,
        }

        if self.tlSampling.value > 0.01:
            self._show("[warning] Timelapse sampling > 0.01 can be slow or unstable on some datasets.")
        if self.msSampling.value > 0.01:
            self._show("[warning] Multistack sampling > 0.01 can be slow or unstable on some datasets.")

        pair_mode = str(self.analysisPairModeCombo.currentData or "adjacent")
        if pair_mode not in {"adjacent", "baseline", "all_pairs"}:
            pair_mode = "adjacent"
        if multistack_enabled is None:
            multistack_enabled = self._selected_profile_enables_multistack() or bool(
                getattr(self, "useMultistackCheck", None) and self.useMultistackCheck.checked
            )
        multistack_metric = self._selected_profile_multistack_metric() or self.regMetric.currentText

        selected_profile = self._selected_config_profile()
        profile_cfg = self._selected_profile_config_dict()
        profile_masks_cfg = (profile_cfg.get("masks") or {}) if isinstance(profile_cfg, dict) else {}
        profile_multistack_cfg = (
            (profile_cfg.get("multistack_correction") or {}) if isinstance(profile_cfg, dict) else {}
        )

        mask_method = str(self.maskMethod.currentText or "adaptive")
        if mask_method == "global":
            mask_method = "seg_gauss"
        mask_low = float(self.maskLow.value)
        mask_high = float(self.maskHigh.value)
        contour_support_threshold = float(self.maskContourSupportThreshold.value)
        segmentation_cfg = {"method": mask_method}
        segmentation_cfg["use_segmentation_aligned_contour_support"] = bool(self.maskAlignedContourSupport.checked)
        if mask_method == "seg_gauss":
            segmentation_cfg.update(
                {
                    "trab_threshold": mask_low,
                    "cort_threshold": mask_high,
                    "gaussian_sigma": float(self.maskSigma.value),
                }
            )
        elif mask_method == "laplace_hamming":
            segmentation_cfg.update(
                {
                    "trab_threshold": contour_support_threshold,
                    "cort_threshold": float(getattr(self, "_lh_cort_support_threshold", 450.0)),
                    "gaussian_sigma": float(self.maskContourSigma.value),
                    "laplace_hamming_threshold": float(self.maskLaplaceThreshold.value),
                    "laplace_hamming_low_pass_cutoff": float(self.maskLaplaceLowPass.value),
                    "laplace_hamming_high_pass_cutoff": float(self.maskLaplaceHighPass.value),
                    "laplace_hamming_epsilon": float(self.maskLaplaceEpsilon.value),
                    "laplace_hamming_min_size_voxels": int(mask_high),
                }
            )
        else:
            segmentation_cfg.update(
                {
                    "adaptive_low_threshold": mask_low,
                    "adaptive_high_threshold": mask_high,
                }
            )
        periosteal_contour_method = str(self.maskPeriostealContour.currentData or "standard")
        if selected_profile == "ped-fx":
            periosteal_contour_method = "geodesic_fracture"
        outer_cfg = {
            "contour_method": periosteal_contour_method,
            "fill_holes": bool(self.maskGeodesicFillHoles.checked),
            "periosteal_threshold": contour_support_threshold,
            "geodesic_bone_threshold": float(self.maskGeodesicThreshold.value),
            "geodesic_fill_holes": bool(self.maskGeodesicFillHoles.checked),
            "periosteal_kernelsize": int(self.maskOuterKernel.value),
            "periosteal_open_radius": int(self.maskOuterOpen.value),
            "gaussian_sigma": float(self.maskContourSigma.value),
        }
        masks_override = {
            "outer": outer_cfg,
            "segmentation": segmentation_cfg,
            "generate": False,
            "overwrite": False,
        }
        if isinstance(profile_masks_cfg.get("roles"), list):
            masks_override["roles"] = list(profile_masks_cfg["roles"])
        endosteal_contour_method = str(self.maskEndostealContour.currentData or "standard")
        masks_override["inner"] = {
            "contour_method": endosteal_contour_method,
            "endosteal_threshold": float(self.maskEndostealThreshold.value),
            "endosteal_kernelsize": int(self.maskEndostealKernel.value),
            "gaussian_sigma": float(self.maskContourSigma.value),
        }
        if selected_profile == "ped-fx":
            masks_override["roles"] = ["full"]
            masks_override["inner"] = {"contour_method": "none"}
        initial_translation_voxels = [
            float(self.msInitTx.value),
            float(self.msInitTy.value),
            float(self.msInitTz.value),
        ]
        if selected_profile == "ped-fx":
            profile_initial_translation = profile_multistack_cfg.get("initial_translation_voxels")
            if isinstance(profile_initial_translation, (list, tuple)) and len(profile_initial_translation) >= 3:
                initial_translation_voxels = [float(v) for v in profile_initial_translation[:3]]

        settings = {
            "import": {
                # Do not fail when z-slices are not perfectly divisible by stack depth.
                # Keep the last partial stack to preserve data coverage.
                "on_incomplete_stack": "keep_last",
            },
            "masks": masks_override,
            "timelapsed_registration": {
                "metric": self.regMetric.currentText,
                "sampling_percentage": float(self.tlSampling.value),
                "number_of_resolutions": int(self.tlRes.value),
                "number_of_iterations": int(self.tlIter.value),
            },
            "multistack_correction": {
                "enabled": bool(multistack_enabled),
                "metric": multistack_metric,
                "sampling_percentage": float(self.msSampling.value),
                "number_of_resolutions": int(self.msRes.value),
                "number_of_iterations": int(self.msIter.value),
                "overlap_crop_buffer_voxels": int(self.msOverlapBuffer.value),
                "initial_translation_voxels": initial_translation_voxels,
            },
            "fusion": {
                "enable_filling": False,
            },
        }
        if self._selected_profile_is_custom() or bool(force_analysis_controls):
            settings["analysis"] = self._analysis_config_from_controls(pair_mode)
            settings["visualization"] = {
                "threshold": float(self.analysisThreshold.value),
                "cluster_size": int(self.analysisCluster.value),
                "label_map": label_map,
            }
        return settings

    def _run(self, args):
        try:
            stage = None
            if self._active_stage is not None:
                stage = self._active_stage
                self._set_stage_status(stage, "running")
            self._set_running_ui(True)
            self.logic.run_cli(args, on_output=self._show, on_finished=self._on_finished)
            self._show("[timelapsed-slicer] started: " + " ".join(args))
        except Exception as exc:
            self._set_running_ui(False)
            self.logic.cleanup_temp_files(remove_fallback=False)
            if stage is not None:
                self._set_stage_status(stage, "error")
            self._set_user_message(
                "error",
                "Could not start command",
                "Please verify the selected dataset and installed dependencies."
                f"<br><small>{str(exc)}</small>",
            )
            slicer.util.errorDisplay(str(exc))

    def _run_sequence(self, commands, stages=None):
        if not commands:
            return
        self._queued_commands = [list(cmd) for cmd in commands[1:]]
        self._queued_stages = list(stages[1:]) if stages else []
        self._active_stage = stages[0] if stages else None
        self._run(commands[0])

    def _set_running_ui(self, is_running):
        running = bool(is_running)
        for btn in [
            self.runMasksBtn,
            self.runTimelapseBtn,
            self.sceneRunButton,
            self.applyAnalysisSettingsBtn,
            self.runAnalysisBtn,
            self.clearLoadedResultsBtn,
            self.sceneExportCsvButton,
            self.sceneClearLoadedButton,
            self.saveAnalysisScenarioBtn,
        ]:
            self._set_widget_enabled_safe(btn, not running)
        self._set_widget_enabled_safe(getattr(self, "doNotGenerateMasksCheck", None), not running)
        self._set_widget_enabled_safe(getattr(self, "cancelRunBtn", None), running)
        self._set_widget_enabled_safe(getattr(self, "sceneInterruptButton", None), running)

    def _set_interactive_preview_busy(self, is_busy, message=None):
        busy = bool(is_busy)
        widgets = [
            self.analysisThresholdSlider,
            self.analysisThreshold,
            self.analysisClusterSlider,
            self.analysisCluster,
            self.analysisMethodCombo,
            self.analysisRestrictBoneSupportCheck,
            self.analysisBinaryReclassificationCheck,
            self.analysisFullMaskDilation,
            self.analysisMarrowMaskDilation,
            self.analysisMarrowMaskErosion,
            self.analysisGaussianFilterCheck,
            self.analysisGaussianSigma,
            self.applyAnalysisSettingsBtn,
            self.remodellingApplyInteractiveBtn,
            self.remodellingAutoUpdateCheck,
            self.saveAnalysisScenarioBtn,
        ]
        for widget in widgets:
            self._set_widget_enabled_safe(widget, not busy)
        if message is None:
            message = "Updating..." if busy else "Ready"
        label = getattr(self, "analysisStatusLabel", None)
        self._set_label_text_safe(label, message)
        self._set_widget_style_safe(label, "color: #996600;" if busy else "color: #666666;")
        try:
            if busy:
                slicer.app.setOverrideCursor(qt.Qt.WaitCursor)
            else:
                slicer.app.restoreOverrideCursor()
        except Exception:
            pass

    def _on_cancel_run(self):
        cancelled = self.logic.cancel_run()
        killed_external = self.logic.kill_external_runs()
        self.logic.cleanup_temp_files(remove_fallback=False)
        self._queued_commands = []
        self._queued_stages = []
        self._active_stage = None
        self._run_includes_analysis = False
        self._set_running_ui(False)
        if cancelled or killed_external:
            extra = ""
            if killed_external:
                extra = "<br>Also terminated stale process IDs: " + ", ".join(str(p) for p in killed_external)
            self._set_user_message("warn", "Run cancelled", "Current run was cancelled by user." + extra)
            self._show("[timelapsed-slicer] cancellation requested.")
        else:
            self._set_user_message(
                "info",
                "No active run",
                "There was no active process to cancel.",
            )

    def _manual_role_from_filename(self, path):
        stem = self._strip_manual_aim_suffix(Path(path).name)
        upper = stem.upper()
        suffix_roles = [
            ("_TRAB_MASK", "trab"),
            ("_CORT_MASK", "cort"),
            ("_FULL_MASK", "full"),
            ("_REGMASK", "regmask"),
            ("_SEG", "seg"),
        ]
        for suffix, role in suffix_roles:
            if upper.endswith(suffix):
                return stem[: -len(suffix)], role
        return stem, None

    def _strip_manual_aim_suffix(self, name):
        return re.sub(r"(?i)\.aim(?:;\d+)?$", "", str(name))

    def _manual_aim_version(self, path):
        match = re.search(r"(?i)\.aim(?:;(\d+))?$", Path(path).name)
        if not match or match.group(1) is None:
            return 0
        try:
            return int(match.group(1))
        except Exception:
            return 0

    def _prefer_manual_aim_candidate(self, current, candidate):
        if current is None:
            return candidate
        current_version = self._manual_aim_version(current)
        candidate_version = self._manual_aim_version(candidate)
        return candidate if candidate_version > current_version else current

    def _manual_site_from_token(self, token):
        site_aliases = {
            "DR": "radius",
            "RAD": "radius",
            "RADIUS": "radius",
            "DT": "tibia",
            "TIB": "tibia",
            "TIBIA": "tibia",
            "KN": "knee",
            "KNEE": "knee",
            "RL": "radiusleft",
            "RADIUS_LEFT": "radiusleft",
            "RR": "radiusright",
            "RADIUS_RIGHT": "radiusright",
            "TL": "tibialeft",
            "TIBIA_LEFT": "tibialeft",
            "TR": "tibiaright",
            "TIBIA_RIGHT": "tibiaright",
            "KL": "kneeleft",
            "KNEE_LEFT": "kneeleft",
            "KR": "kneeright",
            "KNEE_RIGHT": "kneeright",
        }
        return site_aliases.get(str(token or "").strip().upper(), "radius")

    def _manual_metadata_from_filename(self, base):
        # Example: STRAMBO_0003_TR_Y04.AIM -> sub STRAMBO_0003, site tibia_right, ses Y04.
        match = re.match(
            r"(?i)^(?P<subject>[A-Za-z][A-Za-z0-9]+_[0-9]+)_(?P<site>DR|DT|KN|RL|RR|TL|TR|KL|KR)_(?P<session>Y[0-9]+|T[0-9]+|C[0-9]+|BL|FL[0-9]*|FU[0-9]*)$",
            str(base or ""),
        )
        if not match:
            return {}
        session_id = str(match.group("session")).upper()
        return {
            "subject_id": match.group("subject"),
            "site": self._manual_site_from_token(match.group("site")),
            "session_id": session_id,
            "source_session_id": session_id,
        }

    def _manual_sessions_from_input_files(self, root):
        try:
            from timelapsedhrpqct.dataset.models import RawSession
        except Exception:
            return []

        root = Path(root)
        aim_paths = sorted(
            path for path in root.rglob("*")
            if path.is_file() and re.search(r"(?i)\.aim(?:;\d+)?$", path.name)
        )
        if not aim_paths:
            return []

        grouped = {}
        for path in aim_paths:
            base, role = self._manual_role_from_filename(path)
            entry = grouped.setdefault(
                base,
                {
                    "image": None,
                    "masks": {},
                    "seg": None,
                },
            )
            if role == "seg":
                entry["seg"] = path
            elif role:
                entry["masks"][role] = path
            else:
                existing = entry.get("image")
                entry["image"] = self._prefer_manual_aim_candidate(existing, path)

        sessions = []
        for idx, (base, entry) in enumerate(sorted(grouped.items()), start=1):
            image_path = entry.get("image")
            if image_path is None:
                continue
            metadata = self._manual_metadata_from_filename(base)
            sessions.append(
                RawSession(
                    subject_id=metadata.get("subject_id", "MANUAL"),
                    session_id=metadata.get("session_id", f"T{idx}"),
                    raw_image_path=Path(image_path),
                    source_session_id=metadata.get("source_session_id", str(base)),
                    site=metadata.get("site", "radius"),
                    stack_index=None,
                    raw_mask_paths=dict(entry.get("masks") or {}),
                    raw_seg_path=entry.get("seg"),
                )
            )

        if sessions:
            return sessions

        return [
            RawSession(
                subject_id="MANUAL",
                session_id=f"T{idx}",
                raw_image_path=path,
                source_session_id=path.stem,
                site="radius",
                stack_index=None,
                raw_mask_paths={},
                raw_seg_path=None,
            )
            for idx, path in enumerate(aim_paths, start=1)
        ]

    def _manual_sessions_need_correction(self, sessions):
        if not sessions:
            return True
        for session in sessions:
            subject_id = str(getattr(session, "subject_id", "") or "").strip().upper()
            session_id = str(getattr(session, "session_id", "") or "").strip().upper()
            source_session_id = str(getattr(session, "source_session_id", "") or "").strip()
            if subject_id == "MANUAL" or not subject_id:
                return True
            if not session_id or re.fullmatch(r"T\d+", session_id) and source_session_id not in {"", session_id}:
                return True
        return False

    def _on_parse(self):
        if not self.logic.is_pipeline_available():
            slicer.util.errorDisplay("Please install timelapsed-hrpqct first.")
            return
        root = self._dataset_root()
        if root is None:
            self._set_user_message(
                "warn",
                "Select dataset folder",
                "Choose a dataset root directory before parsing.",
            )
            slicer.util.errorDisplay("Select a dataset folder first.")
            return

        self._clear_user_message()
        self._set_stage_status("parse", "running")
        sessions, err, mode_used = self.logic.parse_input(
            root,
            parse_mode=self._selected_parse_mode(),
        )
        if err:
            manual_sessions = self._manual_sessions_from_input_files(root)
            if manual_sessions:
                needs_correction = self._manual_sessions_need_correction(manual_sessions)
                self._manual_parse_active = True
                self._last_parsed_sessions = list(manual_sessions)
                self._parsed_baseline_rows = []
                self._populate_parse_table(manual_sessions)
                if needs_correction:
                    try:
                        self.parseBox.collapsed = False
                    except Exception:
                        pass
                self.parseSummaryLabel.text = (
                    "Parse summary: fallback filename parse "
                    f"({len(manual_sessions)} AIM image row(s) prepared)."
                    if not needs_correction
                    else (
                        "Parse summary: manual correction needed "
                        f"({len(manual_sessions)} AIM image row(s) prepared)."
                    )
                )
            else:
                needs_correction = True
                self._manual_parse_active = False
                self.parseTable.setRowCount(0)
                self._last_parsed_sessions = []
                self._parsed_baseline_rows = []
                self.parseSummaryLabel.text = "Parse summary: failed"
            self.parseSummaryLabel.styleSheet = "color: #cc5500;" if needs_correction else "color: #228b22;"
            self._refresh_processing_subjects()
            self._set_stage_status("parse", "error" if needs_correction else "done")
            if needs_correction:
                detail = (
                    "Could not parse filenames automatically. "
                    "Use Parse Details to correct Subject, Site, Session, and Stack values, "
                    "then run the pipeline from the corrected table."
                    if manual_sessions
                    else "Could not parse filenames automatically and no AIM files were found for manual correction."
                )
                self._set_user_message("error", "Parse needs correction", detail)
            else:
                self._set_user_message(
                    "success",
                    "Parse used fallback",
                    "Automatic parser failed, but Slicer recovered structured filename rows. "
                    "Review Parse Details if needed, then run the pipeline.",
                )
            self._show("[parse] automatic parse failed:")
            self._show(str(err))
            return

        self._last_parse_mode_used = mode_used
        self._manual_parse_active = False
        self._show(f"[parse] discovered {len(sessions)} sessions under {root}")
        self._last_parsed_sessions = list(sessions)
        self._parsed_baseline_rows = [
            (
                str(getattr(s, "subject_id", "")).strip(),
                str(getattr(s, "site", "")).strip().lower(),
                str(getattr(s, "session_id", "")).strip(),
                "-" if getattr(s, "stack_index", None) is None else str(int(getattr(s, "stack_index"))),
            )
            for s in sessions
        ]
        self._refresh_processing_subjects()
        self._populate_parse_table(sessions)
        self._refresh_patient_list()
        self._set_stage_status("parse", "done")
        mode_label = f" via {mode_used} mode" if mode_used else ""
        self._set_user_message(
            "success",
            "Parse successful",
            f"Discovered {len(sessions)} session(s){mode_label}.",
        )

    def _site_options(self):
        return [
            "radius",
            "tibia",
            "knee",
            "radiusleft",
            "radiusright",
            "tibialeft",
            "tibiaright",
            "kneeleft",
            "kneeright",
        ]

    def _session_options(self, sessions):
        default = ["T1", "T2", "T3", "T4", "T5", "BL", "FL", "C1", "C2", "C3"]
        seen = set(default)
        out = list(default)
        for session in sessions:
            sid = str(getattr(session, "session_id", "")).strip()
            if sid and sid not in seen:
                out.append(sid)
                seen.add(sid)
        return out

    def _on_parse_table_item_changed(self, item):
        if self._updating_parse_table or item is None:
            return
        row = int(item.row())
        col = int(item.column())
        if row < 0 or row >= len(self._last_parsed_sessions):
            return
        text = str(item.text() or "").strip()
        if col == 0:
            if text:
                self._last_parsed_sessions[row].subject_id = text
        elif col == 3:
            if text == "-" or text == "":
                self._last_parsed_sessions[row].stack_index = None
            else:
                try:
                    stack_value = int(text)
                    self._last_parsed_sessions[row].stack_index = stack_value if stack_value > 0 else None
                except Exception:
                    pass
        self._refresh_patient_list()
        self._refresh_processing_subjects()

    def _on_parse_site_changed(self, row, text):
        if self._updating_parse_table:
            return
        if row < 0 or row >= len(self._last_parsed_sessions):
            return
        site = str(text or "").strip().lower()
        if not site:
            return
        self._last_parsed_sessions[row].site = site
        self._refresh_patient_list()
        self._refresh_processing_subjects()

    def _on_parse_session_changed(self, row, text):
        if self._updating_parse_table:
            return
        if row < 0 or row >= len(self._last_parsed_sessions):
            return
        session_id = str(text or "").strip()
        if not session_id:
            return
        self._last_parsed_sessions[row].session_id = session_id
        self._refresh_patient_list()
        self._refresh_processing_subjects()

    def _selected_processing_subject(self):
        combo = getattr(self, "processingSubjectCombo", None)
        if not self._qt_object_alive(combo):
            return None
        try:
            text = str(getattr(combo, "currentText", "")).strip()
        except (RuntimeError, ValueError):
            return None
        if not text or text == "All subjects":
            return None
        return text

    def _selected_processing_site(self):
        combo = getattr(self, "processingSiteCombo", None)
        if not self._qt_object_alive(combo):
            return None
        try:
            text = str(getattr(combo, "currentText", "")).strip()
        except (RuntimeError, ValueError):
            return None
        if not text or text == "All sites":
            return None
        return text

    def _sessions_for_processing_scope(self):
        self._sync_sessions_from_parse_table()
        sessions = list(self._last_parsed_sessions or [])
        subject = self._selected_processing_subject()
        site = self._selected_processing_site()
        if subject is None and site is None:
            return sessions, None, None
        scoped = [
            s for s in sessions
            if (
                subject is None
                or str(getattr(s, "subject_id", "")).strip() == subject
            )
            and (
                site is None
                or str(getattr(s, "site", "")).strip().lower() == site
            )
        ]
        return scoped, subject, site

    def _refresh_processing_subjects(self):
        combo = getattr(self, "processingSubjectCombo", None)
        if not self._qt_object_alive(combo):
            return
        prev = self._selected_processing_subject()
        combo.clear()
        combo.addItem("All subjects")
        seen = set()
        for s in (self._last_parsed_sessions or []):
            subject = str(getattr(s, "subject_id", "")).strip()
            if subject and subject not in seen:
                seen.add(subject)
                combo.addItem(subject)
        if prev and prev in seen:
            idx = combo.findText(prev)
            if idx >= 0:
                combo.setCurrentIndex(idx)
        self._refresh_processing_sites()

    def _refresh_processing_sites(self, *_args):
        combo = getattr(self, "processingSiteCombo", None)
        if not self._qt_object_alive(combo):
            return
        prev = self._selected_processing_site()
        subject = self._selected_processing_subject()
        combo.clear()
        combo.addItem("All sites")
        seen = set()
        for s in (self._last_parsed_sessions or []):
            if subject is not None and str(getattr(s, "subject_id", "")).strip() != subject:
                continue
            site = str(getattr(s, "site", "")).strip().lower()
            if site and site not in seen:
                seen.add(site)
                combo.addItem(site)
        if prev and prev in seen:
            idx = combo.findText(prev)
            if idx >= 0:
                combo.setCurrentIndex(idx)

    def _sanitize_name_token(self, text):
        token = re.sub(r"[^A-Za-z0-9]+", "_", str(text or "").strip())
        token = token.strip("_")
        return token or "UNKNOWN"

    def _site_to_token(self, site):
        site_norm = str(site or "").strip().lower()
        mapping = {
            "radius": "DR",
            "tibia": "DT",
            "knee": "KN",
            "radiusleft": "RL",
            "radius_left": "RL",
            "radiusright": "RR",
            "radius_right": "RR",
            "tibialeft": "TL",
            "tibia_left": "TL",
            "tibiaright": "TR",
            "tibia_right": "TR",
            "kneeleft": "KL",
            "knee_left": "KL",
            "kneeright": "KR",
            "knee_right": "KR",
        }
        return mapping.get(site_norm, self._sanitize_name_token(site).upper())

    def _mask_role_suffix(self, role):
        role_norm = str(role or "").strip().lower()
        if role_norm == "trab":
            return "_TRAB_MASK"
        if role_norm == "cort":
            return "_CORT_MASK"
        if role_norm == "full":
            return "_FULL_MASK"
        if role_norm == "regmask":
            return "_REGMASK"
        return "_" + self._sanitize_name_token(role).upper()

    def _reset_temp_input_root(self):
        if not self._temp_input_root:
            return
        try:
            shutil.rmtree(self._temp_input_root, ignore_errors=True)
        except Exception:
            pass
        self._temp_input_root = None

    def _sync_sessions_from_parse_table(self):
        if not self._last_parsed_sessions:
            return
        rows = min(int(self.parseTable.rowCount), len(self._last_parsed_sessions))
        for row in range(rows):
            session = self._last_parsed_sessions[row]
            subj_item = self.parseTable.item(row, 0)
            stack_item = self.parseTable.item(row, 3)
            site_widget = self.parseTable.cellWidget(row, 1)
            session_widget = self.parseTable.cellWidget(row, 2)

            subject = str(subj_item.text() if subj_item else "").strip()
            site = str(site_widget.currentText if site_widget else "").strip().lower()
            session_id = str(session_widget.currentText if session_widget else "").strip()
            stack_text = str(stack_item.text() if stack_item else "").strip()

            if subject:
                session.subject_id = subject
            if site:
                session.site = site
            if session_id:
                session.session_id = session_id
            if stack_text in {"", "-"}:
                session.stack_index = None
            else:
                try:
                    stack_value = int(stack_text)
                    session.stack_index = stack_value if stack_value > 0 else None
                except Exception:
                    pass

    def _has_parse_overrides(self):
        if self._manual_parse_active:
            return True
        if not self._last_parsed_sessions or not self._parsed_baseline_rows:
            return False
        table_rows = int(self.parseTable.rowCount)
        if table_rows != len(self._parsed_baseline_rows):
            return False
        for row in range(table_rows):
            subj_item = self.parseTable.item(row, 0)
            stack_item = self.parseTable.item(row, 3)
            site_widget = self.parseTable.cellWidget(row, 1)
            session_widget = self.parseTable.cellWidget(row, 2)
            subj_ui = str(subj_item.text() if subj_item else "").strip()
            site_ui = str(site_widget.currentText if site_widget else "").strip().lower()
            ses_ui = str(session_widget.currentText if session_widget else "").strip()
            stack_ui = str(stack_item.text() if stack_item else "").strip()
            subj0, site0, ses0, stack0 = self._parsed_baseline_rows[row]
            if (subj_ui, site_ui, ses_ui, stack_ui) != (subj0, site0, ses0, stack0):
                return True
        return False

    def _raw_ingest_mode(self):
        copy_raw = bool(getattr(self.copyRawInputsCheck, "checked", False))
        restructure_raw = bool(getattr(self.restructureRawCheck, "checked", False))
        if copy_raw and restructure_raw:
            raise ValueError("Select either Copy raw inputs or Restructure raw inputs, not both.")
        if restructure_raw:
            return "restructure"
        if copy_raw:
            return "copy"
        return "none"

    def _raw_ingest_cli_flags(self, mode: str):
        if mode == "copy":
            return ["--copy-raw-inputs"]
        if mode == "restructure":
            return ["--restructure-raw"]
        return []

    def _storage_cli_flags(self):
        if not self.logic.run_cli_supports_option("--storage-mode"):
            return []
        mode = str(getattr(self.storageModeCombo, "currentData", "minimal") or "minimal")
        if mode == "full":
            return ["--storage-mode", "full"]
        return ["--storage-mode", "minimal"]

    def _selected_parse_mode(self):
        mode = str(getattr(self.parseModeCombo, "currentText", "auto")).strip().lower()
        if mode not in {"auto", "filename", "header"}:
            return "auto"
        return mode

    def _effective_force_header_for_run(self):
        mode = self._selected_parse_mode()
        if mode == "header":
            return True
        if mode == "filename":
            return False
        # auto: if parse already resolved to header fallback, keep run consistent.
        return str(self._last_parse_mode_used or "").strip().lower() == "header"

    def _raw_discovery_cli_flags(self):
        if self._effective_force_header_for_run():
            return ["--force-header-discovery"]
        return []

    def _make_run_input_root(self, dataset_root: Path, ingest_mode: str = "none"):
        return self._make_run_input_root_for_sessions(dataset_root, ingest_mode=ingest_mode)

    def _make_run_input_root_for_sessions(
        self,
        dataset_root: Path,
        ingest_mode: str = "none",
        sessions: list | None = None,
        force_virtual_root: bool = False,
    ):
        self._sync_sessions_from_parse_table()
        selected_sessions = sessions if sessions is not None else list(self._last_parsed_sessions or [])
        if not selected_sessions:
            return dataset_root

        has_overrides = self._has_parse_overrides()
        if not force_virtual_root and not has_overrides:
            return dataset_root

        if ingest_mode == "restructure":
            self._set_user_message(
                "warn",
                "Restructure requires direct source paths",
                (
                    "Parse table edits create a temporary virtual input root. "
                    "Disable 'Restructure raw inputs' or keep parse values as-is before running."
                ),
            )
            return None

        self._reset_temp_input_root()
        tmp_root = Path(tempfile.mkdtemp(prefix="timelapsed_slicer_input_"))
        created = 0

        def _link_or_copy(src: Path, dst: Path):
            dst.parent.mkdir(parents=True, exist_ok=True)
            try:
                os.symlink(str(src), str(dst))
            except Exception:
                shutil.copy2(src, dst)

        for session in selected_sessions:
            subject_id = self._sanitize_name_token(getattr(session, "subject_id", ""))
            site_token = self._site_to_token(getattr(session, "site", "radius"))
            session_id = self._sanitize_name_token(getattr(session, "session_id", ""))
            stack_index = getattr(session, "stack_index", None)
            stack_chunk = ""
            if stack_index is not None:
                try:
                    stack_chunk = f"_STACK{int(stack_index):02d}"
                except Exception:
                    stack_chunk = f"_STACK{self._sanitize_name_token(stack_index)}"

            base = f"{subject_id}_{site_token}{stack_chunk}_{session_id}"
            image_src = Path(getattr(session, "raw_image_path"))
            image_dst = tmp_root / f"{base}.AIM"
            _link_or_copy(image_src, image_dst)
            created += 1

            raw_masks = getattr(session, "raw_mask_paths", {}) or {}
            for role, mask_path in raw_masks.items():
                mask_src = Path(mask_path)
                suffix = self._mask_role_suffix(role)
                mask_dst = tmp_root / f"{base}{suffix}.AIM"
                _link_or_copy(mask_src, mask_dst)
                created += 1

            seg_path = getattr(session, "raw_seg_path", None)
            if seg_path:
                seg_src = Path(seg_path)
                seg_dst = tmp_root / f"{base}_SEG.AIM"
                _link_or_copy(seg_src, seg_dst)
                created += 1

        self._temp_input_root = str(tmp_root)
        mode_label = "scoped" if force_virtual_root and not has_overrides else "corrected"
        self._show(f"[parse] using {mode_label} virtual input root: {tmp_root} ({created} file links)")
        return tmp_root

    def _populate_parse_table(self, sessions):
        self._updating_parse_table = True
        try:
            self.parseTable.setRowCount(len(sessions))
            self.parseSummaryLabel.text = (
                f"Parse summary: {len(sessions)} session(s) discovered "
                "(Subject is editable. Site/Session are dropdown-correctable; original session is read-only.)"
            )
            self.parseSummaryLabel.styleSheet = "color: #228b22;"
            site_options = self._site_options()
            session_options = self._session_options(sessions)

            for row, session in enumerate(sessions):
                subject = str(getattr(session, "subject_id", ""))
                site = str(getattr(session, "site", ""))
                session_id = str(getattr(session, "session_id", ""))
                source_session_id = str(
                    getattr(session, "source_session_id", "") or session_id
                )
                stack_index = getattr(session, "stack_index", None)
                stack_text = "-" if stack_index is None else str(stack_index)

                raw_image = getattr(session, "raw_image_path", None)
                image_name = Path(raw_image).name if raw_image else "-"

                raw_masks = getattr(session, "raw_mask_paths", {}) or {}
                mask_roles = ", ".join(sorted(str(k) for k in raw_masks.keys())) if raw_masks else "-"

                seg_path = getattr(session, "raw_seg_path", None)
                seg_text = "yes" if seg_path else "no"

                # Subject (editable text)
                subject_item = qt.QTableWidgetItem(subject)
                subject_item.setFlags(subject_item.flags() | qt.Qt.ItemIsEditable)
                self.parseTable.setItem(row, 0, subject_item)

                # Site (dropdown)
                site_combo = qt.QComboBox()
                site_combo.addItems(site_options)
                site_current = site if site in site_options else "radius"
                site_combo.setCurrentText(site_current)
                site_combo.currentTextChanged.connect(
                    lambda text, r=row: self._on_parse_site_changed(r, text)
                )
                self.parseTable.setCellWidget(row, 1, site_combo)

                # Session (dropdown + editable)
                ses_combo = qt.QComboBox()
                ses_combo.setEditable(True)
                ses_combo.addItems(session_options)
                ses_combo.setCurrentText(session_id)
                ses_combo.currentTextChanged.connect(
                    lambda text, r=row: self._on_parse_session_changed(r, text)
                )
                self.parseTable.setCellWidget(row, 2, ses_combo)

                # Stack
                stack_item = qt.QTableWidgetItem(stack_text)
                stack_item.setFlags(stack_item.flags() | qt.Qt.ItemIsEditable)
                self.parseTable.setItem(row, 3, stack_item)

                # Read-only informative columns
                for col, value in [
                    (4, image_name),
                    (5, mask_roles),
                    (6, seg_text),
                    (7, source_session_id),
                ]:
                    item = qt.QTableWidgetItem(value)
                    item.setFlags(item.flags() & ~qt.Qt.ItemIsEditable)
                    self.parseTable.setItem(row, col, item)

            self.parseTable.resizeColumnsToContents()
        finally:
            self._updating_parse_table = False

    def _require_pipeline_installed(self) -> bool:
        if self.logic.is_pipeline_available():
            return True
        slicer.util.errorDisplay("Please install timelapsed-hrpqct first.")
        return False

    def _require_dataset_root(self) -> Path | None:
        root = self._dataset_root()
        if root is not None:
            return root
        slicer.util.errorDisplay("Select a dataset root first.")
        return None

    def _require_results_root(self, message: str = "Could not resolve results dataset path.") -> Path | None:
        imported = self._imported_dataset_root()
        if imported is not None:
            return imported
        slicer.util.errorDisplay(message)
        return None

    def _on_run_masks(self):
        self._set_user_message(
            "info",
            "Mask generation moved to Bone Contouring",
            "Timelapsed now only consumes existing registration masks, analysis ROIs, and bone segmentations.",
        )
        slicer.util.infoDisplay(
            "Mask generation moved to Bone Contouring. Timelapsed will not generate masks."
        )

    def _on_run_timelapse(self):
        if not self._require_pipeline_installed():
            return
        source_root = self._require_dataset_root()
        if source_root is None:
            return
        try:
            raw_ingest_mode = self._raw_ingest_mode()
        except ValueError as exc:
            slicer.util.errorDisplay(str(exc))
            return
        scoped_sessions, scoped_subject, scoped_site = self._sessions_for_processing_scope()
        if (scoped_subject or scoped_site) and not scoped_sessions:
            slicer.util.errorDisplay("No parsed sessions available for the selected processing scope.")
            return
        run_root = self._make_run_input_root_for_sessions(
            source_root,
            ingest_mode=raw_ingest_mode,
            sessions=scoped_sessions,
            force_virtual_root=bool(scoped_subject or scoped_site),
        )
        if run_root is None:
            return
        mode = self._selected_run_mode(scoped_sessions)

        imported = self._require_results_root()
        if imported is None:
            return
        self._set_stage_status("registration", "pending")
        self._set_stage_status("analysis", "pending")
        self._active_stage = "registration"
        self._is_full_pipeline_run = False
        self._run_skips_mask_generation = True
        self._run_includes_analysis = True
        cfg = self.logic.create_override_config(
            self._settings_override(multistack_enabled=(mode == "multistack")),
            results_root=imported,
        )
        self._run(
            [
                "run",
                str(run_root),
                "--output-root",
                str(imported),
                "--mode",
                mode,
                "--skip-mask-generation",
                *self._raw_ingest_cli_flags(raw_ingest_mode),
                *self._storage_cli_flags(),
                *self._raw_discovery_cli_flags(),
                *self._profile_cli_args(),
                "--config",
                cfg,
            ]
        )

    def _on_run_analysis(self):
        if not self._require_pipeline_installed():
            return
        root = self._require_dataset_root()
        if root is None:
            return
        scoped_subject = self._selected_processing_subject()
        scoped_site = self._selected_processing_site()

        imported = self._require_results_root("Could not resolve imported dataset path.")
        if imported is None:
            return
        self._set_stage_status("analysis", "pending")
        self._active_stage = "analysis"
        self._is_full_pipeline_run = False
        self._run_skips_mask_generation = True
        self._run_includes_analysis = False
        cfg = self.logic.create_override_config(self._settings_override(), results_root=imported)
        run_args = [
            "analyse",
            str(imported),
            *(["--subject", str(scoped_subject)] if scoped_subject is not None else []),
            *(["--site", str(scoped_site)] if scoped_site else []),
            *self._profile_cli_args(),
            "--config",
            cfg,
        ]
        if self._selected_profile_is_custom():
            run_args.extend([
                "--thr",
                str(float(self.analysisThreshold.value)),
                "--clusters",
                str(int(self.analysisCluster.value)),
            ])
        self._run(run_args)

    def _current_analysis_requires_segmentation(self):
        return self._current_analysis_method() in {"grayscale_and_binary", "grayscale_marrow_mask"}

    def _raw_sessions_have_segmentation_inputs(self, sessions):
        sessions = list(sessions or [])
        if not sessions:
            return False
        for session in sessions:
            seg_path = getattr(session, "raw_seg_path", None)
            if not seg_path or not Path(seg_path).exists():
                return False
        return True

    def _existing_imported_segmentations_complete(self, imported, sessions=None, scoped_subject=None, scoped_site=None):
        try:
            from timelapsedhrpqct.dataset.artifacts import iter_imported_stack_records
        except Exception as exc:
            self._show(f"[masks] could not inspect existing imported segmentations: {exc}")
            return False

        session_keys = set()
        for session in sessions or []:
            subject = str(getattr(session, "subject_id", "")).strip()
            site = str(getattr(session, "site", "")).strip().lower()
            session_id = str(getattr(session, "session_id", "")).strip()
            stack_index = getattr(session, "stack_index", None)
            try:
                stack_index = int(stack_index) if stack_index is not None else None
            except Exception:
                stack_index = None
            if subject and site and session_id:
                session_keys.add((subject, site, session_id, stack_index))

        records = []
        for record in iter_imported_stack_records(imported):
            if scoped_subject is not None and str(getattr(record, "subject_id", "")) != str(scoped_subject):
                continue
            if scoped_site is not None and str(getattr(record, "site", "")).strip().lower() != str(scoped_site):
                continue
            if session_keys:
                key = (
                    str(getattr(record, "subject_id", "")).strip(),
                    str(getattr(record, "site", "")).strip().lower(),
                    str(getattr(record, "session_id", "")).strip(),
                    int(getattr(record, "stack_index", 0)),
                )
                # Single-stack parsed sessions may have no explicit stack index.
                key_no_stack = (key[0], key[1], key[2], None)
                if key not in session_keys and key_no_stack not in session_keys:
                    continue
            records.append(record)

        if not records:
            return False
        return all(getattr(record, "seg_path", None) and Path(record.seg_path).exists() for record in records)

    def _can_skip_mask_generation(self, imported, sessions=None, scoped_subject=None, scoped_site=None):
        if not self._current_analysis_requires_segmentation():
            return True
        if self._raw_sessions_have_segmentation_inputs(sessions):
            return True
        return self._existing_imported_segmentations_complete(
            imported,
            sessions=sessions,
            scoped_subject=scoped_subject,
            scoped_site=scoped_site,
        )

    def _batch_required_analysis_roles(self, settings):
        analysis_cfg = settings.get("analysis") or {}
        if not analysis_cfg:
            profile_cfg = self._selected_profile_config_dict()
            analysis_cfg = (profile_cfg.get("analysis") or {}) if isinstance(profile_cfg, dict) else {}
        roles = analysis_cfg.get("compartments")
        if not isinstance(roles, list) or not roles:
            roles = ((settings.get("masks") or {}).get("roles") or ["full", "trab", "cort"])
        normalized = []
        for role in roles:
            role_text = str(role or "").strip().lower()
            if role_text and role_text not in normalized:
                normalized.append(role_text)
        return normalized or ["full", "trab", "cort"]

    def _session_preflight_key(self, obj):
        subject = str(getattr(obj, "subject_id", "") or "").strip()
        site = str(getattr(obj, "site", "") or "").strip().lower()
        session_id = str(getattr(obj, "session_id", "") or "").strip()
        stack_index = getattr(obj, "stack_index", None)
        try:
            stack_index = int(stack_index) if stack_index is not None else None
        except Exception:
            stack_index = None
        return subject, site, session_id, stack_index

    def _session_preflight_label(self, key):
        subject, site, session_id, stack_index = key
        stack_text = "" if stack_index is None else f" stack-{int(stack_index):02d}"
        return f"sub-{subject} site-{site} ses-{session_id}{stack_text}"

    def _matching_preflight_entry(self, entries, key):
        if key in entries:
            return entries[key]
        subject, site, session_id, stack_index = key
        candidates = []
        if stack_index is None:
            candidates.append((subject, site, session_id, 1))
        else:
            candidates.append((subject, site, session_id, None))
        for candidate in candidates:
            if candidate in entries:
                return entries[candidate]
        return None

    def _mask_role_exists_for_preflight(self, masks, role):
        role = str(role or "").strip().lower()
        if role == "full":
            return (
                self._path_exists(masks.get("full"))
                or self._path_exists(masks.get("regmask"))
                or any(str(name).lower().startswith("roi") and self._path_exists(path) for name, path in masks.items())
                or (self._path_exists(masks.get("trab")) and self._path_exists(masks.get("cort")))
            )
        return self._path_exists(masks.get(role))

    def _registration_mask_exists_for_preflight(self, masks):
        return (
            self._path_exists(masks.get("regmask"))
            or self._path_exists(masks.get("full"))
            or (self._path_exists(masks.get("trab")) and self._path_exists(masks.get("cort")))
            or any(str(role).lower().startswith("roi") and self._path_exists(path) for role, path in masks.items())
        )

    def _resolve_batch_effective_analysis_roles(self, entries, configured_roles):
        role_sets = []
        for entry in entries:
            roles = {
                str(role).lower()
                for role, path in (entry.get("masks") or {}).items()
                if self._path_exists(path)
            }
            role_sets.append(roles)
        if not role_sets:
            return []
        common_roles = set(role_sets[0])
        for roles in role_sets[1:]:
            common_roles &= roles
        roi_roles = sorted(role for role in common_roles if role.startswith("roi"))
        if roi_roles:
            return roi_roles
        if "regmask" in common_roles:
            return ["regmask"]
        available_configured = [
            role for role in configured_roles
            if all(self._mask_role_exists_for_preflight(entry.get("masks") or {}, role) for entry in entries)
        ]
        if available_configured:
            return available_configured
        fallback = [
            role for role in ("trab", "cort", "full")
            if all(self._mask_role_exists_for_preflight(entry.get("masks") or {}, role) for entry in entries)
        ]
        return fallback

    def _missing_batch_required_inputs(self, imported, sessions=None, scoped_subject=None, scoped_site=None, settings=None):
        settings = settings or self._settings_override()
        configured_roles = self._batch_required_analysis_roles(settings)
        sessions = list(sessions or self._last_parsed_sessions or [])
        entries = {}

        for session in sessions:
            key = self._session_preflight_key(session)
            if not key[0] or not key[1] or not key[2]:
                continue
            raw_masks = {
                str(role).lower(): Path(path)
                for role, path in (getattr(session, "raw_mask_paths", {}) or {}).items()
                if path is not None
            }
            entries[key] = {
                "label": self._session_preflight_label(key),
                "masks": raw_masks,
                "seg": Path(getattr(session, "raw_seg_path")) if getattr(session, "raw_seg_path", None) else None,
            }

        try:
            from timelapsedhrpqct.dataset.artifacts import iter_imported_stack_records
        except Exception as exc:
            self._show(f"[preflight] could not inspect imported masks: {exc}")
            imported_records = []
        else:
            imported_records = list(iter_imported_stack_records(imported))

        for record in imported_records:
            key = self._session_preflight_key(record)
            if scoped_subject is not None and str(key[0]) != str(scoped_subject):
                continue
            if scoped_site is not None and str(key[1]).strip().lower() != str(scoped_site):
                continue
            entry = self._matching_preflight_entry(entries, key)
            if entry is None:
                if sessions:
                    continue
                entry = {
                    "label": self._session_preflight_label(key),
                    "masks": {},
                    "seg": None,
                }
                entries[key] = entry
            for role, path in (getattr(record, "mask_paths", {}) or {}).items():
                if self._path_exists(path):
                    entry["masks"][str(role).lower()] = Path(path)
            seg_path = getattr(record, "seg_path", None)
            if self._path_exists(seg_path):
                entry["seg"] = Path(seg_path)

        missing = []
        for key, entry in sorted(entries.items(), key=lambda item: item[0]):
            if not self._registration_mask_exists_for_preflight(entry.get("masks") or {}):
                missing.append(f"{entry['label']}: registration mask/ROI")
            if self._current_analysis_requires_segmentation() and not self._path_exists(entry.get("seg")):
                missing.append(f"{entry['label']}: bone segmentation")

        groups = {}
        for key, entry in entries.items():
            subject, site, _session_id, stack_index = key
            groups.setdefault((subject, site, stack_index), []).append(entry)
        for (subject, site, stack_index), group_entries in sorted(groups.items(), key=lambda item: item[0]):
            if len(group_entries) < 2:
                continue
            effective_roles = self._resolve_batch_effective_analysis_roles(group_entries, configured_roles)
            if not effective_roles:
                stack_text = "" if stack_index is None else f" stack-{int(stack_index):02d}"
                missing.append(
                    f"sub-{subject} site-{site}{stack_text}: analysis ROI "
                    f"(configured roles: {', '.join(configured_roles)})"
                )
        return missing

    def _auto_mode_from_sessions(self, sessions=None):
        candidate_sessions = sessions if sessions is not None else (self._last_parsed_sessions or [])
        has_multistack = any(
            getattr(s, "stack_index", None) is not None and int(getattr(s, "stack_index", 0)) > 1
            for s in candidate_sessions
        )
        return "multistack" if has_multistack else "regular"

    def _on_run_full_pipeline(self):
        if not self._require_pipeline_installed():
            return
        source_root = self._require_dataset_root()
        if source_root is None:
            self._set_user_message("warn", "Select dataset folder", "Choose a dataset root first.")
            return
        try:
            raw_ingest_mode = self._raw_ingest_mode()
        except ValueError as exc:
            slicer.util.errorDisplay(str(exc))
            return
        scoped_sessions, scoped_subject, scoped_site = self._sessions_for_processing_scope()
        if (scoped_subject or scoped_site) and not scoped_sessions:
            slicer.util.errorDisplay("No parsed sessions available for the selected processing scope.")
            return
        run_root = self._make_run_input_root_for_sessions(
            source_root,
            ingest_mode=raw_ingest_mode,
            sessions=scoped_sessions,
            force_virtual_root=bool(scoped_subject or scoped_site),
        )
        if run_root is None:
            return
        imported = self._require_results_root("Could not resolve imported dataset path.")
        if imported is None:
            return
        mode = self._selected_run_mode(scoped_sessions)
        settings_override = self._settings_override(multistack_enabled=(mode == "multistack"))
        missing_inputs = self._missing_batch_required_inputs(
            imported,
            sessions=scoped_sessions,
            scoped_subject=scoped_subject,
            scoped_site=scoped_site,
            settings=settings_override,
        )
        if missing_inputs:
            details = "\n".join(f"- {item}" for item in missing_inputs[:12])
            if len(missing_inputs) > 12:
                details += f"\n- ... {len(missing_inputs) - 12} more"
            message = (
                "Missing required Timelapsed input(s):\n"
                f"{details}\n\n"
                "Generate missing masks, ROIs, or bone segmentations in Bone Contouring before running Timelapsed."
            )
            self._set_user_message(
                "error",
                "Missing required inputs",
                (
                    "Timelapsed no longer generates masks automatically. "
                    "Prepare the missing masks, ROIs, or segmentations in Bone Contouring first."
                ),
            )
            self._show("[preflight] " + message.replace("\n", " "))
            slicer.util.errorDisplay(message)
            return
        cfg = self.logic.create_override_config(
            settings_override,
            results_root=imported,
        )
        self._set_user_message(
            "info",
            "Running full pipeline",
            (
                f"Mode: <b>{mode}</b>. "
                "Using existing/provided masks, segmentations, and ROI inputs. "
                "Generate missing inputs in Bone Contouring before running Timelapsed."
            ),
        )
        for s in ("registration", "analysis"):
            self._set_stage_status(s, "pending")
        self._active_stage = "registration"
        self._is_full_pipeline_run = True
        self._run_skips_mask_generation = True
        self._run_includes_analysis = True
        run_args = [
            "run",
            str(run_root),
            "--output-root",
            str(imported),
            "--mode",
            mode,
            *self._raw_ingest_cli_flags(raw_ingest_mode),
            *self._storage_cli_flags(),
            *self._raw_discovery_cli_flags(),
            *self._profile_cli_args(),
            "--config",
            cfg,
        ]
        run_args.append("--skip-mask-generation")
        self._run(run_args)

    def _on_finished(self, exit_code, exit_status):
        self._show(f"[timelapsed-slicer] finished with exit code {exit_code}")
        self._set_running_ui(False)
        if self._active_stage is not None:
            self._set_stage_status(self._active_stage, "done" if int(exit_code) == 0 else "error")
        if int(exit_code) != 0:
            self.logic.cleanup_temp_files(remove_fallback=False)
            self._set_user_message(
                "error",
                "Pipeline step failed",
                "Check the log below for the failing command and verify filenames/config."
                " You can rerun individual steps after fixing the issue.",
            )
            self._queued_commands = []
            self._queued_stages = []
            self._active_stage = None
            self._is_full_pipeline_run = False
            self._run_skips_mask_generation = False
            self._run_includes_analysis = False
            self._last_scene_plan = None
            self._refresh_patient_list()
            return
        if self._queued_commands and exit_code == 0:
            next_cmd = self._queued_commands.pop(0)
            self._active_stage = self._queued_stages.pop(0) if self._queued_stages else None
            self._run(next_cmd)
            return
        self._queued_commands = []
        self._queued_stages = []
        if self._is_full_pipeline_run and int(exit_code) == 0:
            for s in ("registration", "analysis"):
                self._set_stage_status(s, "done")
        elif self._run_includes_analysis and int(exit_code) == 0:
            # Timelapse pipeline commands internally perform discovery/import
            # and analysis, so mark all stages complete for clear 100% progress.
            for s in ("dataset", "parse", "registration", "analysis"):
                self._set_stage_status(s, "done")
        self._active_stage = None
        scene_plan = self._last_scene_plan
        self._is_full_pipeline_run = False
        self._run_skips_mask_generation = True
        self._run_includes_analysis = False
        self._last_scene_plan = None
        self.logic.cleanup_temp_files(remove_fallback=False)
        if scene_plan is not None:
            self._adopt_scene_run_as_current_dataset(scene_plan)
            self._load_scene_run_outputs(scene_plan)
            for s in ("dataset", "parse", "registration", "analysis"):
                self._set_stage_status(s, "done")
            self._set_scene_stage_message("Current step: complete")
        else:
            self._refresh_patient_list()
        self._set_user_message("success", "Completed", "Requested step(s) finished successfully.")

    def _on_install_pipeline(self):
        self._show("[dependency] Installing/updating timelapsed-hrpqct and contour dependencies ...")
        try:
            slicer.app.setOverrideCursor(qt.Qt.WaitCursor)
            self.logic.install_or_update_pipeline()
            self._show("[dependency] Installation finished.")
            slicer.util.infoDisplay(
                "timelapsed-hrpqct and contour dependency installation finished.\\n"
                "If import problems persist, restart Slicer."
            )
        except Exception as exc:
            slicer.util.errorDisplay(f"Install failed: {exc}")
        finally:
            slicer.app.restoreOverrideCursor()
            self._update_dependency_ui()

    def _on_check_pipeline(self):
        self._update_dependency_ui()

    def _dependency_status_text(self, detail):
        detail = str(detail or "")
        if detail.startswith("Installed ("):
            return detail.split(" from ", 1)[0]
        if detail.startswith("Out of date"):
            return detail.split(". Imported from ", 1)[0]
        if detail.startswith("Not installed"):
            return detail[:120] + "..." if len(detail) > 120 else detail
        return detail[:80] + "..." if len(detail) > 80 else detail

    def _update_dependency_ui(self):
        available, detail = self.logic.pipeline_status()
        self.pipelineStatusLabel.text = self._dependency_status_text(detail)
        self.pipelineStatusLabel.toolTip = str(detail or "")
        if available:
            self.pipelineStatusLabel.styleSheet = "color: #228b22;"
            if hasattr(self, "dependencyBox"):
                self.dependencyBox.collapsed = True
        else:
            self.pipelineStatusLabel.styleSheet = "color: #cc5500;"
            if hasattr(self, "dependencyBox"):
                self.dependencyBox.collapsed = False

    def _refresh_patient_list(self):
        if not self._qt_object_alive(getattr(self, "patientCombo", None)):
            return
        self.patientCombo.clear()
        self._patient_keys = []

        imported = self._imported_dataset_root()
        keys = set()
        if imported is not None and imported.exists():
            try:
                from timelapsedhrpqct.dataset.artifacts import (
                    iter_filled_session_records,
                    iter_fused_session_records,
                    iter_imported_stack_records,
                )

                for rec in iter_imported_stack_records(imported):
                    keys.add((str(rec.subject_id), str(rec.site)))
                for rec in iter_fused_session_records(imported):
                    keys.add((str(rec.subject_id), str(rec.site)))
                for rec in iter_filled_session_records(imported):
                    keys.add((str(rec.subject_id), str(rec.site)))
            except Exception as exc:
                self._show(f"[patients] artifact lookup failed: {exc}")

        if not keys and self._last_parsed_sessions:
            for s in self._last_parsed_sessions:
                subject = str(getattr(s, "subject_id", "")).strip()
                site = str(getattr(s, "site", "")).strip() or "radius"
                if subject:
                    keys.add((subject, site))

        self._patient_keys = sorted(keys)
        for subject, site in self._patient_keys:
            self.patientCombo.addItem(f"sub-{subject} | site-{site}")
        self._refresh_remodelling_comparison_list()

    def _refresh_remodelling_comparison_list(self):
        if not self._qt_object_alive(getattr(self, "remodellingComparisonCombo", None)):
            return
        self.remodellingComparisonCombo.clear()
        self._remodelling_comparison_items = []

        is_remodelling_load = (
            hasattr(self, "loadTypeCombo")
            and self.loadTypeCombo.currentText == "remodelling image"
        )
        self.remodellingComparisonCombo.enabled = bool(is_remodelling_load)
        if not is_remodelling_load:
            self.remodellingComparisonCombo.addItem("Not used for this data type")
            self._rebuild_series_summary_pair_selector([])
            return

        patient_key = self._current_patient_key()
        imported = self._imported_dataset_root()
        if patient_key is None or imported is None or not imported.exists():
            self.remodellingComparisonCombo.addItem("No remodelling comparisons found")
            self._rebuild_series_summary_pair_selector([])
            return

        subject_id, site = patient_key
        candidates = []
        try:
            from timelapsedhrpqct.dataset.derivative_paths import analysis_visualize_dir

            viz_dir = analysis_visualize_dir(imported, subject_id, site)
        except Exception:
            viz_dir = (
                imported
                / f"sub-{subject_id}"
                / f"site-{site}"
                / "analysis"
                / "visualize"
            )

        if viz_dir.exists():
            candidates = sorted(viz_dir.glob("*_remodelling.nii.gz")) + sorted(
                viz_dir.glob("*_remodelling.mha")
            )

        remodelling_candidates = []
        for path in candidates:
            ctx = self._parse_remodelling_source_context(path)
            if ctx is None:
                continue
            remodelling_candidates.append((ctx, Path(path)))

        def remodelling_sort_key(item):
            ctx, path = item
            return (
                str(ctx.get("t0", "")),
                str(ctx.get("t1", "")),
                0 if str(ctx.get("compartment", "")).strip().lower() == "full" else 1,
                str(ctx.get("compartment", "")),
                str(path),
            )

        for ctx, path in sorted(remodelling_candidates, key=remodelling_sort_key):
            label = f"{ctx['t0']} -> {ctx['t1']} ({self._scene_display_compartment_name(ctx.get('compartment', 'full'))})"
            self._remodelling_comparison_items.append((label, Path(path)))
            self.remodellingComparisonCombo.addItem(label)

        if not self._remodelling_comparison_items:
            self.remodellingComparisonCombo.addItem("No remodelling comparisons found")
            self.remodellingComparisonCombo.enabled = False
            self._rebuild_series_summary_pair_selector([])
            return

        self._rebuild_series_summary_pair_selector(self._adjacent_pairs_from_remodelling_items())

    def _current_remodelling_comparison_path(self):
        idx = int(self.remodellingComparisonCombo.currentIndex)
        if idx < 0 or idx >= len(self._remodelling_comparison_items):
            return None
        return self._remodelling_comparison_items[idx][1]

    def _adjacent_pairs_from_remodelling_items(self):
        ordered_ids = []
        seen = set()
        for _label, path in self._remodelling_comparison_items:
            ctx = self._parse_remodelling_source_context(path)
            if ctx is None:
                continue
            t0 = str(ctx["t0"])
            t1 = str(ctx["t1"])
            if t0 not in seen:
                ordered_ids.append(t0)
                seen.add(t0)
            if t1 not in seen:
                ordered_ids.append(t1)
                seen.add(t1)
        return ordered_ids

    def _current_patient_key(self):
        if not self._qt_object_alive(getattr(self, "patientCombo", None)):
            return None
        idx = int(self.patientCombo.currentIndex)
        if idx < 0 or idx >= len(self._patient_keys):
            return None
        return self._patient_keys[idx]

    def _subject_hierarchy(self):
        return slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)

    def _install_subject_hierarchy_selection_hook(self):
        if self._sh_tree_hooks_installed:
            return
        main = slicer.util.mainWindow()
        if main is None:
            return

        try:
            trees = slicer.util.findChildren(main, className="qMRMLSubjectHierarchyTreeView")
        except Exception:
            trees = []

        hooked = False
        for tree in trees:
            if tree.property("timelapsedHooked"):
                continue
            connected = False
            try:
                tree.currentItemChanged.connect(self._on_subject_hierarchy_item_selected)
                connected = True
            except Exception:
                pass
            if not connected:
                try:
                    tree.connect("currentItemChanged(vtkIdType)", self._on_subject_hierarchy_item_selected)
                    connected = True
                except Exception:
                    pass
            if connected:
                tree.setProperty("timelapsedHooked", True)
                hooked = True

        if hooked:
            self._sh_tree_hooks_installed = True
            self._show("[load] Subject hierarchy stack-selection hook enabled.")
        else:
            qt.QTimer.singleShot(1500, self._install_subject_hierarchy_selection_hook)

    def _first_scalar_volume_under_item(self, item_id):
        sh = self._subject_hierarchy()
        if sh is None:
            return None

        node = sh.GetItemDataNode(item_id)
        if node is not None and node.IsA("vtkMRMLScalarVolumeNode"):
            return node

        child_ids = vtk.vtkIdList()
        try:
            sh.GetItemChildren(item_id, child_ids, True)
        except TypeError:
            try:
                sh.GetItemChildren(item_id, child_ids)
            except Exception:
                return None
        except Exception:
            return None

        for i in range(child_ids.GetNumberOfIds()):
            child_id = child_ids.GetId(i)
            child_node = sh.GetItemDataNode(child_id)
            if child_node is not None and child_node.IsA("vtkMRMLScalarVolumeNode"):
                return child_node
        return None

    def _on_subject_hierarchy_item_selected(self, item_id):
        # Disabled for now: keep subject hierarchy interaction passive/predictable.
        return

    def _ensure_folder_item(self, parent_item_id, name):
        sh = self._subject_hierarchy()
        child_id = sh.GetItemChildWithName(parent_item_id, name)
        if child_id:
            return child_id
        return sh.CreateFolderItem(parent_item_id, name)

    def _ensure_load_folder(self, subject_id, site, session_id=None, stack_index=None):
        sh = self._subject_hierarchy()
        scene_id = sh.GetSceneItemID()
        root_id = self._ensure_folder_item(scene_id, "TimelapsedHRpQCT Loaded")
        subj_id = self._ensure_folder_item(root_id, f"sub-{subject_id}")
        site_id = self._ensure_folder_item(subj_id, f"site-{site}")
        if session_id is None:
            return site_id
        ses_id = self._ensure_folder_item(site_id, f"ses-{session_id}")
        if stack_index is None:
            return ses_id
        return self._ensure_folder_item(ses_id, f"stack-{int(stack_index):02d}")

    def _set_item_visibility_safe(self, item_id, visible):
        sh = self._subject_hierarchy()
        if sh is None:
            return
        try:
            sh.SetItemDisplayVisibility(int(item_id), 1 if bool(visible) else 0)
        except Exception:
            pass

    def _collect_stack_items(self, subject_id=None, site=None):
        sh = self._subject_hierarchy()
        if sh is None:
            return []
        scene_id = sh.GetSceneItemID()
        root_id = sh.GetItemChildWithName(scene_id, "TimelapsedHRpQCT Loaded")
        if not root_id:
            return []

        base_id = root_id
        if subject_id is not None:
            subj_id = sh.GetItemChildWithName(base_id, f"sub-{subject_id}")
            if not subj_id:
                return []
            base_id = subj_id
        if site is not None:
            site_id = sh.GetItemChildWithName(base_id, f"site-{site}")
            if not site_id:
                return []
            base_id = site_id

        child_ids = vtk.vtkIdList()
        try:
            sh.GetItemChildren(base_id, child_ids, True)
        except TypeError:
            sh.GetItemChildren(base_id, child_ids)
        except Exception:
            return []

        stack_items = []
        for i in range(child_ids.GetNumberOfIds()):
            cid = child_ids.GetId(i)
            name = str(sh.GetItemName(cid) or "").lower()
            if name.startswith("stack-"):
                stack_items.append(cid)
        return stack_items

    def _stack_sort_key(self, item_id):
        sh = self._subject_hierarchy()
        if sh is None:
            return (9999, 9999)
        name = str(sh.GetItemName(item_id) or "")
        stack_match = re.search(r"stack-(\d+)", name, flags=re.IGNORECASE)
        stack_num = int(stack_match.group(1)) if stack_match else 9999
        parent_id = sh.GetItemParent(item_id)
        parent_name = str(sh.GetItemName(parent_id) or "")
        ses_match = re.search(r"ses-([A-Za-z]*)(\d+)", parent_name)
        ses_num = int(ses_match.group(2)) if ses_match else 9999
        return (ses_num, stack_num)

    def _set_exclusive_stack_visibility(self, active_stack_item_id, subject_id=None, site=None):
        stack_items = self._collect_stack_items(subject_id=subject_id, site=site)
        if not stack_items:
            return
        for sid in stack_items:
            self._set_item_visibility_safe(sid, int(sid) == int(active_stack_item_id))

    def _apply_default_stack_visibility(self, subject_id, site):
        stack_items = self._collect_stack_items(subject_id=subject_id, site=site)
        if not stack_items:
            return
        first_stack = sorted(stack_items, key=self._stack_sort_key)[0]
        self._set_exclusive_stack_visibility(first_stack, subject_id=subject_id, site=site)
        volume_node = self._first_scalar_volume_under_item(first_stack)
        if volume_node is not None:
            slicer.util.setSliceViewerLayers(background=volume_node, fit=False)

    def _place_node_in_folder(self, node, folder_item_id):
        if node is None:
            return
        sh = self._subject_hierarchy()
        item_id = sh.GetItemByDataNode(node)
        if item_id:
            sh.SetItemParent(item_id, folder_item_id)

    def _clear_loaded_review_nodes(self):
        sh = self._subject_hierarchy()
        if sh is None:
            return
        scene_id = sh.GetSceneItemID()
        root_id = sh.GetItemChildWithName(scene_id, "TimelapsedHRpQCT Loaded")
        if not root_id:
            self._interactive_preview_cache = {}
            return

        child_ids = vtk.vtkIdList()
        try:
            sh.GetItemChildren(root_id, child_ids, True)
        except TypeError:
            sh.GetItemChildren(root_id, child_ids)
        except Exception:
            child_ids = vtk.vtkIdList()

        nodes_to_remove = []
        for i in range(child_ids.GetNumberOfIds()):
            item_id = child_ids.GetId(i)
            node = sh.GetItemDataNode(item_id)
            if node is not None:
                nodes_to_remove.append(node)

        for node in nodes_to_remove:
            try:
                slicer.mrmlScene.RemoveNode(node)
            except Exception:
                pass

        try:
            sh.RemoveItem(root_id)
        except Exception:
            pass

        self._interactive_preview_cache = {}
        self._rebuild_series_summary_pair_selector([])
        self._set_pair_metric_labels(None, None)
        self._set_series_summary_labels(None)

    def _on_clear_loaded_timelapsed_results(self):
        if self.logic.is_running():
            slicer.util.warningDisplay("Stop the running Timelapsed pipeline before clearing loaded results.")
            return
        self._clear_loaded_review_nodes()
        self._remove_scene_run_nonlinear_transform_nodes()
        self._interactive_preview_cache = {}
        self._last_scene_results_plan = None
        if hasattr(self, "sceneStatusLabel"):
            self.sceneStatusLabel.text = "Cleared loaded Timelapsed results and preview cache."
        try:
            self._set_scene_comparison_rows([])
        except Exception:
            pass
        gc.collect()
        self._show("[scene] cleared loaded Timelapsed result nodes and preview cache.")

    def _session_base_color(self, session_id):
        token = str(session_id).upper()
        if token in {"T1", "BL", "BASELINE"}:
            return (0.95, 0.35, 0.35)
        if token in {"T2", "FL", "FU", "FOLLOWUP", "FOLLOWUP1"}:
            return (0.20, 0.65, 0.95)
        if token in {"T3", "FU2", "FOLLOWUP2"}:
            return (0.25, 0.75, 0.40)

        m = re.search(r"(\d+)$", token)
        if m:
            idx = int(m.group(1))
            palette = [
                (0.95, 0.35, 0.35),
                (0.20, 0.65, 0.95),
                (0.25, 0.75, 0.40),
                (0.95, 0.75, 0.25),
                (0.65, 0.45, 0.90),
            ]
            return palette[(max(idx, 1) - 1) % len(palette)]
        return (0.85, 0.85, 0.30)

    def _create_scalar_node_from_array(self, name, array_zyx, spacing_xyz, origin_xyz):
        node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLScalarVolumeNode", name)
        slicer.util.updateVolumeFromArray(node, array_zyx)
        node.SetSpacing(float(spacing_xyz[0]), float(spacing_xyz[1]), float(spacing_xyz[2]))
        node.SetOrigin(float(origin_xyz[0]), float(origin_xyz[1]), float(origin_xyz[2]))
        return node

    def _copy_volume_geometry(self, source_node, target_node):
        if source_node is None or target_node is None:
            return
        try:
            matrix = vtk.vtkMatrix4x4()
            source_node.GetIJKToRASMatrix(matrix)
            target_node.SetIJKToRASMatrix(matrix)
            return
        except Exception:
            pass
        try:
            target_node.SetSpacing(*source_node.GetSpacing())
            target_node.SetOrigin(*source_node.GetOrigin())
        except Exception:
            pass

    def _configure_segmentation_display(self, seg_node):
        if seg_node is None:
            return
        display = seg_node.GetDisplayNode()
        if display is None:
            seg_node.CreateDefaultDisplayNodes()
            display = seg_node.GetDisplayNode()
        if display is None:
            return
        display.SetVisibility(True)
        display.SetVisibility2D(True)
        display.SetVisibility3D(False)
        display.SetOpacity2DFill(0.35)
        display.SetOpacity2DOutline(0.0)
        display.SetSliceIntersectionThickness(2)
        # Ensure segmentations are shown in all slice views (not bound to a specific view).
        if hasattr(display, "RemoveAllViewNodeIDs"):
            display.RemoveAllViewNodeIDs()

        segmentation = seg_node.GetSegmentation()
        if segmentation is not None:
            segment_ids = vtk.vtkStringArray()
            segmentation.GetSegmentIDs(segment_ids)
            # Some Slicer versions support all-at-once visibility toggles.
            if hasattr(display, "SetAllSegmentsVisibility"):
                display.SetAllSegmentsVisibility(True)
            # Also force each segment visible for compatibility.
            if hasattr(display, "SetSegmentVisibility"):
                for i in range(segment_ids.GetNumberOfValues()):
                    seg_id = segment_ids.GetValue(i)
                    display.SetSegmentVisibility(seg_id, True)

    def _remodelling_color_node(self):
        name = "TimelapsedHRpQCT_RemodellingColors"
        existing = slicer.mrmlScene.GetFirstNodeByName(name)
        if existing is not None:
            return existing
        color_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLColorTableNode", name)
        if hasattr(color_node, "SetTypeToUser"):
            color_node.SetTypeToUser()
        if hasattr(color_node, "SetNumberOfColors"):
            color_node.SetNumberOfColors(6)
        colors = {
            0: ("background", 0.0, 0.0, 0.0, 0.0),
            1: ("resorption", 1.00, 0.05, 0.70, 1.0),
            2: ("quiescent", 0.62, 0.62, 0.62, 0.32),
            3: ("formation", 1.00, 0.48, 0.00, 1.0),
            4: ("formation", 1.00, 0.48, 0.00, 1.0),
            5: ("quiescent", 0.62, 0.62, 0.62, 0.32),
        }
        for value, color in colors.items():
            try:
                color_node.SetColor(int(value), color[0], float(color[1]), float(color[2]), float(color[3]), float(color[4]))
            except Exception:
                pass
        try:
            color_node.HideFromEditorsOn()
        except Exception:
            pass
        return color_node

    def _create_labelmap_from_label_array(
        self,
        name,
        label_arr_zyx,
        spacing_xyz,
        origin_xyz,
        folder_item_id=None,
        activate_display=True,
    ):
        node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode", name)
        slicer.util.updateVolumeFromArray(node, label_arr_zyx.astype(np.uint8, copy=False))
        node.SetSpacing(float(spacing_xyz[0]), float(spacing_xyz[1]), float(spacing_xyz[2]))
        node.SetOrigin(float(origin_xyz[0]), float(origin_xyz[1]), float(origin_xyz[2]))
        node.CreateDefaultDisplayNodes()
        self._style_remodelling_labelmap(node, activate_display=activate_display)
        if folder_item_id is not None:
            self._place_node_in_folder(node, folder_item_id)
        return node

    def _create_remodelling_scalar_node_from_label_array(
        self,
        name,
        label_arr_zyx,
        spacing_xyz,
        origin_xyz,
        folder_item_id=None,
        activate_display=True,
    ):
        node = self._create_scalar_node_from_array(
            name,
            label_arr_zyx.astype(np.uint8, copy=False),
            spacing_xyz,
            origin_xyz,
        )
        node.CreateDefaultDisplayNodes()
        self._style_remodelling_scalar_volume(node, activate_display=activate_display)
        if folder_item_id is not None:
            self._place_node_in_folder(node, folder_item_id)
        return node

    def _style_remodelling_scalar_volume(self, volume_node, *, activate_display=True):
        if volume_node is None:
            return
        display = volume_node.GetDisplayNode()
        if display is None:
            volume_node.CreateDefaultDisplayNodes()
            display = volume_node.GetDisplayNode()
        if display is None:
            return
        color_node = self._remodelling_color_node()
        if color_node is not None and hasattr(display, "SetAndObserveColorNodeID"):
            display.SetAndObserveColorNodeID(color_node.GetID())
        display.SetVisibility(True)
        if hasattr(display, "SetOpacity"):
            display.SetOpacity(1.0)
        if hasattr(display, "SetInterpolate"):
            display.SetInterpolate(False)
        elif hasattr(display, "InterpolateOff"):
            display.InterpolateOff()
        if hasattr(display, "AutoWindowLevelOff"):
            display.AutoWindowLevelOff()
        if hasattr(display, "SetWindowLevel"):
            display.SetWindowLevel(5.0, 2.5)
        elif hasattr(display, "SetWindow") and hasattr(display, "SetLevel"):
            display.SetWindow(5.0)
            display.SetLevel(2.5)
        if activate_display:
            try:
                slicer.util.setSliceViewerLayers(foreground=volume_node, foregroundOpacity=0.65, fit=False)
            except Exception:
                pass

    def _style_remodelling_labelmap(self, label_node, *, activate_display=True):
        self._style_remodelling_scalar_volume(label_node, activate_display=activate_display)

    def _on_interactive_preview_control_changed(self, *_args):
        if getattr(self, "_suppress_interactive_preview_updates", False):
            return
        if self.logic.is_running():
            return
        if not self.remodellingAutoUpdateCheck.checked:
            self._mark_analysis_settings_dirty()
            return
        self._interactivePreviewTimer.start()

    def _current_analysis_pair_mode(self):
        pair_mode = str(self.analysisPairModeCombo.currentData or "adjacent")
        return pair_mode if pair_mode in {"adjacent", "baseline", "all_pairs"} else "adjacent"

    def _on_analysis_pair_mode_changed(self, *_args):
        if getattr(self, "_suppress_interactive_preview_updates", False):
            return
        self._mark_analysis_settings_dirty()

    def _interactive_preview_label_map(self):
        return {
            "resorption": 1,
            "demineralisation": 2,
            "quiescent": 2,
            "formation": 3,
            "mineralisation": 2,
        }

    def _current_comparison_table_row(self, metric_row):
        formation = metric_row.get("formation_frac_bv0")
        resorption = metric_row.get("resorption_frac_bv0")
        try:
            activity = float(formation) + float(resorption)
        except Exception:
            activity = ""
        try:
            net = float(formation) - float(resorption)
        except Exception:
            net = ""
        return [
            self._scene_display_compartment_name(metric_row.get("compartment", "full")),
            self._format_scene_result_fraction(formation),
            self._format_scene_result_fraction(resorption),
            self._format_scene_result_fraction(activity),
            self._format_scene_result_fraction(net),
        ]

    def _update_current_comparison_table(self, rows=None):
        if not hasattr(self, "currentComparisonTable"):
            return
        seen_row_keys = set()
        normalized_rows = []
        for row in list(rows or []):
            row_key = str(row.get("compartment", "")) if isinstance(row, dict) else str(row)
            if row_key in seen_row_keys:
                continue
            seen_row_keys.add(row_key)
            normalized_rows.append(row)
        display_rows = [self._current_comparison_table_row(row) for row in normalized_rows]
        if not display_rows:
            display_rows = [["N/A", "N/A", "N/A", "N/A", "N/A"]]
        self.currentComparisonTable.clearContents()
        self.currentComparisonTable.setRowCount(len(display_rows))
        for row_idx, values in enumerate(display_rows):
            for col_idx, value in enumerate(values):
                item = qt.QTableWidgetItem(str(value))
                self.currentComparisonTable.setItem(row_idx, col_idx, item)
        try:
            self.currentComparisonTable.resizeColumnsToContents()
            self.currentComparisonTable.horizontalHeader().setStretchLastSection(True)
        except Exception:
            pass

    def _set_pair_metric_labels(self, formation_frac=None, resorption_frac=None, compartment="full"):
        rows = []
        if formation_frac is not None or resorption_frac is not None:
            rows.append(
                {
                    "compartment": str(compartment),
                    "formation_frac_bv0": formation_frac,
                    "resorption_frac_bv0": resorption_frac,
                }
            )
        self._latest_pair_metric_rows = list(rows)
        self._update_current_comparison_table(rows)

    def _set_pair_metric_rows(self, rows=None):
        normalized_rows = list(rows or [])
        self._latest_pair_metric_rows = normalized_rows
        self._update_current_comparison_table(normalized_rows)

    def _csv_float_or_nan(self, value):
        text = str(value if value is not None else "").strip()
        if not text:
            return float("nan")
        return float(text)

    def _metric_rows_have_finite_fractions(self, rows):
        for row in list(rows or []):
            for key in ("formation_frac_bv0", "resorption_frac_bv0"):
                try:
                    if np.isfinite(float(row.get(key))):
                        return True
                except Exception:
                    continue
        return False

    def _saved_pair_metric_rows_for_context(self, ctx):
        if not ctx:
            return []
        imported = self._imported_dataset_root()
        if imported is None:
            return []
        subject_id = str(ctx.get("subject_id") or "").strip()
        site = str(ctx.get("site") or "").strip()
        t0 = str(ctx.get("t0") or "").strip()
        t1 = str(ctx.get("t1") or "").strip()
        if not subject_id or not site or not t0 or not t1:
            return []
        try:
            from timelapsedhrpqct.dataset.derivative_paths import pairwise_remodelling_csv_path

            pairwise_path = pairwise_remodelling_csv_path(imported, subject_id, site)
        except Exception as exc:
            self._show(f"[load] could not resolve saved remodelling metrics: {exc}")
            return []
        if not pairwise_path.exists():
            return []

        rows = []
        try:
            with pairwise_path.open("r", encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if str(row.get("t0") or "").strip() != t0:
                        continue
                    if str(row.get("t1") or "").strip() != t1:
                        continue
                    compartment = str(row.get("compartment") or "").strip()
                    if not compartment:
                        continue
                    rows.append(
                        {
                            "compartment": compartment,
                            "formation_frac_bv0": self._csv_float_or_nan(row.get("formation_frac_bv0")),
                            "resorption_frac_bv0": self._csv_float_or_nan(row.get("resorption_frac_bv0")),
                        }
                    )
        except Exception as exc:
            self._show(f"[load] could not read saved remodelling metrics: {exc}")
            return []
        return rows

    def _scene_result_row_from_metric(self, ctx, metric_row):
        formation = metric_row.get("formation_frac_bv0")
        resorption = metric_row.get("resorption_frac_bv0")
        try:
            activity = float(formation) + float(resorption)
        except Exception:
            activity = ""
        try:
            net = float(formation) - float(resorption)
        except Exception:
            net = ""
        return [
            f"{str(ctx.get('t0', '')).strip()} -> {str(ctx.get('t1', '')).strip()}",
            self._scene_display_compartment_name(metric_row.get("compartment", "full")),
            self._format_scene_result_fraction(formation),
            self._format_scene_result_fraction(resorption),
            self._format_scene_result_fraction(activity),
            self._format_scene_result_fraction(net),
        ]

    def _set_series_summary_labels(self, summary=None):
        self._latest_series_summary = summary

    def _set_series_summary_saved_state(self, text):
        return

    def _current_saved_analysis_matches_preview(self, metadata):
        if not metadata:
            return False
        saved_method = str(metadata.get("method", ""))
        saved_source = str(metadata.get("change_region_source", "")).strip().lower()
        saved_binary = metadata.get("binary_reclassification_enabled", None)
        saved_thresholds = metadata.get("thresholds") or []
        saved_clusters = metadata.get("cluster_sizes") or []
        return (
            saved_method == self._current_analysis_method()
            and (not saved_source or saved_source == ("bone_union" if self.analysisRestrictBoneSupportCheck.checked else "common_mask"))
            and (saved_binary is None or bool(saved_binary) == bool(self.analysisBinaryReclassificationCheck.checked))
            and saved_thresholds[:1] == [float(self.analysisThreshold.value)]
            and saved_clusters[:1] == [int(self.analysisCluster.value)]
            and bool(metadata.get("gaussian_filter", False)) == bool(self.analysisGaussianFilterCheck.checked)
            and float(metadata.get("gaussian_sigma", 0.0)) == float(self.analysisGaussianSigma.value)
            and int(metadata.get("full_mask_dilation_voxels", 2)) == int(self.analysisFullMaskDilation.value)
            and int(metadata.get("marrow_mask_dilation_voxels", 2)) == int(self.analysisMarrowMaskDilation.value)
            and int(metadata.get("marrow_mask_erosion_voxels", 0)) == int(self.analysisMarrowMaskErosion.value)
        )

    def _selected_series_adjacent_pairs(self):
        selected = []
        for key, item in sorted(self._series_summary_pair_checks.items()):
            try:
                state = item.checkState()
            except Exception:
                state = qt.Qt.Unchecked
            if state == qt.Qt.Checked:
                selected.append(key)
        return selected

    def _rebuild_series_summary_pair_selector(self, session_ids):
        self._series_summary_pair_checks = {}

    def _apply_preview_label_filters(self, label_arr_zyx, valid_mask_zyx=None):
        original = np.asarray(label_arr_zyx)
        arr = original.copy()
        if valid_mask_zyx is not None:
            valid = np.asarray(valid_mask_zyx, dtype=bool)
            if valid.shape == arr.shape and np.any(valid):
                arr[~valid] = 0
        # Collapse legacy 5-label remodelling images into the default 3-label display.
        if np.any(arr == 4) or np.any(arr == 5):
            arr[arr == 2] = 2
            arr[arr == 3] = 2
            arr[arr == 4] = 3
            arr[arr == 5] = 2
        return arr

    def _get_valid_mask_for_source(self, source_path):
        if not source_path:
            return None
        cache_key = str(Path(source_path).resolve())
        cached = self._interactive_preview_cache.get(cache_key)
        if cached is not None:
            return self._display_valid_mask_for_preview_inputs(cached)
        try:
            preview_inputs = self._get_interactive_preview_inputs(source_path)
            return self._display_valid_mask_for_preview_inputs(preview_inputs)
        except Exception:
            return None

    def _display_valid_mask_for_preview_inputs(self, preview_inputs):
        valid_mask = preview_inputs.get("valid_mask")
        if str(preview_inputs.get("context", {}).get("compartment", "")) != "full":
            return valid_mask

        support0 = preview_inputs.get("support_mask_t0")
        support1 = preview_inputs.get("support_mask_t1")
        if support0 is None or support1 is None:
            return valid_mask

        try:
            from timelapsedhrpqct.analysis import dilate_mask_xy, erode_mask

            support0 = np.asarray(support0, dtype=bool)
            support1 = np.asarray(support1, dtype=bool)
            if valid_mask is not None and (
                support0.shape != np.asarray(valid_mask).shape
                or support1.shape != np.asarray(valid_mask).shape
            ):
                return valid_mask
            if int(self.analysisFullMaskDilation.value) > 0:
                support0 = dilate_mask_xy(support0, int(self.analysisFullMaskDilation.value))
                support1 = dilate_mask_xy(support1, int(self.analysisFullMaskDilation.value))
            return erode_mask(support0 & support1, int(self._analysis_erosion_voxels))
        except Exception:
            return valid_mask

    def _infer_results_root_from_path(self, path_obj):
        p = Path(path_obj).resolve()
        for candidate in [p] + list(p.parents):
            if candidate.name == "Timelapse":
                if candidate.parent.name == "derivatives":
                    return candidate.parent.parent
                return candidate
        return None

    def _parse_remodelling_source_context(self, source_path):
        name = Path(source_path).name
        patterns = [
            re.compile(
                r"^sub-(?P<subject_id>.+?)_voi-(?P<site>.+?)_desc-(?P<compartment>.+?)_"
                r"t0-(?P<t0>.+?)_t1-(?P<t1>.+?)_thr-(?P<threshold>.+?)_cluster-(?P<cluster>\d+)_remodelling\.(?:nii\.gz|mha)$"
            ),
            re.compile(
                r"^sub-(?P<subject_id>.+?)_site-(?P<site>.+?)_comp-(?P<compartment>.+?)_"
                r"t0-(?P<t0>.+?)_t1-(?P<t1>.+?)_thr-(?P<threshold>.+?)_cluster-(?P<cluster>\d+)_remodelling\.(?:nii\.gz|mha)$"
            ),
            re.compile(
                r"^sub-(?P<subject_id>.+?)_comp-(?P<compartment>.+?)_"
                r"t0-(?P<t0>.+?)_t1-(?P<t1>.+?)_thr-(?P<threshold>.+?)_cluster-(?P<cluster>\d+)_remodelling\.(?:nii\.gz|mha)$"
            ),
        ]
        for pattern in patterns:
            match = pattern.match(name)
            if match is None:
                continue
            data = match.groupdict()
            data.setdefault("site", "radius")
            compact_sites = {
                "radius_left": "radiusleft",
                "radiusright": "radiusright",
                "radius_right": "radiusright",
                "tibialeft": "tibialeft",
                "tibia_left": "tibialeft",
                "tibiaright": "tibiaright",
                "tibia_right": "tibiaright",
                "kneeleft": "kneeleft",
                "knee_left": "kneeleft",
                "kneeright": "kneeright",
                "knee_right": "kneeright",
            }
            data["site"] = compact_sites.get(str(data["site"]).lower(), str(data["site"]).lower())
            data["threshold"] = float(str(data["threshold"]).replace("p", "."))
            data["cluster"] = int(data["cluster"])
            data["source_path"] = str(Path(source_path).resolve())
            return data
        return None

    def _load_support_mask_array(self, mask_paths, reference_image_path):
        def _read_bool(path_obj):
            return (sitk.GetArrayFromImage(sitk.ReadImage(str(path_obj))) > 0).astype(bool, copy=False)

        if "full" in mask_paths and Path(mask_paths["full"]).exists():
            return _read_bool(mask_paths["full"])
        if "regmask" in mask_paths and Path(mask_paths["regmask"]).exists():
            return _read_bool(mask_paths["regmask"])

        roi_paths = [
            Path(path)
            for role, path in sorted((mask_paths or {}).items())
            if str(role).lower().startswith("roi") and Path(path).exists()
        ]
        if roi_paths:
            union = None
            for roi_path in roi_paths:
                arr = _read_bool(roi_path)
                union = arr if union is None else (union | arr)
            if union is not None:
                return union

        if (
            "trab" in mask_paths
            and "cort" in mask_paths
            and Path(mask_paths["trab"]).exists()
            and Path(mask_paths["cort"]).exists()
        ):
            trab = _read_bool(mask_paths["trab"])
            cort = _read_bool(mask_paths["cort"])
            return trab | cort

        ref_arr = sitk.GetArrayFromImage(sitk.ReadImage(str(reference_image_path)))
        return np.zeros_like(ref_arr, dtype=bool)

    def _infer_stack_seg_path_from_image_path(self, image_path):
        path = Path(str(image_path))
        name = path.name
        replacements = [
            ("_image.nii.gz", "_seg.nii.gz"),
            ("_image.mha", "_seg.mha"),
            ("_image.nrrd", "_seg.nrrd"),
        ]
        for old, new in replacements:
            if name.endswith(old):
                candidate = path.with_name(name[: -len(old)] + new)
                if candidate.exists():
                    return candidate
        return None

    def _fused_metadata_path_from_image_path(self, image_path):
        path = Path(str(image_path))
        name = path.name
        replacements = [
            ("_image_fused.nii.gz", "_fused.json"),
            ("_image_fused.mha", "_fused.json"),
            ("_image_fused.nrrd", "_fused.json"),
        ]
        for old, new in replacements:
            if name.endswith(old):
                candidate = path.with_name(name[: -len(old)] + new)
                if candidate.exists():
                    return candidate
        return None

    def _read_seg_array_for_preview(self, session, reference_image):
        def _read_nonempty_seg(path):
            if path is None or not Path(path).exists():
                return None
            seg_img = sitk.ReadImage(str(path))
            seg_arr = (sitk.GetArrayFromImage(seg_img) > 0).astype(bool, copy=False)
            if np.any(seg_arr):
                return seg_arr
            return None

        seg_path = getattr(session, "seg_path", None)
        direct_seg = _read_nonempty_seg(seg_path)
        if direct_seg is not None:
            return direct_seg

        metadata_path = getattr(session, "metadata_path", None)
        if metadata_path is None or not Path(metadata_path).exists():
            metadata_path = self._fused_metadata_path_from_image_path(getattr(session, "image_path", ""))
        if metadata_path is None:
            return None
        try:
            metadata = json.loads(Path(metadata_path).read_text(encoding="utf-8"))
        except Exception:
            return None

        seg_union = None
        for contributor in metadata.get("contributors", []) or []:
            stack_seg_path = contributor.get("seg_path")
            if stack_seg_path:
                stack_seg_path = Path(str(stack_seg_path))
            else:
                stack_seg_path = self._infer_stack_seg_path_from_image_path(contributor.get("image_path", ""))
            if stack_seg_path is None or not Path(stack_seg_path).exists():
                continue

            transform_path = contributor.get("transform_source")
            if transform_path and Path(str(transform_path)).exists():
                transform = sitk.ReadTransform(str(transform_path))
            else:
                transform = sitk.Transform(3, sitk.sitkIdentity)

            seg_img = sitk.Cast(sitk.ReadImage(str(stack_seg_path)) > 0, sitk.sitkUInt8)
            seg_tx = sitk.Resample(
                seg_img,
                reference_image,
                transform,
                sitk.sitkNearestNeighbor,
                0,
                sitk.sitkUInt8,
            )
            if seg_union is None:
                seg_union = sitk.Cast(seg_tx > 0, sitk.sitkUInt8)
            else:
                seg_union = seg_union | sitk.Cast(seg_tx > 0, sitk.sitkUInt8)

        if seg_union is None:
            return None
        return (sitk.GetArrayFromImage(seg_union) > 0).astype(bool, copy=False)

    def _get_interactive_preview_inputs(self, source_path):
        cache_key = str(Path(source_path).resolve())
        cached = self._interactive_preview_cache.get(cache_key)
        if cached is not None:
            return cached

        ctx = self._parse_remodelling_source_context(source_path)
        if ctx is None:
            raise ValueError(
                f"Could not parse remodelling visualization filename for interactive preview: {Path(source_path).name}"
            )
        imported_root = self._infer_results_root_from_path(source_path)
        if imported_root is None:
            raise ValueError(f"Could not infer Timelapse root from {source_path}")

        from timelapsedhrpqct.analysis import build_series_common_masks
        from timelapsedhrpqct.processing.analysis_io import discover_analysis_sessions

        sessions = discover_analysis_sessions(
            imported_root,
            ctx["subject_id"],
            ctx["site"],
            use_filled_images=False,
            require_seg=False,
        )
        sessions_by_id = {str(s.session_id): s for s in sessions}
        t0 = sessions_by_id.get(str(ctx["t0"]))
        t1 = sessions_by_id.get(str(ctx["t1"]))
        if t0 is None or t1 is None:
            raise ValueError(
                f"Could not locate transformed sessions for t0={ctx['t0']} and t1={ctx['t1']} "
                f"in sub-{ctx['subject_id']} site-{ctx['site']}."
            )

        ref_img = sitk.ReadImage(str(t0.image_path))
        img_t0 = sitk.GetArrayFromImage(ref_img).astype(np.float32, copy=False)
        img_t1_ref = sitk.ReadImage(str(t1.image_path))
        img_t1 = sitk.GetArrayFromImage(img_t1_ref).astype(np.float32, copy=False)
        delta_zyx = (img_t1 - img_t0).astype(np.float32, copy=False)
        seg_t0 = self._read_seg_array_for_preview(t0, ref_img)
        seg_t1 = self._read_seg_array_for_preview(t1, img_t1_ref)

        support_t0 = self._load_support_mask_array(t0.mask_paths, t0.image_path)
        support_t1 = self._load_support_mask_array(t1.mask_paths, t1.image_path)
        compartment = str(ctx["compartment"])
        compartment_for_valid_mask = "full"
        if compartment == "roi_union":
            compartment_for_valid_mask = "full"
        elif compartment:
            compartment_for_valid_mask = compartment
        if compartment_for_valid_mask == "full":
            comp_t0 = support_t0
            comp_t1 = support_t1
        else:
            comp_path_t0 = t0.mask_paths.get(compartment_for_valid_mask)
            comp_path_t1 = t1.mask_paths.get(compartment_for_valid_mask)
            if comp_path_t0 is None or comp_path_t1 is None:
                raise ValueError(
                    f"Interactive preview requires mask role '{compartment_for_valid_mask}' in both selected sessions."
                )
            comp_t0 = (sitk.GetArrayFromImage(sitk.ReadImage(str(comp_path_t0))) > 0).astype(bool, copy=False)
            comp_t1 = (sitk.GetArrayFromImage(sitk.ReadImage(str(comp_path_t1))) > 0).astype(bool, copy=False)

        valid_mask = build_series_common_masks(
            {
                "full": [support_t0, support_t1],
                compartment_for_valid_mask: [comp_t0, comp_t1],
            },
            [compartment_for_valid_mask],
            int(self._analysis_erosion_voxels),
        )[compartment_for_valid_mask]

        cached = {
            "cache_key": cache_key,
            "context": ctx,
            "spacing_xyz": tuple(float(x) for x in ref_img.GetSpacing()),
            "origin_xyz": tuple(float(x) for x in ref_img.GetOrigin()),
            "image_arr_t0": img_t0,
            "image_arr_t1": img_t1,
            "delta_zyx": delta_zyx,
            "seg_arr_t0": seg_t0,
            "seg_arr_t1": seg_t1,
            "support_mask_t0": support_t0,
            "support_mask_t1": support_t1,
            "t0_mask_paths": {str(k): str(v) for k, v in (t0.mask_paths or {}).items()},
            "t1_mask_paths": {str(k): str(v) for k, v in (t1.mask_paths or {}).items()},
            "compartment_mask_cache": {},
            "valid_mask": valid_mask,
            "current_label_arr": None,
        }
        self._interactive_preview_cache[cache_key] = cached
        return cached

    def _pair_metric_compartments(self, preview_inputs):
        roles = ["full"]
        mask_paths_t0 = preview_inputs.get("t0_mask_paths") or {}
        mask_paths_t1 = preview_inputs.get("t1_mask_paths") or {}
        shared_roles = sorted(set(mask_paths_t0) & set(mask_paths_t1))
        for role in shared_roles:
            role_token = str(role).strip().lower()
            if role_token in {"", "full", "regmask"}:
                continue
            path_t0 = Path(str(mask_paths_t0.get(role, "")))
            path_t1 = Path(str(mask_paths_t1.get(role, "")))
            if path_t0.exists() and path_t1.exists():
                roles.append(str(role))
        return roles

    def _preview_compartment_masks(self, preview_inputs, compartment):
        if str(compartment) == "full":
            return (
                np.asarray(preview_inputs["support_mask_t0"], dtype=bool),
                np.asarray(preview_inputs["support_mask_t1"], dtype=bool),
            )
        cache = preview_inputs.setdefault("compartment_mask_cache", {})
        if compartment in cache:
            return cache[compartment]
        mask_paths_t0 = preview_inputs.get("t0_mask_paths") or {}
        mask_paths_t1 = preview_inputs.get("t1_mask_paths") or {}
        path_t0 = Path(str(mask_paths_t0.get(compartment, "")))
        path_t1 = Path(str(mask_paths_t1.get(compartment, "")))
        if not path_t0.exists() or not path_t1.exists():
            raise ValueError(f"Mask role '{compartment}' is not available for the current pair.")
        mask_t0 = (sitk.GetArrayFromImage(sitk.ReadImage(str(path_t0))) > 0).astype(bool, copy=False)
        mask_t1 = (sitk.GetArrayFromImage(sitk.ReadImage(str(path_t1))) > 0).astype(bool, copy=False)
        cache[compartment] = (mask_t0, mask_t1)
        return cache[compartment]

    def _compute_pair_remodelling_preview_compat(self, compute_pair_remodelling_preview, **kwargs):
        import inspect

        try:
            params = inspect.signature(compute_pair_remodelling_preview).parameters
        except Exception:
            params = {}
        if "marrow_mask_dilation_voxels" not in params:
            kwargs.pop("marrow_mask_dilation_voxels", None)
            if bool(self.analysisRestrictBoneSupportCheck.checked) and int(self.analysisMarrowMaskDilation.value) > 0:
                self._show(
                    "[preview] installed timelapsed-hrpqct does not support bone support dilation yet; "
                    "update the core package to apply this option."
                )
        return compute_pair_remodelling_preview(**kwargs)

    def _preview_delta_for_current_settings(self, preview_inputs):
        delta_cache = preview_inputs.setdefault("delta_cache", {})
        gaussian_enabled = bool(self.analysisGaussianFilterCheck.checked)
        sigma = float(self.analysisGaussianSigma.value)
        cache_key = ("gaussian", round(sigma, 6)) if gaussian_enabled else ("raw", 0.0)
        if cache_key in delta_cache:
            return delta_cache[cache_key]
        raw_delta = preview_inputs.get("delta_zyx")
        if raw_delta is None:
            raw_delta = np.asarray(preview_inputs["image_arr_t1"], dtype=np.float32) - np.asarray(
                preview_inputs["image_arr_t0"],
                dtype=np.float32,
            )
            preview_inputs["delta_zyx"] = raw_delta.astype(np.float32, copy=False)
        if gaussian_enabled:
            from timelapsedhrpqct.analysis import maybe_smooth_density

            delta = maybe_smooth_density(
                np.asarray(raw_delta, dtype=np.float32),
                gaussian_filter=True,
                gaussian_sigma=sigma,
            )
        else:
            delta = np.asarray(raw_delta, dtype=np.float32)
        delta_cache[cache_key] = delta
        return delta

    def _compute_pair_remodelling_preview_from_cached_delta(self, preview_inputs, *, valid_mask, label_map):
        try:
            from timelapsedhrpqct.analysis import compute_pair_remodelling_preview_from_delta

            delta = self._preview_delta_for_current_settings(preview_inputs)
            return self._compute_pair_remodelling_preview_compat(
                compute_pair_remodelling_preview_from_delta,
                delta=delta,
                seg_arr_t0=preview_inputs["seg_arr_t0"],
                seg_arr_t1=preview_inputs["seg_arr_t1"],
                valid_mask=valid_mask,
                threshold=float(self.analysisThreshold.value),
                cluster_size=int(self.analysisCluster.value),
                method=self._current_analysis_method(),
                label_map=label_map,
                support_mask_t0=preview_inputs.get("support_mask_t0"),
                support_mask_t1=preview_inputs.get("support_mask_t1"),
                marrow_mask_dilation_voxels=int(self.analysisMarrowMaskDilation.value),
                marrow_mask_erosion_voxels=int(self.analysisMarrowMaskErosion.value),
            )
        except ImportError:
            from timelapsedhrpqct.analysis import compute_pair_remodelling_preview

            return self._compute_pair_remodelling_preview_compat(
                compute_pair_remodelling_preview,
                image_arr_t0=preview_inputs["image_arr_t0"],
                image_arr_t1=preview_inputs["image_arr_t1"],
                seg_arr_t0=preview_inputs["seg_arr_t0"],
                seg_arr_t1=preview_inputs["seg_arr_t1"],
                valid_mask=valid_mask,
                threshold=float(self.analysisThreshold.value),
                cluster_size=int(self.analysisCluster.value),
                method=self._current_analysis_method(),
                gaussian_filter=bool(self.analysisGaussianFilterCheck.checked),
                gaussian_sigma=float(self.analysisGaussianSigma.value),
                label_map=label_map,
                support_mask_t0=preview_inputs.get("support_mask_t0"),
                support_mask_t1=preview_inputs.get("support_mask_t1"),
                marrow_mask_dilation_voxels=int(self.analysisMarrowMaskDilation.value),
                marrow_mask_erosion_voxels=int(self.analysisMarrowMaskErosion.value),
            )

    def _compute_pair_metric_rows(self, preview_inputs):
        from timelapsedhrpqct.analysis import build_series_common_masks

        support_t0 = np.asarray(preview_inputs["support_mask_t0"], dtype=bool)
        support_t1 = np.asarray(preview_inputs["support_mask_t1"], dtype=bool)
        rows = []
        for compartment in self._pair_metric_compartments(preview_inputs):
            comp_t0, comp_t1 = self._preview_compartment_masks(preview_inputs, compartment)
            valid_mask = build_series_common_masks(
                {
                    "full": [support_t0, support_t1],
                    compartment: [comp_t0, comp_t1],
                },
                [compartment],
                int(self._analysis_erosion_voxels),
                full_mask_dilation_voxels=int(self.analysisFullMaskDilation.value),
            )[compartment]
            preview = self._compute_pair_remodelling_preview_from_cached_delta(
                preview_inputs,
                valid_mask=valid_mask,
                label_map=self._interactive_preview_label_map(),
            )
            rows.append(
                {
                    "compartment": str(compartment),
                    "formation_frac_bv0": float(preview.formation_frac_bv0),
                    "resorption_frac_bv0": float(preview.resorption_frac_bv0),
                }
            )
        return rows

    def _refresh_pair_metrics_for_current_selection(self, *_args):
        node_id = self.remodellingFullSegCombo.currentData
        full_seg = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id is not None else None
        source_path = str(full_seg.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "") if full_seg is not None else ""
        if not source_path:
            self._set_pair_metric_labels(None, None)
            return
        ctx = self._parse_remodelling_source_context(source_path)
        if ctx is None:
            ctx = {}
        saved_rows = self._saved_pair_metric_rows_for_context(ctx)
        if self._metric_rows_have_finite_fractions(saved_rows):
            self._set_pair_metric_rows(saved_rows)
            return
        if not self._scene_compartment_is_interactive_source(ctx.get("compartment", "")):
            self._set_pair_metric_labels(None, None, compartment=self._scene_display_compartment_name(ctx.get("compartment", "")))
            return
        try:
            preview_inputs = self._get_interactive_preview_inputs(source_path)
            preview = self._compute_pair_remodelling_preview_from_cached_delta(
                preview_inputs,
                valid_mask=self._display_valid_mask_for_preview_inputs(preview_inputs),
                label_map=self._interactive_preview_label_map(),
            )
            compartment = str((preview_inputs.get("context") or {}).get("compartment", "full"))
            self._set_pair_metric_labels(
                formation_frac=preview.formation_frac_bv0,
                resorption_frac=preview.resorption_frac_bv0,
                compartment=compartment,
            )
        except Exception as exc:
            self._set_pair_metric_labels(None, None)
            self._show(f"[preview] pair metrics unavailable: {exc}")

    def _on_remodelling_selection_changed(self, *_args):
        self._activate_remodelling_display_for_current_selection()
        self._refresh_pair_metrics_for_current_selection()

    def _activate_remodelling_display_for_current_selection(self):
        if not hasattr(self, "remodellingFullSegCombo"):
            return False
        node_id = self.remodellingFullSegCombo.currentData
        node = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id is not None else None
        if node is None:
            return False
        try:
            scene = slicer.mrmlScene
            for class_name in ("vtkMRMLScalarVolumeNode", "vtkMRMLLabelMapVolumeNode", "vtkMRMLSegmentationNode"):
                for index in range(scene.GetNumberOfNodesByClass(class_name)):
                    other_node = scene.GetNthNodeByClass(index, class_name)
                    if other_node is None:
                        continue
                    if str(other_node.GetAttribute("TimelapsedHRpQCT.RemodellingFull") or "") != "1":
                        continue
                    display = other_node.GetDisplayNode()
                    if display is None:
                        other_node.CreateDefaultDisplayNodes()
                        display = other_node.GetDisplayNode()
                    if display is not None:
                        display.SetVisibility(other_node is node)
            self._style_remodelling_scalar_volume(node, activate_display=False)
            slicer.util.setSliceViewerLayers(foreground=node, foregroundOpacity=0.65, fit=False)
            self._center_slices_on_node(node, fit_to_bounds=True)
            return True
        except Exception as exc:
            self._show(f"[preview] could not activate selected remodelling image: {exc}")
            return False

    def _get_subject_series_preview_inputs(self, subject_id, site):
        from timelapsedhrpqct.analysis import (
            adjacent_pair_key,
            compute_pair_remodelling_preview,
            compute_pair_trajectory_summary,
            dilate_mask_xy,
            erode_mask,
        )
        from timelapsedhrpqct.processing.analysis_io import discover_analysis_sessions

        imported_root = self._imported_dataset_root()
        if imported_root is None or not imported_root.exists():
            raise ValueError("Imported dataset root is not available.")
        sessions = discover_analysis_sessions(
            imported_root,
            subject_id,
            site,
            use_filled_images=False,
            require_seg=False,
        )
        if len(sessions) < 2:
            raise ValueError("Need at least two transformed sessions for series summary.")
        ordered = []
        for session in sessions:
            image = sitk.ReadImage(str(session.image_path))
            image_arr = sitk.GetArrayFromImage(image).astype(np.float32, copy=False)
            seg_arr = self._read_seg_array_for_preview(session, image)
            support = self._load_support_mask_array(session.mask_paths, session.image_path)
            ordered.append(
                {
                    "session_id": str(session.session_id),
                    "spacing_xyz": tuple(float(x) for x in image.GetSpacing()),
                    "origin_xyz": tuple(float(x) for x in image.GetOrigin()),
                    "image_arr": image_arr,
                    "seg_arr": seg_arr,
                    "support_mask": support,
                }
            )
        return ordered

    def _create_remodelling_display_from_array(
        self,
        segmentation_name,
        label_arr_zyx,
        spacing_xyz,
        origin_xyz,
        folder_item_id=None,
        create_full=True,
        source_path=None,
        interactive_cache_key=None,
        valid_mask_zyx=None,
        geometry_source_node=None,
        center_slices=True,
        activate_display=True,
    ):
        filtered_arr = self._apply_preview_label_filters(label_arr_zyx, valid_mask_zyx=valid_mask_zyx)
        full_seg = None

        if create_full:
            full_seg = self._create_remodelling_scalar_node_from_label_array(
                name=f"{segmentation_name}_full",
                label_arr_zyx=filtered_arr,
                spacing_xyz=spacing_xyz,
                origin_xyz=origin_xyz,
                folder_item_id=folder_item_id,
                activate_display=activate_display,
            )
            self._copy_volume_geometry(geometry_source_node, full_seg)
            full_seg.SetAttribute("TimelapsedHRpQCT.RemodellingFull", "1")
            if source_path is not None:
                full_seg.SetAttribute("TimelapsedHRpQCT.RemodellingSourcePath", str(Path(source_path).resolve()))
            if interactive_cache_key is not None:
                full_seg.SetAttribute("TimelapsedHRpQCT.RemodellingInteractiveCacheKey", str(interactive_cache_key))
            if center_slices:
                self._center_slices_on_segmentation(full_seg)

        self._refresh_remodelling_full_selector()
        return full_seg, None

    def _create_remodelling_segmentations_from_array(self, *args, **kwargs):
        return self._create_remodelling_display_from_array(*args, **kwargs)

    def _selected_remodelling_source_path(self):
        if not hasattr(self, "remodellingFullSegCombo"):
            return ""
        node_id = self.remodellingFullSegCombo.currentData
        node = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id is not None else None
        if node is None:
            return ""
        return str(node.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "")

    def _set_remodelling_selector_by_source_path(self, source_path):
        source = str(source_path or "")
        if not source:
            return False
        for index in range(int(self.remodellingFullSegCombo.count)):
            node_id = self.remodellingFullSegCombo.itemData(index)
            node = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id is not None else None
            if node is None:
                continue
            if str(node.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "") == source:
                self.remodellingFullSegCombo.setCurrentIndex(index)
                return True
        return False

    def _remodelling_session_sort_key(self, session_id):
        text = str(session_id or "")
        match = re.search(r"(\d+)$", text)
        if match:
            return (0, int(match.group(1)), text)
        return (1, text)

    def _remodelling_source_sort_key(self, source_path):
        ctx = self._parse_remodelling_source_context(source_path)
        if ctx is None:
            return (str(source_path), "", "", "")
        return (
            str(ctx.get("subject_id", "")),
            str(ctx.get("site", "")),
            self._remodelling_session_sort_key(str(ctx.get("t0", ""))),
            self._remodelling_session_sort_key(str(ctx.get("t1", ""))),
            str(ctx.get("compartment", "")),
            str(source_path),
        )

    def _warn_missing_scene_baseline_pairs(self, pairs, session_ids):
        self._last_missing_scene_baseline_pairs = []
        if self._current_analysis_pair_mode() != "baseline" or len(session_ids) < 3:
            return
        baseline_session = session_ids[0]
        expected_pairs = {
            (baseline_session, session_id)
            for session_id in session_ids[1:]
        }
        missing_pairs = sorted(
            expected_pairs.difference(set(pairs)),
            key=lambda pair: (
                self._remodelling_session_sort_key(pair[0]),
                self._remodelling_session_sort_key(pair[1]),
            ),
        )
        if not missing_pairs:
            return
        self._last_missing_scene_baseline_pairs = list(missing_pairs)
        missing_text = ", ".join(f"{t0} -> {t1}" for t0, t1 in missing_pairs)
        message = (
            f"[scene] baseline comparison(s) not yet computed: {missing_text}. "
            "Running analysis with Pair mode = Baseline to compute the actual baseline comparisons."
        )
        self._show(message)
        if hasattr(self, "sceneStatusLabel") and self.sceneStatusLabel is not None:
            self.sceneStatusLabel.text = message.replace("[scene] ", "")

    def _detect_missing_scene_baseline_pairs(self, rows):
        row_list = list(rows or [])
        pairs = []
        for row in row_list:
            try:
                t0, t1 = [part.strip() for part in str(row[0]).split("->", 1)]
            except Exception:
                continue
            pairs.append((t0, t1))
        session_ids = sorted(
            {session for pair in pairs for session in pair},
            key=self._remodelling_session_sort_key,
        )
        self._warn_missing_scene_baseline_pairs(pairs, session_ids)
        return row_list

    def _clear_scene_analysis_outputs_for_refresh(self, plan):
        output_root = Path(plan.output_root)
        patterns = [
            "*_remodelling.nii.gz",
            "*_remodelling.mha",
            "*_pairwise_remodelling.csv",
            "*_trajectory_metrics.csv",
        ]
        removed = 0
        for pattern in patterns:
            for path in output_root.rglob(pattern):
                if not path.is_file():
                    continue
                try:
                    path.unlink()
                    removed += 1
                except Exception as exc:
                    self._show(f"[scene] could not remove stale analysis output {path}: {exc}")
        if removed:
            self._show(f"[scene] removed {removed} stale analysis output(s) before scene analysis refresh.")

    def _run_scene_analysis_for_missing_pair_mode(self):
        missing_pairs = list(getattr(self, "_last_missing_scene_baseline_pairs", []) or [])
        if not missing_pairs or self.logic.is_running():
            return False
        plan = getattr(self, "_last_scene_results_plan", None)
        if plan is None:
            return False
        if not self._require_pipeline_installed():
            return False
        self._last_missing_scene_baseline_pairs = []
        self._clear_scene_analysis_outputs_for_refresh(plan)
        cfg = self.logic.create_override_config(
            self._scene_settings_override(),
            results_root=plan.output_root,
        )
        pair_mode_label = str(self.analysisPairModeCombo.currentText or self._current_analysis_pair_mode())
        missing_text = ", ".join(f"{t0} -> {t1}" for t0, t1 in missing_pairs)
        self.sceneStatusLabel.text = (
            f"Running analysis refresh for {missing_text} with Pair mode = {pair_mode_label}..."
        )
        self._show(
            f"[scene] running analysis refresh for missing comparison(s) {missing_text} "
            f"with Pair mode = {pair_mode_label}."
        )
        self._set_stage_status("analysis", "pending")
        self._active_stage = "analysis"
        self._is_full_pipeline_run = False
        self._run_skips_mask_generation = True
        self._run_includes_analysis = True
        self._last_scene_plan = plan
        self._run([
            "analyse",
            str(plan.output_root),
            "--thr",
            str(float(self.analysisThreshold.value)),
            "--clusters",
            str(int(self.analysisCluster.value)),
            *self._profile_cli_args(),
            "--config",
            cfg,
        ])
        return True

    def _refresh_remodelling_full_selector(self):
        selected_source_path = self._selected_remodelling_source_path()
        self.remodellingFullSegCombo.clear()
        scene = slicer.mrmlScene
        entries = []
        for class_name in ("vtkMRMLScalarVolumeNode", "vtkMRMLLabelMapVolumeNode", "vtkMRMLSegmentationNode"):
            for i in range(scene.GetNumberOfNodesByClass(class_name)):
                node = scene.GetNthNodeByClass(i, class_name)
                if node is None:
                    continue
                if not str(node.GetAttribute("TimelapsedHRpQCT.RemodellingFull") or "") == "1":
                    continue
                source_path = str(node.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "")
                entries.append((self._remodelling_source_sort_key(source_path), node.GetName(), node.GetID()))
        for _sort_key, name, node_id in sorted(entries):
            self.remodellingFullSegCombo.addItem(name, node_id)
        if not self._set_remodelling_selector_by_source_path(selected_source_path):
            if self.remodellingFullSegCombo.count > 0 and self.remodellingFullSegCombo.currentIndex < 0:
                self.remodellingFullSegCombo.setCurrentIndex(0)

    def _on_apply_interactive_remodelling(self):
        if self.logic.is_running():
            self._show("[preview] interactive remodelling update skipped while pipeline is running.")
            return
        if getattr(self, "_last_scene_results_plan", None) is not None:
            self._refresh_scene_results_table_from_loaded_remodelling()
            if self._run_scene_analysis_for_missing_pair_mode():
                return

        node_id = self.remodellingFullSegCombo.currentData
        if node_id is None:
            return
        full_seg = slicer.mrmlScene.GetNodeByID(str(node_id))
        if full_seg is None:
            return

        source_path = str(full_seg.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "")
        if not source_path:
            slicer.util.warningDisplay("Selected remodelling segmentation is missing source metadata.")
            return
        ctx = self._parse_remodelling_source_context(source_path)
        if ctx is None:
            ctx = {}
        if not self._scene_compartment_is_interactive_source(ctx.get("compartment", "")):
            view_state = self._capture_slice_view_state()
            self._set_interactive_preview_busy(True, "Updating remodelling ROI union...")
            try:
                preview, preview_inputs, _source_node, compartments = self._compute_pair_union_remodelling_preview(
                    ctx,
                    source_path=source_path,
                )
                sh = self._subject_hierarchy()
                folder_id = None
                if sh is not None:
                    item_id = sh.GetItemByDataNode(full_seg)
                    if item_id:
                        folder_id = sh.GetItemParent(item_id)
                base_name = str(full_seg.GetName() or "")
                if base_name.endswith("_full"):
                    base_name = base_name[:-5]
                new_full, _preview = self._create_remodelling_display_from_array(
                    segmentation_name=base_name,
                    label_arr_zyx=preview.label_image,
                    spacing_xyz=preview_inputs["spacing_xyz"],
                    origin_xyz=preview_inputs["origin_xyz"],
                    folder_item_id=folder_id,
                    create_full=True,
                    source_path=source_path,
                    interactive_cache_key=preview_inputs["cache_key"],
                    valid_mask_zyx=preview.valid_mask,
                    geometry_source_node=full_seg,
                    center_slices=False,
                )
                slicer.mrmlScene.RemoveNode(full_seg)
                if new_full is not None:
                    idx = self.remodellingFullSegCombo.findData(new_full.GetID())
                    if idx >= 0:
                        self.remodellingFullSegCombo.setCurrentIndex(idx)
                    self._activate_remodelling_display_for_current_selection()
                self._restore_slice_view_state(view_state)
            except Exception as exc:
                self._set_interactive_preview_busy(False, "Update failed")
                slicer.util.warningDisplay(f"Interactive remodelling ROI-union update failed:\n{exc}")
                return
            metric_rows = self._compute_pair_metric_rows(preview_inputs)
            self._set_pair_metric_rows(metric_rows)
            self._refresh_scene_results_table_from_loaded_remodelling()
            self._set_interactive_preview_busy(False, "Ready")
            self._show(
                "[preview] remodelling ROI union updated "
                f"from {len(compartments)} ROI(s) "
                f"(thr={float(self.analysisThreshold.value):g}, "
                f"cluster={int(self.analysisCluster.value)}, "
                f"method={self._current_analysis_method()}, "
                f"gauss={'on' if self.analysisGaussianFilterCheck.checked else 'off'}, "
                f"sigma={float(self.analysisGaussianSigma.value):g})."
            )
            return

        view_state = self._capture_slice_view_state()
        self._set_interactive_preview_busy(True, "Updating remodelling preview...")
        try:
            preview_inputs = self._get_interactive_preview_inputs(source_path)

            valid_mask = self._display_valid_mask_for_preview_inputs(preview_inputs)

            preview = self._compute_pair_remodelling_preview_from_cached_delta(
                preview_inputs,
                valid_mask=valid_mask,
                label_map=self._interactive_preview_label_map(),
            )
        except Exception as exc:
            self._set_interactive_preview_busy(False, "Update failed")
            slicer.util.warningDisplay(f"Interactive remodelling update failed:\n{exc}")
            return
        try:
            preview_inputs["current_label_arr"] = preview.label_image
            sh = self._subject_hierarchy()
            folder_id = None
            if sh is not None:
                item_id = sh.GetItemByDataNode(full_seg)
                if item_id:
                    folder_id = sh.GetItemParent(item_id)
            base_name = str(full_seg.GetName() or "")
            if base_name.endswith("_full"):
                base_name = base_name[:-5]

            new_full, _preview = self._create_remodelling_display_from_array(
                segmentation_name=base_name,
                label_arr_zyx=preview.label_image,
                spacing_xyz=preview_inputs["spacing_xyz"],
                origin_xyz=preview_inputs["origin_xyz"],
                folder_item_id=folder_id,
                create_full=True,
                source_path=source_path,
                interactive_cache_key=preview_inputs["cache_key"],
                valid_mask_zyx=preview.valid_mask,
                geometry_source_node=full_seg,
                center_slices=False,
            )
            slicer.mrmlScene.RemoveNode(full_seg)
            if new_full is not None:
                idx = self.remodellingFullSegCombo.findData(new_full.GetID())
                if idx >= 0:
                    self.remodellingFullSegCombo.setCurrentIndex(idx)
                self._activate_remodelling_display_for_current_selection()
            self._restore_slice_view_state(view_state)
        except Exception as exc:
            self._set_interactive_preview_busy(False, "Update failed")
            slicer.util.warningDisplay(f"Interactive remodelling display update failed:\n{exc}")
            return
        self._set_pair_metric_labels(
            formation_frac=preview.formation_frac_bv0,
            resorption_frac=preview.resorption_frac_bv0,
            compartment=str((preview_inputs.get("context") or {}).get("compartment", "full")),
        )
        self._refresh_scene_results_table_from_loaded_remodelling()
        self._set_interactive_preview_busy(False, "Ready")
        self._show(
            "[preview] remodelling updated "
            f"(thr={float(self.analysisThreshold.value):g}, "
            f"cluster={int(self.analysisCluster.value)}, "
            f"method={self._current_analysis_method()}, "
            f"gauss={'on' if self.analysisGaussianFilterCheck.checked else 'off'}, "
            f"sigma={float(self.analysisGaussianSigma.value):g})."
        )

    def _capture_slice_view_state(self):
        state = {}
        try:
            lm = slicer.app.layoutManager()
            if lm is None:
                return state
            for name in ("Red", "Yellow", "Green"):
                widget = lm.sliceWidget(name)
                node = widget.mrmlSliceNode() if widget is not None else None
                if node is None:
                    continue
                entry = {"field_of_view": tuple(float(v) for v in node.GetFieldOfView())}
                if hasattr(node, "GetSliceOffset"):
                    entry["slice_offset"] = float(node.GetSliceOffset())
                state[name] = entry
        except Exception:
            return state
        return state

    def _restore_slice_view_state(self, state):
        if not state:
            return
        try:
            lm = slicer.app.layoutManager()
            if lm is None:
                return
            for name, entry in state.items():
                widget = lm.sliceWidget(name)
                node = widget.mrmlSliceNode() if widget is not None else None
                if node is None:
                    continue
                fov = entry.get("field_of_view")
                if fov is not None and len(fov) >= 3:
                    node.SetFieldOfView(float(fov[0]), float(fov[1]), float(fov[2]))
                if "slice_offset" in entry and hasattr(node, "SetSliceOffset"):
                    node.SetSliceOffset(float(entry["slice_offset"]))
            self._ensure_slice_scale_bars()
        except Exception:
            pass

    def _compute_series_summary_for_current_subject(self):
        patient_key = self._current_patient_key()
        if patient_key is None:
            raise ValueError("No processed patient selected.")
        subject_id, site = patient_key
        selected_pairs = self._selected_series_adjacent_pairs()
        if not selected_pairs:
            raise ValueError("Select at least one adjacent interval.")

        from timelapsedhrpqct.analysis import (
            adjacent_pair_key,
            compute_pair_remodelling_preview,
            compute_pair_trajectory_summary,
            dilate_mask_xy,
            erode_mask,
        )

        series_inputs = self._get_subject_series_preview_inputs(subject_id, site)
        adjacent_events = []
        pair_rows = []
        for i in range(len(series_inputs) - 1):
            t0 = series_inputs[i]
            t1 = series_inputs[i + 1]
            key = adjacent_pair_key(t0["session_id"], t1["session_id"])
            support0 = np.asarray(t0["support_mask"], dtype=bool)
            support1 = np.asarray(t1["support_mask"], dtype=bool)
            if int(self.analysisFullMaskDilation.value) > 0:
                support0 = dilate_mask_xy(support0, int(self.analysisFullMaskDilation.value))
                support1 = dilate_mask_xy(support1, int(self.analysisFullMaskDilation.value))
            valid_mask = erode_mask(support0 & support1, int(self._analysis_erosion_voxels))
            preview = self._compute_pair_remodelling_preview_compat(
                compute_pair_remodelling_preview,
                image_arr_t0=t0["image_arr"],
                image_arr_t1=t1["image_arr"],
                seg_arr_t0=t0["seg_arr"],
                seg_arr_t1=t1["seg_arr"],
                valid_mask=valid_mask,
                threshold=float(self.analysisThreshold.value),
                cluster_size=int(self.analysisCluster.value),
                method=self._current_analysis_method(),
                gaussian_filter=bool(self.analysisGaussianFilterCheck.checked),
                gaussian_sigma=float(self.analysisGaussianSigma.value),
                label_map=self._interactive_preview_label_map(),
                support_mask_t0=t0["support_mask"],
                support_mask_t1=t1["support_mask"],
                marrow_mask_dilation_voxels=int(self.analysisMarrowMaskDilation.value),
                marrow_mask_erosion_voxels=int(self.analysisMarrowMaskErosion.value),
            )
            if key in selected_pairs:
                adjacent_events.append((t0["session_id"], t1["session_id"], preview.formation.copy(), preview.resorption.copy()))
            pair_rows.append(
                {
                    "subject_id": subject_id,
                    "site": site,
                    "compartment": "full",
                    "t0": t0["session_id"],
                    "t1": t1["session_id"],
                    "threshold": float(self.analysisThreshold.value),
                    "cluster_min_size": int(self.analysisCluster.value),
                    "formation_vox": int(preview.formation_vox),
                    "resorption_vox": int(preview.resorption_vox),
                    "BV0_vox": int(preview.bv0_vox),
                    "formation_frac_bv0": float(preview.formation_frac_bv0),
                    "resorption_frac_bv0": float(preview.resorption_frac_bv0),
                }
            )
        summary = compute_pair_trajectory_summary(
            compartment="full",
            threshold=float(self.analysisThreshold.value),
            cluster_size=int(self.analysisCluster.value),
            common_region_path="interactive",
            valid_shape=series_inputs[0]["image_arr"].shape,
            adjacent_events=adjacent_events,
            selected_adjacent_pairs=selected_pairs,
        )
        selected_pair_rows = [
            row for row in pair_rows
            if f"{row['t0']}->{row['t1']}" in set(selected_pairs)
        ]
        if selected_pair_rows:
            summary["mean_formation_frac_bv0"] = float(
                np.nanmean([float(row["formation_frac_bv0"]) for row in selected_pair_rows])
            )
            summary["mean_resorption_frac_bv0"] = float(
                np.nanmean([float(row["resorption_frac_bv0"]) for row in selected_pair_rows])
            )
        else:
            summary["mean_formation_frac_bv0"] = float("nan")
            summary["mean_resorption_frac_bv0"] = float("nan")
        summary["subject_id"] = subject_id
        summary["site"] = site
        return pair_rows, summary

    def _refresh_saved_cohort_summary(self):
        imported = self._imported_dataset_root()
        if imported is None:
            self._latest_study_summary_rows = []
            self._set_series_summary_labels(None)
            self._set_series_summary_saved_state("Results root not available.")
            return
        selected_pairs = self._selected_series_adjacent_pairs()
        if not selected_pairs:
            self._latest_study_summary_rows = []
            self._set_series_summary_labels(None)
            self._set_series_summary_saved_state("Select at least one comparison pair.")
            return
        try:
            cohort_rows = self._read_saved_pairwise_cohort_rows(
                imported,
                selected_pairs=set(selected_pairs),
            )
        except Exception as exc:
            self._latest_study_summary_rows = []
            self._set_series_summary_labels(None)
            self._set_series_summary_saved_state(f"Could not read cohort analysis: {exc}")
            return

        if not cohort_rows:
            self._latest_study_summary_rows = []
            self._set_series_summary_labels(None)
            self._set_series_summary_saved_state("No saved cohort rows found for the selected pairs.")
            return

        rows_by_compartment = {}
        for row in cohort_rows:
            rows_by_compartment.setdefault(row["compartment"], []).append(row)
        summary_rows = []
        for compartment in sorted(rows_by_compartment.keys()):
            rows_for_compartment = rows_by_compartment[compartment]
            summary_rows.append(
                {
                    "compartment": compartment,
                    "mean_formation_frac_bv0": float(
                        np.nanmean([row["formation_frac_bv0"] for row in rows_for_compartment])
                    ),
                    "mean_resorption_frac_bv0": float(
                        np.nanmean([row["resorption_frac_bv0"] for row in rows_for_compartment])
                    ),
                    "mean_net_change_frac_bv0": float(
                        np.nanmean([row["NV_BV"] for row in rows_for_compartment])
                    ),
                    "mean_active_frac_bv0": float(
                        np.nanmean([row["AV_BV"] for row in rows_for_compartment])
                    ),
                    "n_subjects": len(
                        {(row["subject_id"], row["site"]) for row in rows_for_compartment}
                    ),
                }
            )
        summary = {
            "rows": summary_rows,
            "trajectory_selected_adjacent_pairs": list(selected_pairs),
            "cohort_rows": cohort_rows,
        }
        self._latest_study_summary_rows = cohort_rows
        self._set_series_summary_labels(summary)
        self._set_series_summary_saved_state(
            "Showing saved cohort means from existing pairwise analysis outputs for all available masks."
        )

    def _on_update_series_summary(self):
        self._refresh_saved_cohort_summary()

    def _read_saved_pairwise_cohort_rows(self, imported, selected_pairs=None):
        from timelapsedhrpqct.dataset.derivative_paths import (
            analysis_metadata_path,
            pairwise_remodelling_csv_path,
        )

        cohort_rows = []
        selected_pairs = set(selected_pairs or [])
        if not self._patient_keys:
            self._refresh_patient_list()
        for subject_id, site in list(self._patient_keys):
            pairwise_path = pairwise_remodelling_csv_path(imported, subject_id, site)
            if not pairwise_path.exists():
                continue
            metadata_profile = ""
            try:
                meta_path = analysis_metadata_path(imported, subject_id, site)
                if meta_path.exists():
                    metadata_profile = str((json.loads(meta_path.read_text(encoding="utf-8")) or {}).get("profile") or "")
            except Exception as exc:
                self._show(f"[export] could not read analysis metadata profile for sub-{subject_id} site-{site}: {exc}")
            with pairwise_path.open("r", encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    compartment = str(row.get("compartment", "")).strip()
                    if not compartment:
                        continue
                    pair_key = f"{row.get('t0')}->{row.get('t1')}"
                    if selected_pairs and pair_key not in selected_pairs:
                        continue
                    base_row = dict(row)
                    base_row.update(
                        {
                            "subject_id": subject_id,
                            "site": site,
                            "compartment": compartment,
                            "profile": row.get("profile") or row.get("config_profile") or metadata_profile or self._cohort_export_profile(),
                            "t0": row.get("t0"),
                            "t1": row.get("t1"),
                            "pair_key": pair_key,
                            "formation_frac_bv0": float(row.get("formation_frac_bv0", "nan")),
                            "resorption_frac_bv0": float(row.get("resorption_frac_bv0", "nan")),
                        }
                    )
                    cohort_rows.append(enrich_cohort_export_row(base_row))
        return cohort_rows

    def _csv_fieldnames(self, rows):
        fields = []
        seen = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    fields.append(key)
                    seen.add(key)
        return fields

    def _write_csv_rows(self, path, rows):
        rows = list(rows)
        if not rows:
            return
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=self._csv_fieldnames(rows))
            writer.writeheader()
            writer.writerows(rows)

    def _write_xlsx_rows(self, path, sheets):
        from openpyxl import Workbook

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        for title, rows in sheets:
            rows = list(rows)
            if not rows:
                continue
            ws = wb.create_sheet(title=title[:31])
            fieldnames = self._csv_fieldnames(rows)
            ws.append(fieldnames)
            for row in rows:
                ws.append([row.get(field) for field in fieldnames])
        if not wb.sheetnames:
            ws = wb.create_sheet(title="summary")
            ws.append(["message"])
            ws.append(["No rows available"])
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

    def _select_cohort_export_options(self, rows, default_path):
        available = set()
        for row in rows:
            available.update(row.keys())

        default_fields = list(COHORT_DEFAULT_EXPORT_FIELDS)

        extra_specs = [
            (field, tooltip)
            for field, tooltip in COHORT_EXTRA_EXPORT_FIELD_SPECS
            if field in available and field not in default_fields
        ]
        dialog = qt.QDialog(slicer.util.mainWindow())
        dialog.setWindowTitle("Export Timelapsed HR-pQCT Results")
        layout = qt.QVBoxLayout(dialog)

        message = qt.QLabel(
            "The CSV will include the standard reporting fields. "
            "Optionally add extra diagnostic or provenance fields."
        )
        message.wordWrap = True
        layout.addWidget(message)

        path_row = qt.QHBoxLayout()
        path_edit = qt.QLineEdit()
        path_edit.text = str(default_path)
        browse_button = qt.QPushButton("Browse...")
        path_row.addWidget(path_edit)
        path_row.addWidget(browse_button)
        layout.addLayout(path_row)

        def _browse_output_path():
            result = qt.QFileDialog.getSaveFileName(
                dialog,
                "Export Timelapsed HR-pQCT Results",
                path_edit.text,
                "CSV files (*.csv);;Excel workbook (*.xlsx)",
            )
            if isinstance(result, (tuple, list)):
                selected = result[0] if result else ""
            else:
                selected = result
            if str(selected).strip():
                path_edit.text = str(selected)

        browse_button.clicked.connect(_browse_output_path)

        checkboxes = []
        if extra_specs:
            scroll = qt.QScrollArea()
            scroll.setWidgetResizable(True)
            content = qt.QWidget()
            form = qt.QFormLayout(content)
            for field, tooltip in extra_specs:
                checkbox = qt.QCheckBox(field)
                checkbox.toolTip = tooltip
                form.addRow(checkbox)
                checkboxes.append((field, checkbox))
            scroll.setWidget(content)
            scroll.setMinimumHeight(260)
            layout.addWidget(scroll)

        button_row = qt.QHBoxLayout()
        button_row.addStretch(1)
        cancel_button = qt.QPushButton("Cancel")
        save_button = qt.QPushButton("Export CSV")
        cancel_button.connect("clicked()", dialog, "reject()")
        save_button.connect("clicked()", dialog, "accept()")
        button_row.addWidget(cancel_button)
        button_row.addWidget(save_button)
        layout.addLayout(button_row)

        if dialog.exec_() != qt.QDialog.Accepted:
            return None
        selected_path = Path(str(path_edit.text).strip())
        if not str(selected_path).strip():
            return None
        fields = default_fields + [field for field, checkbox in checkboxes if bool(checkbox.checked)]
        return fields, selected_path

    def _on_export_study_summary(self):
        imported = self._imported_dataset_root()
        if imported is None:
            self._set_series_summary_saved_state("Results root not available.")
            return
        try:
            selected_pairs = self._selected_series_adjacent_pairs()
            export_cohort_rows = self._read_saved_pairwise_cohort_rows(imported, selected_pairs=selected_pairs)
        except Exception as exc:
            self._set_series_summary_saved_state(f"Could not read cohort rows: {exc}")
            slicer.util.warningDisplay(f"Could not read cohort rows:\n{exc}")
            return
        if not export_cohort_rows:
            self._set_series_summary_saved_state("No saved cohort rows found to export.")
            return

        default_dir = imported if imported is not None else Path.home()
        default_path = default_dir / default_export_filename("timelapsed_hrpqct_results")
        export_options = self._select_cohort_export_options(export_cohort_rows, default_path)
        if export_options is None:
            return
        export_fields, selected_path = export_options
        export_rows = project_rows_to_fields(export_cohort_rows, export_fields)

        if selected_path.suffix.lower() == ".xlsx":
            csv_path = selected_path.with_suffix(".csv")
        elif selected_path.suffix:
            csv_path = selected_path
        else:
            csv_path = selected_path.with_suffix(".csv")

        try:
            self._write_csv_rows(csv_path, export_rows)
        except Exception as exc:
            self._set_series_summary_saved_state(f"CSV export failed: {exc}")
            return

        xlsx_message = ""
        if selected_path.suffix.lower() == ".xlsx":
            try:
                self._write_xlsx_rows(
                    selected_path,
                    [
                        ("timelapsed_results", export_rows),
                    ],
                )
                xlsx_message = f" XLSX: {selected_path}"
            except Exception as exc:
                xlsx_message = f" XLSX skipped ({exc})."

        self._set_series_summary_saved_state(
            f"Exported cohort rows CSV: {csv_path}.{xlsx_message}"
        )
        self._show(f"[export] study cohort rows CSV written to {csv_path}")
        if selected_path.suffix.lower() == ".xlsx" and xlsx_message.startswith(" XLSX:"):
            self._show(f"[export] study cohort rows XLSX written to {selected_path}")

    def _scene_comparison_table_rows(self):
        if not hasattr(self, "sceneComparisonTable"):
            return []
        table = self.sceneComparisonTable
        headers = [
            str(table.horizontalHeaderItem(col_idx).text())
            for col_idx in range(int(table.columnCount))
        ]
        rows = []
        for row_idx in range(int(table.rowCount)):
            row = {}
            for col_idx, header in enumerate(headers):
                item = table.item(row_idx, col_idx)
                row[header] = str(item.text()) if item is not None else ""
            if row.get("Pair") == "N/A" and row.get("Mask") == "N/A":
                continue
            rows.append(row)
        return rows

    def _on_export_scene_comparison_csv(self):
        rows = self._scene_comparison_table_rows()
        if not rows:
            self.sceneStatusLabel.text = "No scene comparison rows available to export."
            self._show("[scene] no scene comparison rows available to export.")
            return
        default_dir = Path(self._path_text(getattr(self, "sceneResultsRootPath", None)) or ".").expanduser()
        if not default_dir.exists():
            default_dir = Path.home()
        default_path = default_dir / default_export_filename("timelapsed_scene_comparisons")
        result = qt.QFileDialog.getSaveFileName(
            slicer.util.mainWindow(),
            "Export Scene Timelapsed Results",
            str(default_path),
            "CSV files (*.csv)",
        )
        selected = result[0] if isinstance(result, tuple) else result
        if not selected:
            return
        csv_path = Path(str(selected))
        if csv_path.suffix.lower() != ".csv":
            csv_path = csv_path.with_suffix(".csv")
        try:
            self._write_csv_rows(csv_path, rows)
        except Exception as exc:
            self.sceneStatusLabel.text = f"Scene CSV export failed: {exc}"
            slicer.util.warningDisplay(f"Scene CSV export failed:\n{exc}")
            return
        self.sceneStatusLabel.text = f"Exported scene comparison CSV: {csv_path}"
        self._show(f"[scene] scene comparison CSV written to {csv_path}")

    def _on_save_analysis_scenario(self):
        patient_key = self._current_patient_key()
        if patient_key is None:
            slicer.util.warningDisplay("No patient selected.")
            return
        subject_id, site = patient_key
        dialog_result = qt.QInputDialog.getText(
            slicer.util.mainWindow(),
            "Save Analysis Scenario",
            "Scenario name:",
            qt.QLineEdit.Normal,
            f"thr-{int(self.analysisThreshold.value)}_cluster-{int(self.analysisCluster.value)}",
        )
        if isinstance(dialog_result, (tuple, list)):
            name = dialog_result[0] if len(dialog_result) >= 1 else ""
            ok = bool(dialog_result[1]) if len(dialog_result) >= 2 else bool(str(name).strip())
        else:
            name = dialog_result
            ok = bool(str(name).strip())
        if not ok or not str(name).strip():
            return
        scenario_name = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name).strip())
        imported = self._imported_dataset_root()
        if imported is None:
            slicer.util.warningDisplay("Could not resolve results root.")
            return
        scenario_dir = (
            imported
            / f"sub-{subject_id}"
            / f"site-{site}"
            / "analysis"
            / "scenarios"
            / scenario_name
        )
        scenario_dir.mkdir(parents=True, exist_ok=True)
        try:
            pairwise_csv = scenario_dir / "pairwise_preview_metrics.csv"
            cohort_summary_json = scenario_dir / "cohort_summary.json"

            node_id = self.remodellingFullSegCombo.currentData
            full_seg = slicer.mrmlScene.GetNodeByID(str(node_id)) if node_id is not None else None
            source_path = str(full_seg.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "") if full_seg is not None else ""
            cache_key = str(full_seg.GetAttribute("TimelapsedHRpQCT.RemodellingInteractiveCacheKey") or "") if full_seg is not None else ""
            cache_entry = self._interactive_preview_cache.get(cache_key) if cache_key else None
            label_arr = cache_entry.get("current_label_arr") if cache_entry is not None else None
            pair_rows = []
            if cache_entry is not None:
                ctx = cache_entry.get("context") or {}
                for metric_row in self._compute_pair_metric_rows(cache_entry):
                    pair_rows.append(
                        {
                            "subject_id": subject_id,
                            "site": site,
                            "compartment": str(metric_row.get("compartment", "full")),
                            "t0": str(ctx.get("t0", "")),
                            "t1": str(ctx.get("t1", "")),
                            "threshold": float(self.analysisThreshold.value),
                            "cluster_min_size": int(self.analysisCluster.value),
                            "formation_frac_bv0": float(metric_row.get("formation_frac_bv0", float("nan"))),
                            "resorption_frac_bv0": float(metric_row.get("resorption_frac_bv0", float("nan"))),
                        }
                    )
            if pair_rows:
                with pairwise_csv.open("w", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=list(pair_rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(pair_rows)
            if self._latest_series_summary is not None:
                cohort_summary_json.write_text(
                    json.dumps(self._latest_series_summary, indent=2),
                    encoding="utf-8",
                )
            if label_arr is not None and cache_entry is not None:
                img = sitk.GetImageFromArray(np.asarray(label_arr, dtype=np.uint8))
                img.SetSpacing(cache_entry["spacing_xyz"])
                img.SetOrigin(cache_entry["origin_xyz"])
                sitk.WriteImage(img, str(scenario_dir / "current_pair_remodelling.mha"))

            metadata = {
                "kind": "slicer_analysis_scenario",
                "subject_id": subject_id,
                "site": site,
                "scenario_name": scenario_name,
                "method": self._current_analysis_method(),
                "change_region_source": "bone_union" if self.analysisRestrictBoneSupportCheck.checked else "common_mask",
                "binary_reclassification_enabled": bool(self.analysisBinaryReclassificationCheck.checked),
                "threshold": float(self.analysisThreshold.value),
                "cluster_size": int(self.analysisCluster.value),
                "gaussian_filter": bool(self.analysisGaussianFilterCheck.checked),
                "gaussian_sigma": float(self.analysisGaussianSigma.value),
                "full_mask_dilation_voxels": int(self.analysisFullMaskDilation.value),
                "marrow_mask_dilation_voxels": int(self.analysisMarrowMaskDilation.value),
                "marrow_mask_erosion_voxels": int(self.analysisMarrowMaskErosion.value),
                "trajectory_selected_adjacent_pairs": self._selected_series_adjacent_pairs(),
                "source_remodelling_path": source_path or None,
                "pairwise_csv": str(pairwise_csv) if pair_rows else None,
                "cohort_summary_json": str(cohort_summary_json) if self._latest_series_summary is not None else None,
            }
            (scenario_dir / "analysis_scenario.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
        except Exception as exc:
            slicer.util.warningDisplay(f"Saving analysis scenario failed:\n{exc}")
            return
        self._show(f"[save] analysis scenario written to {scenario_dir}")

    def _load_remodelling_as_segmentation(
        self,
        segmentation_name,
        labelmap_path,
        folder_item_id=None,
        create_full=True,
        activate_display=True,
    ):
        try:
            ok, remodelling_node = self._load_volume_node(labelmap_path)
        except Exception as exc:
            self._show(f"[load] failed to load remodelling volume from {labelmap_path}: {exc}")
            return False
        if not ok or remodelling_node is None:
            self._show(f"[load] failed to load remodelling volume from {labelmap_path}")
            return False
        remodelling_node.SetName(f"{segmentation_name}_full" if create_full else str(segmentation_name))
        remodelling_node.SetAttribute("TimelapsedHRpQCT.RemodellingFull", "1")
        remodelling_node.SetAttribute("TimelapsedHRpQCT.RemodellingSourcePath", str(Path(labelmap_path).resolve()))
        self._style_remodelling_scalar_volume(remodelling_node, activate_display=activate_display)
        if folder_item_id is not None:
            self._place_node_in_folder(remodelling_node, folder_item_id)
        if remodelling_node is not None and activate_display:
            self._center_slices_on_node(remodelling_node, fit_to_bounds=True)
        return True

    def _center_slices_on_segmentation(self, seg_node):
        self._center_slices_on_node(seg_node, fit_to_bounds=True)

    def _center_slices_on_node(self, node_to_center, fit_to_bounds=False):
        if node_to_center is None:
            return
        try:
            bounds = [0.0] * 6
            node_to_center.GetBounds(bounds)
            if not all(np.isfinite(bounds)):
                return
            cx = 0.5 * (bounds[0] + bounds[1])
            cy = 0.5 * (bounds[2] + bounds[3])
            cz = 0.5 * (bounds[4] + bounds[5])
            lm = slicer.app.layoutManager()
            if lm is None:
                return
            for name in ("Red", "Yellow", "Green"):
                widget = lm.sliceWidget(name)
                if widget is None:
                    continue
                node = widget.mrmlSliceNode()
                if node is not None:
                    node.JumpSliceByCentering(cx, cy, cz)
                    if fit_to_bounds:
                        self._fit_slice_node_to_bounds(node, widget, bounds, name)
            self._ensure_slice_scale_bars()
        except Exception:
            pass

    def _fit_slice_node_to_bounds(self, slice_node, widget, bounds, view_name):
        try:
            dims = {
                "Red": (abs(bounds[1] - bounds[0]), abs(bounds[3] - bounds[2])),
                "Yellow": (abs(bounds[1] - bounds[0]), abs(bounds[5] - bounds[4])),
                "Green": (abs(bounds[3] - bounds[2]), abs(bounds[5] - bounds[4])),
            }.get(str(view_name), (abs(bounds[1] - bounds[0]), abs(bounds[3] - bounds[2])))
            dim_x = max(float(dims[0]), 1.0)
            dim_y = max(float(dims[1]), 1.0)
            view = widget.sliceView() if widget is not None else None
            render_window = view.renderWindow() if view is not None else None
            size = render_window.GetSize() if render_window is not None else (1, 1)
            aspect = max(0.1, float(max(1, int(size[0]))) / float(max(1, int(size[1]))))
            target_x = max(dim_x * 1.18, 8.0)
            target_y = max(dim_y * 1.18, 8.0)
            if target_x / target_y < aspect:
                target_x = target_y * aspect
            else:
                target_y = target_x / aspect
            current_fov = slice_node.GetFieldOfView()
            z_fov = float(current_fov[2]) if current_fov is not None and len(current_fov) >= 3 else 1.0
            slice_node.SetFieldOfView(float(target_x), float(target_y), z_fov)
        except Exception:
            pass

    def _ensure_slice_scale_bars(self):
        try:
            lm = slicer.app.layoutManager()
            if lm is None:
                return
            for name in ("Red", "Yellow", "Green"):
                widget = lm.sliceWidget(name)
                if widget is None:
                    continue
                view = widget.sliceView()
                render_window = view.renderWindow() if view is not None else None
                renderers = render_window.GetRenderers() if render_window is not None else None
                renderer = renderers.GetFirstRenderer() if renderers is not None else None
                if renderer is None:
                    continue
                record = self._slice_scale_bars.get(name)
                if record is None:
                    points = vtk.vtkPoints()
                    points.SetNumberOfPoints(6)
                    cells = vtk.vtkCellArray()
                    for p0, p1 in ((0, 1), (2, 3), (4, 5)):
                        line = vtk.vtkLine()
                        line.GetPointIds().SetId(0, p0)
                        line.GetPointIds().SetId(1, p1)
                        cells.InsertNextCell(line)
                    polydata = vtk.vtkPolyData()
                    polydata.SetPoints(points)
                    polydata.SetLines(cells)
                    coordinate = vtk.vtkCoordinate()
                    coordinate.SetCoordinateSystemToDisplay()
                    mapper = vtk.vtkPolyDataMapper2D()
                    mapper.SetInputData(polydata)
                    mapper.SetTransformCoordinate(coordinate)
                    line_actor = vtk.vtkActor2D()
                    line_actor.SetMapper(mapper)
                    line_actor.GetProperty().SetColor(1.0, 1.0, 1.0)
                    line_actor.GetProperty().SetLineWidth(3.0)
                    text_actor = vtk.vtkTextActor()
                    text_actor.GetTextProperty().SetColor(1.0, 1.0, 1.0)
                    text_actor.GetTextProperty().SetFontSize(13)
                    text_actor.GetTextProperty().BoldOn()
                    renderer.AddActor2D(line_actor)
                    renderer.AddActor2D(text_actor)
                    record = {
                        "points": points,
                        "polydata": polydata,
                        "line_actor": line_actor,
                        "text_actor": text_actor,
                        "observer_node": None,
                        "observer_tag": None,
                    }
                    self._slice_scale_bars[name] = record
                slice_node = widget.mrmlSliceNode()
                if (
                    slice_node is not None
                    and record.get("observer_node") is not slice_node
                ):
                    old_node = record.get("observer_node")
                    old_tag = record.get("observer_tag")
                    if old_node is not None and old_tag is not None:
                        try:
                            old_node.RemoveObserver(old_tag)
                        except Exception:
                            pass
                    try:
                        record["observer_tag"] = slice_node.AddObserver(
                            vtk.vtkCommand.ModifiedEvent,
                            lambda _caller, _event, view_name=name: self._update_slice_scale_bar(view_name),
                        )
                        record["observer_node"] = slice_node
                    except Exception:
                        record["observer_node"] = None
                        record["observer_tag"] = None
                self._update_slice_scale_bar(name, widget, render_window)
        except Exception:
            pass

    def _update_slice_scale_bar(self, name, widget=None, render_window=None):
        try:
            record = self._slice_scale_bars.get(name)
            if record is None:
                return
            if widget is None or render_window is None:
                lm = slicer.app.layoutManager()
                widget = lm.sliceWidget(name) if lm is not None else None
                view = widget.sliceView() if widget is not None else None
                render_window = view.renderWindow() if view is not None else None
            if widget is None or render_window is None:
                return
            slice_node = widget.mrmlSliceNode()
            if slice_node is None:
                return
            size = render_window.GetSize()
            width_px = max(1, int(size[0]))
            fov = slice_node.GetFieldOfView()
            fov_x_mm = abs(float(fov[0])) if fov is not None else 0.0
            if fov_x_mm <= 0:
                return
            mm_per_px = fov_x_mm / float(width_px)
            candidates_mm = (0.5, 1.0, 2.0, 5.0, 10.0)
            target_px = 105.0
            visible_candidates = [
                mm for mm in candidates_mm if 45.0 <= (mm / mm_per_px) <= 170.0
            ]
            if visible_candidates:
                bar_mm = min(visible_candidates, key=lambda mm: abs((mm / mm_per_px) - target_px))
            else:
                bar_mm = min(candidates_mm, key=lambda mm: abs((mm / mm_per_px) - target_px))
            bar_px = max(24.0, min(180.0, bar_mm / mm_per_px))
            x0 = 24.0
            y0 = 28.0
            x1 = x0 + bar_px
            tick = 5.0
            points = record["points"]
            for idx, point in enumerate(
                (
                    (x0, y0, 0.0),
                    (x1, y0, 0.0),
                    (x0, y0 - tick, 0.0),
                    (x0, y0 + tick, 0.0),
                    (x1, y0 - tick, 0.0),
                    (x1, y0 + tick, 0.0),
                )
            ):
                points.SetPoint(idx, point)
            points.Modified()
            record["polydata"].Modified()
            label = f"{int(bar_mm)} mm" if float(bar_mm).is_integer() else f"{bar_mm:g} mm"
            text_actor = record["text_actor"]
            text_actor.SetInput(label)
            text_actor.SetPosition(x0, y0 + 8.0)
            record["line_actor"].SetVisibility(True)
            text_actor.SetVisibility(True)
            render_window.Render()
        except Exception:
            pass

    def _set_3d_background_black(self):
        try:
            lm = slicer.app.layoutManager()
            if lm is None:
                return
            for i in range(int(lm.threeDViewCount)):
                view = lm.threeDWidget(i).threeDView()
                view_node = view.mrmlViewNode() if view is not None else None
                if view_node is not None:
                    view_node.SetBackgroundColor(0.0, 0.0, 0.0)
                    view_node.SetBackgroundColor2(0.0, 0.0, 0.0)
        except Exception:
            pass

    def _create_segmentation_node_from_role_arrays(
        self,
        segmentation_name,
        role_to_array,
        spacing_xyz,
        origin_xyz,
        session_id=None,
        folder_item_id=None,
    ):
        seg_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode", segmentation_name)
        seg_node.CreateDefaultDisplayNodes()
        seg_logic = slicer.modules.segmentations.logic()
        loaded_any = False

        base_color = self._session_base_color(session_id or "")
        role_tint = {
            "full": 1.00,
            "trab": 0.85,
            "cort": 0.70,
            "regmask": 0.55,
        }

        for role in sorted(role_to_array.keys()):
            arr = (np.asarray(role_to_array[role]) > 0).astype(np.uint8)
            if int(arr.max()) == 0:
                continue

            label_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLLabelMapVolumeNode",
                f"{segmentation_name}_{role}_tmp",
            )
            slicer.util.updateVolumeFromArray(label_node, arr)
            label_node.SetSpacing(float(spacing_xyz[0]), float(spacing_xyz[1]), float(spacing_xyz[2]))
            label_node.SetOrigin(float(origin_xyz[0]), float(origin_xyz[1]), float(origin_xyz[2]))

            existing_ids = vtk.vtkStringArray()
            seg_node.GetSegmentation().GetSegmentIDs(existing_ids)
            before = {existing_ids.GetValue(i) for i in range(existing_ids.GetNumberOfValues())}

            seg_logic.ImportLabelmapToSegmentationNode(label_node, seg_node)
            slicer.mrmlScene.RemoveNode(label_node)

            updated_ids = vtk.vtkStringArray()
            seg_node.GetSegmentation().GetSegmentIDs(updated_ids)
            after = {updated_ids.GetValue(i) for i in range(updated_ids.GetNumberOfValues())}
            new_ids = list(after - before)
            for seg_id in new_ids:
                segment = seg_node.GetSegmentation().GetSegment(seg_id)
                if segment:
                    segment.SetName(str(role))
                    tint = role_tint.get(str(role).lower(), 0.9)
                    segment.SetColor(
                        float(min(max(base_color[0] * tint, 0.0), 1.0)),
                        float(min(max(base_color[1] * tint, 0.0), 1.0)),
                        float(min(max(base_color[2] * tint, 0.0), 1.0)),
                    )
            loaded_any = loaded_any or bool(new_ids)

        if not loaded_any:
            slicer.mrmlScene.RemoveNode(seg_node)
            return False
        self._configure_segmentation_display(seg_node)
        if folder_item_id is not None:
            self._place_node_in_folder(seg_node, folder_item_id)
        return True

    def _merge_raw_session_records(self, records):
        if not records:
            return None

        recs = list(records)

        # First pass: read geometry so we can compute robust z placement.
        geom = []
        for rec in recs:
            img = sitk.ReadImage(str(rec.image_path))
            sx, sy, sz = img.GetSpacing()
            ox, oy, oz = img.GetOrigin()
            slice_range = getattr(rec, "slice_range", None)
            z_start_meta = getattr(slice_range, "z_start", None) if slice_range is not None else None
            geom.append(
                {
                    "rec": rec,
                    "img": img,
                    "spacing": (float(sx), float(sy), float(sz)),
                    "origin": (float(ox), float(oy), float(oz)),
                    "z_start_meta": int(z_start_meta) if z_start_meta is not None else None,
                }
            )

        # Prefer metadata z_start when available and non-degenerate; otherwise use origin-derived z.
        meta_starts = [g["z_start_meta"] for g in geom if g["z_start_meta"] is not None]
        use_meta_z = len(set(meta_starts)) > 1

        images = []
        spacing_xyz = None
        origin_xyz = None
        xmax = 0
        ymax = 0
        zmax = 0
        all_roles = set()
        min_oz = min(g["origin"][2] for g in geom) if geom else 0.0

        for g in geom:
            rec = g["rec"]
            img = g["img"]
            arr = sitk.GetArrayFromImage(img)  # z,y,x
            sx, sy, sz = g["spacing"]
            ox, oy, oz = g["origin"]
            if use_meta_z:
                z_start = int(g["z_start_meta"] or 0)
            else:
                z_start = int(round((float(oz) - float(min_oz)) / float(sz))) if sz > 0 else 0
            z_stop = z_start + int(arr.shape[0])

            if spacing_xyz is None:
                spacing_xyz = (sx, sy, sz)
            if origin_xyz is None:
                origin_xyz = (ox, oy, min_oz)

            ymax = max(ymax, int(arr.shape[1]))
            xmax = max(xmax, int(arr.shape[2]))
            zmax = max(zmax, z_stop)

            role_arrays = {}
            for role, mask_path in (getattr(rec, "mask_paths", {}) or {}).items():
                if mask_path and Path(mask_path).exists():
                    m_arr = sitk.GetArrayFromImage(sitk.ReadImage(str(mask_path)))
                    role_arrays[str(role)] = (m_arr > 0).astype(np.uint8)
                    all_roles.add(str(role))

            images.append((z_start, arr, role_arrays))

        images.sort(key=lambda item: int(item[0]))

        merged_img = np.zeros((zmax, ymax, xmax), dtype=images[0][1].dtype)
        merged_roles = {role: np.zeros((zmax, ymax, xmax), dtype=np.uint8) for role in sorted(all_roles)}

        for z_start, arr, role_arrays in images:
            z0 = int(z_start)
            z1 = z0 + int(arr.shape[0])
            y1 = int(arr.shape[1])
            x1 = int(arr.shape[2])
            merged_img[z0:z1, :y1, :x1] = arr

            for role, mask_arr in role_arrays.items():
                my1 = int(mask_arr.shape[1])
                mx1 = int(mask_arr.shape[2])
                mz1 = z0 + int(mask_arr.shape[0])
                merged_roles[role][z0:mz1, :my1, :mx1] = np.maximum(
                    merged_roles[role][z0:mz1, :my1, :mx1],
                    mask_arr,
                )

        return merged_img, merged_roles, spacing_xyz, origin_xyz

    def _load_volume_node(self, path):
        """Load scalar volume with backward-compatible return handling."""
        loaded = slicer.util.loadVolume(str(path))
        if isinstance(loaded, tuple):
            ok, node = loaded
            return bool(ok), node
        if isinstance(loaded, bool):
            return loaded, None
        return loaded is not None, loaded

    def _load_labelmap_node(self, path):
        """Load labelmap volume with backward-compatible return handling."""
        loaded = slicer.util.loadLabelVolume(str(path))
        if isinstance(loaded, tuple):
            ok, node = loaded
            return bool(ok), node
        if isinstance(loaded, bool):
            return loaded, None
        return loaded is not None, loaded

    def _load_transform_node(self, path):
        loaded = slicer.util.loadTransform(str(path))
        if isinstance(loaded, tuple):
            ok, node = loaded
            return bool(ok), node
        if isinstance(loaded, bool):
            return loaded, None
        return loaded is not None, loaded

    def _scene_baseline_transform_key(self, path):
        if Path(path).suffix.lower() != ".tfm":
            return None
        parsed = self._scene_transform_pair_and_kind_from_path(path)
        if parsed is None:
            return None
        moving_session, _fixed_session, kind = parsed
        if kind not in {"baseline", "final"}:
            return None
        rank = 0 if kind == "final" else 1
        return moving_session, rank

    def _scene_transform_pair_and_kind_from_path(self, path):
        if Path(path).suffix.lower() != ".tfm":
            return None
        match = re.search(
            r"from-ses-(?P<moving>.+?)_to-ses-(?P<fixed>.+?)_(?P<kind>baseline|final|pairwise)\.tfm$",
            Path(path).name,
        )
        if match is None:
            return None
        return str(match.group("moving")), str(match.group("fixed")), str(match.group("kind")).lower()

    def _scene_transform_pair_from_path(self, path):
        parsed = self._scene_transform_pair_and_kind_from_path(path)
        if parsed is None:
            return None
        moving_session, fixed_session, _kind = parsed
        return moving_session, fixed_session

    def _scene_transformable_node_ids_for_timepoint(self, plan, timepoint_index):
        timepoint = plan.timepoints[timepoint_index]
        node_ids = {
            timepoint.image_node_id,
            timepoint.reg_mask_node_id,
            timepoint.full_mask_node_id,
            timepoint.trab_mask_node_id,
            timepoint.cort_mask_node_id,
            timepoint.seg_mask_node_id,
        }
        for roi in plan.rois:
            if timepoint_index < len(roi.node_ids):
                node_ids.add(roi.node_ids[timepoint_index])
        return {str(node_id) for node_id in node_ids if str(node_id or "").strip()}

    def _apply_scene_baseline_transforms(self, plan, loaded_transform_nodes):
        transforms_by_session = {}
        for path, transform_node in (loaded_transform_nodes or {}).items():
            if transform_node is None:
                continue
            parsed = self._scene_baseline_transform_key(path)
            if parsed is None:
                continue
            moving_session, rank = parsed
            current = transforms_by_session.get(moving_session)
            if current is None or rank < current[0]:
                transforms_by_session[moving_session] = (rank, transform_node)

        applied = 0
        applied_node_transforms = {}
        for timepoint_index, timepoint in enumerate(plan.timepoints):
            transform_entry = transforms_by_session.get(str(timepoint.session_id))
            if transform_entry is None:
                continue
            _rank, transform_node = transform_entry
            transform_id = transform_node.GetID()
            for node_id in sorted(self._scene_transformable_node_ids_for_timepoint(plan, timepoint_index)):
                node = slicer.mrmlScene.GetNodeByID(node_id)
                if node is None or node is transform_node:
                    continue
                if not hasattr(node, "SetAndObserveTransformNodeID"):
                    continue
                previous_transform_id = applied_node_transforms.get(node_id)
                if previous_transform_id is not None and previous_transform_id != transform_id:
                    self._show(
                        "[scene] skipped transform assignment for shared node "
                        f"{self._scene_node_name(node_id)}; it is mapped to multiple timepoints."
                    )
                    continue
                try:
                    if hasattr(node, "GetTransformNodeID") and node.GetTransformNodeID() == transform_id:
                        applied_node_transforms[node_id] = transform_id
                        continue
                    node.SetAndObserveTransformNodeID(transform_node.GetID())
                    applied_node_transforms[node_id] = transform_id
                    applied += 1
                except Exception as exc:
                    self._show(f"[scene] could not apply baseline transform to {self._scene_node_name(node_id)}: {exc}")
        return applied

    def _set_path_without_immediate_reset(self, widget, path):
        if not self._qt_object_alive(widget):
            return False
        try:
            previous = widget.blockSignals(True)
            try:
                widget.setCurrentPath(str(path))
            finally:
                widget.blockSignals(previous)
            return True
        except (RuntimeError, ValueError, AttributeError):
            return False

    def _adopt_scene_run_as_current_dataset(self, plan):
        self._set_path_without_immediate_reset(getattr(self, "inputPath", None), plan.input_root)
        self._set_path_without_immediate_reset(getattr(self, "resultsRootPath", None), plan.output_root)
        try:
            self._reset_progress_for_dataset_root()
        except (RuntimeError, ValueError, AttributeError):
            pass
        subject_id, site = self._scene_processed_subject_site(plan)
        if self._qt_object_alive(getattr(self, "patientCombo", None)):
            label = f"sub-{subject_id} | site-{site}"
            try:
                index = self.patientCombo.findText(label)
                if index >= 0:
                    self.patientCombo.setCurrentIndex(index)
            except (RuntimeError, ValueError):
                pass
        if self._qt_object_alive(getattr(self, "loadTypeCombo", None)):
            try:
                idx = self.loadTypeCombo.findText("remodelling image")
                if idx >= 0:
                    self.loadTypeCombo.setCurrentIndex(idx)
            except (RuntimeError, ValueError):
                pass
        self._show(
            "[scene] current dataset set to scene run "
            f"input={plan.input_root} output={plan.output_root}"
        )

    def _normalize_scene_site(self, site):
        try:
            from timelapsedhrpqct.config.models import DiscoveryConfig
            from timelapsedhrpqct.dataset.filename_decoder import normalize_site

            return normalize_site(str(site), DiscoveryConfig()) or str(site)
        except Exception:
            return str(site)

    def _seed_scene_transform_registry(self, plan):
        selected = [
            timepoint
            for timepoint in plan.timepoints
            if timepoint.transform_path is not None and Path(timepoint.transform_path).exists()
        ]
        if not selected:
            return 0
        try:
            from timelapsedhrpqct.dataset.transform_registry import (
                TransformRegistryRecord,
                upsert_transform_registry_record,
            )
        except Exception as exc:
            self._show(f"[scene] could not prepare transform registry: {exc}")
            return 0

        subject_id = plan.subject_id
        site = self._normalize_scene_site(plan.site)
        seeded = 0
        for previous, current in zip(plan.timepoints[:-1], plan.timepoints[1:]):
            transform_path = current.transform_path
            if transform_path is None or not Path(transform_path).exists():
                continue
            moving_session = str(current.session_id)
            fixed_session = str(previous.session_id)
            parsed_transform = self._scene_transform_pair_and_kind_from_path(transform_path)
            if parsed_transform is not None:
                parsed_moving, parsed_fixed, parsed_kind = parsed_transform
                if parsed_kind != "pairwise":
                    self._show(
                        "[scene] skipped selected transform for "
                        f"ses-{moving_session}; file encodes a {parsed_kind} transform, "
                        "but registration reuse needs adjacent pairwise transforms."
                    )
                    continue
                if parsed_moving != moving_session:
                    self._show(
                        "[scene] skipped selected transform for "
                        f"ses-{moving_session}; file encodes ses-{parsed_moving} -> ses-{parsed_fixed}."
                    )
                    continue
                moving_session = parsed_moving
                fixed_session = parsed_fixed
                if fixed_session != str(previous.session_id):
                    self._show(
                        "[scene] registered selected transform "
                        f"{moving_session} -> {fixed_session} from filename; it will only be reused for that pair."
                    )
            try:
                upsert_transform_registry_record(
                    Path(plan.output_root),
                    TransformRegistryRecord(
                        subject_id=subject_id,
                        site=site,
                        stack_index=1,
                        moving_session=moving_session,
                        fixed_session=fixed_session,
                        transform_kind="pairwise",
                        internal_path=Path(transform_path),
                        source_format="slicer_tfm",
                        source_path=Path(transform_path),
                        source_direction="moving_to_fixed",
                        internal_direction="moving_to_fixed",
                        coordinate_convention="SimpleITK_LPS_physical",
                        provenance="slicer_scene_selected_transform",
                        import_timestamp=datetime.now(timezone.utc).isoformat(),
                    ),
                )
                seeded += 1
            except Exception as exc:
                self._show(
                    "[scene] could not seed transform registry for "
                    f"{moving_session} -> {fixed_session}: {exc}"
                )
        return seeded

    def _scene_processed_subject_site(self, plan):
        try:
            from timelapsedhrpqct.dataset.artifacts import (
                iter_fused_session_records,
                iter_imported_stack_records,
            )

            records = list(iter_fused_session_records(Path(plan.output_root))) + list(
                iter_imported_stack_records(Path(plan.output_root))
            )
        except Exception as exc:
            self._show(f"[scene] could not inspect processed scene subject/site: {exc}")
            return plan.subject_id, plan.site

        if not records:
            return plan.subject_id, plan.site
        exact_subject = [record for record in records if str(getattr(record, "subject_id", "")) == plan.subject_id]
        candidates = exact_subject or records
        first = sorted(
            candidates,
            key=lambda record: (
                str(getattr(record, "subject_id", "")),
                str(getattr(record, "site", "")),
                str(getattr(record, "session_id", "")),
            ),
        )[0]
        return str(getattr(first, "subject_id", plan.subject_id)), str(getattr(first, "site", plan.site))

    def _scene_mask_label_name(self, subject_id, site, session_id, role, stack_index=None):
        role_token = {
            "full": "mask-full",
            "trab": "mask-trab",
            "cort": "mask-cort",
            "seg": "mask-seg",
        }.get(str(role), f"mask-{role}")
        stack = f"_stack-{int(stack_index):02d}" if stack_index is not None else ""
        return f"sub-{subject_id}_ses-{session_id}_site-{site}{stack}_{role_token}"

    def _load_scene_mask_labelmap(self, path, name, folder_item_id):
        ok, node = self._load_labelmap_node(path)
        if not ok or node is None:
            return False
        try:
            node.SetName(str(name))
            node.SetAttribute("TimelapsedHRpQCT.GeneratedMask", "1")
            node.SetAttribute("TimelapsedHRpQCT.GeneratedMaskSourcePath", str(Path(path).resolve()))
        except Exception:
            pass
        try:
            display = node.GetDisplayNode()
            if display is not None:
                display.SetVisibility(False)
        except Exception:
            pass
        self._place_node_in_folder(node, folder_item_id)
        return True

    def _load_scene_run_masks(self, plan):
        dataset_root = Path(plan.output_root)
        loaded_masks = 0
        seen_paths = set()
        selected_roles = set(self._scene_requested_mask_roles())
        selected_roles.update({"regmask", "seg"})
        processed_subject_id, processed_site = self._scene_processed_subject_site(plan)
        try:
            from timelapsedhrpqct.dataset.artifacts import iter_imported_stack_records

            # Only native/imported masks are loaded back for scene rediscovery.
            # Fused masks live in transformed space and must not be offered as
            # native scene-run inputs for the original loaded volumes.
            records = list(iter_imported_stack_records(dataset_root))
        except Exception as exc:
            self._show(f"[scene] could not inspect scene-run mask artifacts: {exc}")
            return 0

        for record in records:
            subject_id = str(getattr(record, "subject_id", ""))
            site = str(getattr(record, "site", ""))
            if subject_id != processed_subject_id or site != processed_site:
                continue
            session_id = str(getattr(record, "session_id", ""))
            stack_index = getattr(record, "stack_index", None)
            folder_item_id = self._ensure_load_folder(subject_id, site, session_id, stack_index)

            role_to_path = dict(getattr(record, "mask_paths", {}) or {})
            seg_path = getattr(record, "seg_path", None)
            if seg_path is not None:
                role_to_path.setdefault("seg", Path(seg_path))

            for role, path in sorted(role_to_path.items()):
                if str(role) not in selected_roles:
                    continue
                path = Path(path)
                resolved = path.resolve()
                if resolved in seen_paths or not path.exists():
                    continue
                seen_paths.add(resolved)
                name = self._scene_mask_label_name(
                    subject_id,
                    site,
                    session_id,
                    str(role),
                    stack_index=stack_index,
                )
                try:
                    if self._load_scene_mask_labelmap(path, name, folder_item_id):
                        loaded_masks += 1
                except Exception as exc:
                    self._show(f"[scene] could not load mask {path.name}: {exc}")
        return loaded_masks

    def _format_scene_result_fraction(self, value):
        try:
            number = float(value)
        except Exception:
            return "N/A"
        if not np.isfinite(number):
            return "N/A"
        return f"{number:.5g}"

    def _scene_result_rows(self, plan):
        try:
            from timelapsedhrpqct.dataset.derivative_paths import pairwise_remodelling_csv_path

            subject_id, site = self._scene_processed_subject_site(plan)
            pairwise_path = pairwise_remodelling_csv_path(
                Path(plan.output_root),
                subject_id,
                site,
            )
            self._show(f"[scene] scene_results_table_path={pairwise_path}")
        except Exception as exc:
            self._show(f"[scene] could not resolve scene results table path: {exc}")
            return []
        if not pairwise_path.exists():
            self._show(f"[scene] no pairwise results table found: {pairwise_path}")
            return []

        rows = []
        try:
            with pairwise_path.open("r", encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    t0 = str(row.get("t0") or "").strip()
                    t1 = str(row.get("t1") or "").strip()
                    compartment = str(row.get("compartment") or "").strip()
                    formation = row.get("formation_frac_bv0")
                    resorption = row.get("resorption_frac_bv0")
                    activity = row.get("AV_BV")
                    net = row.get("NV_BV")
                    if activity in (None, ""):
                        try:
                            activity = float(formation) + float(resorption)
                        except Exception:
                            activity = ""
                    if net in (None, ""):
                        try:
                            net = float(formation) - float(resorption)
                        except Exception:
                            net = ""
                    rows.append(
                        [
                            f"{t0} -> {t1}",
                            self._scene_display_compartment_name(compartment),
                            self._format_scene_result_fraction(formation),
                            self._format_scene_result_fraction(resorption),
                            self._format_scene_result_fraction(activity),
                            self._format_scene_result_fraction(net),
                        ]
                    )
        except Exception as exc:
            self._show(f"[scene] could not read scene results table: {exc}")
            return []
        return self._detect_missing_scene_baseline_pairs(rows)

    def _scene_result_rows_from_loaded_remodelling(self):
        source_by_pair = {}
        scene = slicer.mrmlScene
        for class_name in ("vtkMRMLScalarVolumeNode", "vtkMRMLLabelMapVolumeNode", "vtkMRMLSegmentationNode"):
            for index in range(scene.GetNumberOfNodesByClass(class_name)):
                node = scene.GetNthNodeByClass(index, class_name)
                if node is None:
                    continue
                if str(node.GetAttribute("TimelapsedHRpQCT.RemodellingFull") or "") != "1":
                    continue
                source_path = str(node.GetAttribute("TimelapsedHRpQCT.RemodellingSourcePath") or "")
                if not source_path:
                    continue
                ctx = self._parse_remodelling_source_context(source_path)
                if ctx is None:
                    continue
                pair_key = (
                    str(ctx.get("subject_id", "")),
                    str(ctx.get("site", "")),
                    str(ctx.get("t0", "")),
                    str(ctx.get("t1", "")),
                )
                rank = 0 if self._scene_compartment_is_interactive_source(ctx.get("compartment", "")) else 1
                current = source_by_pair.get(pair_key)
                candidate = (rank, source_path, ctx)
                if current is None or candidate[0] < current[0]:
                    source_by_pair[pair_key] = candidate

        rows = []
        for _pair_key, (_rank, source_path, ctx) in sorted(
            source_by_pair.items(),
            key=lambda item: self._remodelling_source_sort_key(item[1][1]),
        ):
            try:
                if self._scene_compartment_is_interactive_source(ctx.get("compartment", "")):
                    preview_inputs = self._get_interactive_preview_inputs(source_path)
                else:
                    _preview, preview_inputs, _source_node, _compartments = self._compute_pair_union_remodelling_preview(
                        ctx,
                        source_path=source_path,
                    )
                for metric_row in self._compute_pair_metric_rows(preview_inputs):
                    row = self._scene_result_row_from_metric(ctx, metric_row)
                    rows.append(row)
            except Exception as exc:
                self._show(f"[scene] could not refresh preview table rows for {Path(source_path).name}: {exc}")
        return self._detect_missing_scene_baseline_pairs(rows)

    def _set_scene_comparison_rows(self, rows=None):
        if not hasattr(self, "sceneComparisonTable"):
            return
        seen_row_keys = set()
        display_rows = []
        for row in list(rows or []):
            row_key = (row[0], row[1]) if len(row) >= 2 else tuple(row)
            if row_key in seen_row_keys:
                continue
            seen_row_keys.add(row_key)
            display_rows.append(row)
        if not display_rows:
            display_rows = [["N/A", "N/A", "N/A", "N/A", "N/A", "N/A"]]
        self.sceneComparisonTable.clearContents()
        self.sceneComparisonTable.setRowCount(len(display_rows))
        for row_idx, row_values in enumerate(display_rows):
            for col_idx, value in enumerate(row_values):
                item = qt.QTableWidgetItem(str(value))
                self.sceneComparisonTable.setItem(row_idx, col_idx, item)
        try:
            self.sceneComparisonTable.resizeColumnsToContents()
            self.sceneComparisonTable.horizontalHeader().setStretchLastSection(True)
        except Exception:
            pass

    def _load_scene_results_table_node(self, rows, plan):
        name = f"TimelapsedHRpQCT Scene Results {plan.run_id}"
        table_node = slicer.mrmlScene.GetFirstNodeByName(name)
        if table_node is None:
            table_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLTableNode", name)
        headers = ["Pair", "Mask", "FV/BV", "RV/BV", "AV/BV", "NV/BV"]
        existing_headers = [
            str(table_node.GetColumnName(col_idx))
            for col_idx in range(int(table_node.GetNumberOfColumns()))
        ]
        if existing_headers == headers:
            for col_idx, _header in enumerate(headers):
                column = table_node.GetTable().GetColumn(col_idx)
                column.SetNumberOfValues(len(rows))
                for row_idx, row in enumerate(rows):
                    column.SetValue(row_idx, str(row[col_idx]))
        else:
            table_node.RemoveAllColumns()
            for col_idx, header in enumerate(headers):
                column = vtk.vtkStringArray()
                column.SetName(header)
                for row in rows:
                    column.InsertNextValue(str(row[col_idx]))
                table_node.AddColumn(column)
        try:
            table_node.Modified()
            table_node.GetTable().Modified()
        except Exception:
            pass
        table_node.SetUseColumnTitleAsColumnHeader(True)
        folder_item_id = self._ensure_load_folder(*self._scene_processed_subject_site(plan))
        self._place_node_in_folder(table_node, folder_item_id)
        return table_node

    def _show_scene_results_table_node(self, table_node):
        try:
            layout_manager = slicer.app.layoutManager()
            layout_with_table = slicer.modules.tables.logic().GetLayoutWithTable(layout_manager.layout)
            layout_manager.setLayout(layout_with_table)
            slicer.app.applicationLogic().GetSelectionNode().SetActiveTableID(table_node.GetID())
            slicer.app.applicationLogic().PropagateTableSelection()
        except Exception as exc:
            self._show(f"[scene] could not show scene results table in Slicer table view: {exc}")

    def _load_scene_results_table(self, plan, *, show=False, prefer_saved=False):
        rows = []
        rows_source = "saved CSV"
        if prefer_saved:
            rows = self._scene_result_rows(plan)
        if not rows:
            rows = self._scene_result_rows_from_loaded_remodelling()
            rows_source = "current scene display"
        if not rows and not prefer_saved:
            rows = self._scene_result_rows(plan)
            rows_source = "saved CSV"
        self._set_scene_comparison_rows(rows)
        if not rows:
            return 0

        if show:
            # Scene result table node is only created when explicitly shown.
            table_node = self._load_scene_results_table_node(rows, plan)
            self._show_scene_results_table_node(table_node)
        self._show(f"[scene] Loaded scene results table with {len(rows)} row(s) from {rows_source}.")
        return len(rows)

    def _refresh_scene_results_table_from_loaded_remodelling(self):
        plan = getattr(self, "_last_scene_results_plan", None)
        if plan is None:
            return 0
        return self._load_scene_results_table(plan, show=False)

    def _select_first_scene_remodelling_output(self):
        if not hasattr(self, "remodellingFullSegCombo"):
            return
        self._refresh_remodelling_full_selector()
        if self.remodellingFullSegCombo.count > 0:
            self.remodellingFullSegCombo.setCurrentIndex(0)
            self._activate_remodelling_display_for_current_selection()

    def _prewarm_selected_scene_preview_cache(self):
        source_path = self._selected_remodelling_source_path()
        if not source_path:
            return False
        ctx = self._parse_remodelling_source_context(source_path)
        if ctx is None:
            return False
        try:
            if self._scene_compartment_is_interactive_source(ctx.get("compartment", "")):
                preview_inputs = self._get_interactive_preview_inputs(source_path)
                _ = self._preview_delta_for_current_settings(preview_inputs)
            else:
                source_node = self._find_loaded_interactive_remodelling_node_for_context(ctx, preferred_compartment="full")
                if source_node is None:
                    self._show(f"[preview] prewarming display-union output from saved scene data: {Path(source_path).name}.")
                self._compute_pair_union_remodelling_preview(ctx, source_path=source_path)
            self._show(f"[preview] prewarmed interactive cache for {Path(source_path).name}.")
            return True
        except Exception as exc:
            self._show(f"[preview] could not prewarm interactive cache: {exc}")
            return False

    def _load_scene_run_outputs(self, plan):
        self._remove_scene_run_nonlinear_transform_nodes()
        output_root = Path(plan.output_root)
        if not output_root.exists():
            self._show(f"[scene] output folder not found for load-back: {output_root}")
            return

        self._clear_loaded_review_nodes()
        self._last_scene_results_plan = plan
        processed_subject_id, processed_site = self._scene_processed_subject_site(plan)
        folder_item_id = self._ensure_load_folder(processed_subject_id, processed_site)
        debug_load_masks = str(os.environ.get("SLICER_TIMELAPSED_DEBUG_LOAD_MASKS", "")).strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }
        loaded_masks = self._load_scene_run_masks(plan) if debug_load_masks else 0
        if not debug_load_masks:
            self._show("[scene] skipped mask load-back; using already loaded scene masks.")
        apply_scene_transforms = str(os.environ.get("SLICER_TIMELAPSED_APPLY_SCENE_TRANSFORMS", "")).strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }
        loaded_transforms = 0
        loaded_transform_nodes = {}
        applied_transforms = 0
        loaded_remodelling = 0
        loaded_result_rows = 0

        transform_paths = sorted(
            path
            for pattern in ("*.tfm",)
            for path in output_root.rglob(pattern)
            if path.is_file()
        )
        for path in transform_paths:
            try:
                ok, node = self._load_transform_node(path)
                if ok and node is not None:
                    self._place_node_in_folder(node, folder_item_id)
                    loaded_transform_nodes[Path(path)] = node
                    loaded_transforms += 1
            except Exception as exc:
                self._show(f"[scene] could not load transform {path.name}: {exc}")
        if apply_scene_transforms:
            applied_transforms = self._apply_scene_baseline_transforms(plan, loaded_transform_nodes)
        else:
            self._show("[scene] skipped transform application to loaded scene nodes; keeping scene inputs native.")

        remodelling_paths = sorted(
            {
                path
                for pattern in ("*_remodelling.nii.gz", "*_remodelling.mha")
                for path in output_root.rglob(pattern)
                if path.is_file()
            }
        )
        for path in remodelling_paths:
            try:
                seg_name = f"{path.stem}_segmentation"
                if self._load_remodelling_as_segmentation(
                    seg_name,
                    path,
                    folder_item_id=folder_item_id,
                    create_full=True,
                    activate_display=False,
                ):
                    loaded_remodelling += 1
            except Exception as exc:
                self._show(f"[scene] could not load remodelling output {path.name}: {exc}")

        loaded_result_rows = self._load_scene_results_table(plan, show=True, prefer_saved=True)
        self.sceneStatusLabel.text = (
            f"Loaded {loaded_masks} mask(s), {loaded_transforms} transform(s), "
            f"applied {applied_transforms} to loaded scene node(s), "
            f"{loaded_remodelling} remodelling output(s), and "
            f"{loaded_result_rows} result row(s) from the scene run."
        )
        self._show(
            f"[scene] loaded {loaded_masks} mask(s), {loaded_transforms} transform(s), "
            f"applied {applied_transforms} to loaded scene node(s), "
            f"{loaded_remodelling} remodelling output(s), "
            f"{loaded_result_rows} result row(s) from {output_root}"
        )
        self._select_first_scene_remodelling_output()
        self._prewarm_selected_scene_preview_cache()
        self._refresh_pair_metrics_for_current_selection()

    def _maybe_apply_raw_stack_offset(self, node, record):
        """If imported raw stacks have zero origin, offset by metadata z_start."""
        slice_range = getattr(record, "slice_range", None)
        if node is None or slice_range is None:
            return
        try:
            z_start = int(getattr(slice_range, "z_start"))
            spacing = tuple(float(x) for x in node.GetSpacing())
            origin = list(float(x) for x in node.GetOrigin())
            target_z = float(z_start) * float(spacing[2])
            if abs(origin[2] - target_z) > 1e-6:
                origin[2] = target_z
                node.SetOrigin(origin)
                self._show(
                    f"[load] applied raw stack z-offset: z_start={z_start}, "
                    f"spacing_z={spacing[2]:.6f}, origin_z={target_z:.6f}"
                )
        except Exception as exc:
            self._show(f"[load] could not apply stack offset: {exc}")

    def _on_load_selected(self):
        root = self._dataset_root()
        if root is None:
            slicer.util.errorDisplay("Select a dataset root first.")
            return

        self._clear_loaded_review_nodes()

        patient_key = self._current_patient_key()
        if patient_key is None:
            slicer.util.errorDisplay("No processed patient available to load.")
            return
        subject_id, site = patient_key

        imported = self._imported_dataset_root()
        if imported is None:
            slicer.util.errorDisplay("Could not resolve derivatives path.")
            return

        data_type = self.loadTypeCombo.currentText
        is_remodelling_load = data_type == "remodelling image"
        load_masks_with_images = data_type in {"raw", "transformed"}

        candidates = []
        image_records = []
        loaded_remodelling_source_path = None
        try:
            from timelapsedhrpqct.dataset.artifacts import (
                iter_fused_session_records,
                iter_imported_stack_records,
            )

            if data_type == "raw":
                for rec in iter_imported_stack_records(imported):
                    if rec.subject_id == subject_id and rec.site == site and rec.image_path.exists():
                        candidates.append(rec.image_path)
                        image_records.append(rec)
            elif data_type == "transformed":
                for rec in iter_fused_session_records(imported):
                    if rec.subject_id == subject_id and rec.site == site and rec.image_path.exists():
                        candidates.append(rec.image_path)
                        image_records.append(rec)
        except Exception as exc:
            self._show(f"[load] artifact-based lookup failed: {exc}")

        if is_remodelling_load:
            selected_path = self._current_remodelling_comparison_path()
            if selected_path is not None and Path(selected_path).exists():
                candidates.append(Path(selected_path))

        candidates = sorted(set(candidates))

        if not candidates:
            slicer.util.warningDisplay(
                f"No files found for '{data_type}' in sub-{subject_id} site-{site}."
            )
            return

        loaded = 0
        first_loaded_node = None

        if image_records:
            for rec in sorted(
                image_records,
                key=lambda r: (
                    str(getattr(r, "session_id", "")),
                    int(getattr(r, "stack_index", 0)),
                    str(getattr(r, "image_path", "")),
                ),
            ):
                p = rec.image_path
                ok, node = self._load_volume_node(p)
                if ok and node is not None:
                    loaded += 1
                    if first_loaded_node is None:
                        first_loaded_node = node
                    session_id = str(getattr(rec, "session_id", ""))
                    stack_index = getattr(rec, "stack_index", None)
                    folder_id = self._ensure_load_folder(subject_id, site, session_id, stack_index)
                    self._place_node_in_folder(node, folder_id)
                    try:
                        origin = tuple(float(x) for x in node.GetOrigin())
                        spacing = tuple(float(x) for x in node.GetSpacing())
                        self._show(
                            f"[load] {Path(p).name} origin={origin} spacing={spacing}"
                        )
                    except Exception:
                        pass
                    if load_masks_with_images:
                        role_to_path = {}
                        for role, mask_path in (getattr(rec, "mask_paths", {}) or {}).items():
                            if mask_path and Path(mask_path).exists():
                                role_to_path[str(role)] = Path(mask_path)
                        if role_to_path:
                            seg_name = (
                                f"sub-{subject_id}_site-{site}_ses-{session_id}_"
                                f"stack-{int(stack_index):02d}_{data_type}_masks"
                                if stack_index is not None
                                else f"sub-{subject_id}_site-{site}_ses-{session_id}_{data_type}_masks"
                            )
                            self._load_masks_as_segmentation(
                                seg_name,
                                role_to_path,
                                session_id=session_id,
                                folder_item_id=folder_id,
                                reference_volume_node=node,
                            )
                        seg_path = getattr(rec, "seg_path", None)
                        if seg_path and Path(seg_path).exists():
                            seg_name = (
                                f"sub-{subject_id}_site-{site}_ses-{session_id}_"
                                f"stack-{int(stack_index):02d}_{data_type}_seg"
                                if stack_index is not None
                                else f"sub-{subject_id}_site-{site}_ses-{session_id}_{data_type}_seg"
                            )
                            self._load_masks_as_segmentation(
                                seg_name,
                                {"seg": Path(seg_path)},
                                session_id=session_id,
                                folder_item_id=folder_id,
                                reference_volume_node=node,
                            )
        else:
            folder_id = self._ensure_load_folder(subject_id, site)
            for p in candidates:
                if is_remodelling_load:
                    seg_name = f"{Path(p).stem}_segmentation"
                    ok = self._load_remodelling_as_segmentation(
                        seg_name,
                        Path(p),
                        folder_item_id=folder_id,
                        create_full=True,
                    )
                    if ok:
                        loaded += 1
                        loaded_remodelling_source_path = str(Path(p).resolve())
                        self._show(f"[load] {Path(p).name} loaded as remodelling segmentation.")
                else:
                    ok, node = self._load_volume_node(p)
                    if ok and node is not None:
                        loaded += 1
                        if first_loaded_node is None:
                            first_loaded_node = node
                        self._place_node_in_folder(node, folder_id)
                        try:
                            origin = tuple(float(x) for x in node.GetOrigin())
                            spacing = tuple(float(x) for x in node.GetSpacing())
                            self._show(
                                f"[load] {Path(p).name} origin={origin} spacing={spacing}"
                            )
                        except Exception:
                            pass

        self._show(
            f"[load] loaded {loaded}/{len(candidates)} files for "
            f"sub-{subject_id} site-{site} ({data_type})"
        )
        if first_loaded_node is not None:
            self._center_slices_on_node(first_loaded_node, fit_to_bounds=True)
        if is_remodelling_load and loaded:
            if loaded_remodelling_source_path:
                self._set_remodelling_selector_by_source_path(loaded_remodelling_source_path)
            if self.remodellingFullSegCombo.count > 0 and self.remodellingFullSegCombo.currentIndex < 0:
                self.remodellingFullSegCombo.setCurrentIndex(0)
            try:
                series_inputs = self._get_subject_series_preview_inputs(subject_id, site)
                self._rebuild_series_summary_pair_selector([entry["session_id"] for entry in series_inputs])
            except Exception as exc:
                self._show(f"[series] could not build adjacent-pair selector: {exc}")
                self._rebuild_series_summary_pair_selector([])
            if loaded_remodelling_source_path:
                ctx = self._parse_remodelling_source_context(loaded_remodelling_source_path)
                saved_rows = self._saved_pair_metric_rows_for_context(ctx)
                if self._metric_rows_have_finite_fractions(saved_rows):
                    self._set_pair_metric_rows(saved_rows)
                else:
                    self._refresh_pair_metrics_for_current_selection()
            else:
                self._refresh_pair_metrics_for_current_selection()
            self._refresh_saved_cohort_summary()

    def _load_masks_as_segmentation(
        self,
        segmentation_name,
        role_to_path,
        session_id=None,
        folder_item_id=None,
        reference_volume_node=None,
    ):
        seg_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode", segmentation_name)
        seg_node.CreateDefaultDisplayNodes()
        if reference_volume_node is not None:
            seg_node.SetReferenceImageGeometryParameterFromVolumeNode(reference_volume_node)
        seg_logic = slicer.modules.segmentations.logic()
        loaded_any = False
        base_color = self._session_base_color(session_id or "")
        role_tint = {
            "full": 1.00,
            "trab": 0.85,
            "cort": 0.70,
            "regmask": 0.55,
        }

        for role in sorted(role_to_path.keys()):
            path = role_to_path[role]
            ok, label_node = self._load_labelmap_node(path)
            if not ok or label_node is None:
                self._show(f"[load] failed to load mask role '{role}' from {path}")
                continue

            existing_ids = vtk.vtkStringArray()
            seg_node.GetSegmentation().GetSegmentIDs(existing_ids)
            before = {existing_ids.GetValue(i) for i in range(existing_ids.GetNumberOfValues())}

            seg_logic.ImportLabelmapToSegmentationNode(label_node, seg_node)
            slicer.mrmlScene.RemoveNode(label_node)

            updated_ids = vtk.vtkStringArray()
            seg_node.GetSegmentation().GetSegmentIDs(updated_ids)
            after = {updated_ids.GetValue(i) for i in range(updated_ids.GetNumberOfValues())}
            new_ids = list(after - before)
            for seg_id in new_ids:
                segment = seg_node.GetSegmentation().GetSegment(seg_id)
                if segment:
                    segment.SetName(str(role))
                    tint = role_tint.get(str(role).lower(), 0.9)
                    segment.SetColor(
                        float(min(max(base_color[0] * tint, 0.0), 1.0)),
                        float(min(max(base_color[1] * tint, 0.0), 1.0)),
                        float(min(max(base_color[2] * tint, 0.0), 1.0)),
                    )
            loaded_any = loaded_any or bool(new_ids)

        if not loaded_any:
            slicer.mrmlScene.RemoveNode(seg_node)
            return False
        self._configure_segmentation_display(seg_node)
        if folder_item_id is not None:
            self._place_node_in_folder(seg_node, folder_item_id)
        return True


class TimelapsedHRpQCTTest(ScriptedLoadableModuleTest):
    """Minimal smoke tests for release readiness."""

    def setUp(self):
        slicer.mrmlScene.Clear()

    def runTest(self):
        self.setUp()
        self.test_logic_config_resolution()
        self.test_override_config_write()

    def test_logic_config_resolution(self):
        logic = TimelapsedHRpQCTLogic()
        available, _detail = logic.pipeline_status()
        if not available:
            self.skipTest("timelapsed-hrpqct pipeline is not installed")
        config_path = logic.default_config_path()
        self.assertTrue(Path(config_path).exists())
        self.assertTrue(str(config_path).endswith(".yml"))

    def test_override_config_write(self):
        try:
            import yaml  # noqa: F401
        except Exception:
            self.skipTest("PyYAML is not available")

        logic = TimelapsedHRpQCTLogic()
        results_root = Path(tempfile.mkdtemp(prefix="timelapsed_slicer_test_"))
        path = logic.create_override_config(
            {
                "analysis": {"thresholds": [225.0], "cluster_sizes": [12]},
                "fusion": {"enable_filling": False},
            },
            results_root=results_root,
        )
        p = Path(path)
        self.assertTrue(p.exists())
        self.assertTrue(p.parent == results_root / "slicer_run_configs")
        text = p.read_text(encoding="utf-8")
        self.assertIn("analysis:", text)
        self.assertIn("fusion:", text)
        logic.cleanup_temp_files(remove_fallback=False)
        self.assertTrue(p.exists())
        shutil.rmtree(results_root, ignore_errors=True)
